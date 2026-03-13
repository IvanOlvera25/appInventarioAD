# app.py - Versión corregida para manejar migraciones
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, current_app
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import func  # <— lo usas en agregaciones
from datetime import datetime, timedelta, timezone
import pandas as pd
import os
import tempfile
import pymysql
import pymysql.cursors
from contextlib import contextmanager
from models import db, User, Project, Material, FabricRoll, Request, RequestItem, ProjectSummary, StockMovement, PurchaseRequest, VerificationCode, Department, Category, Unit
from functools import wraps
import secrets  # <- si estás generando códigos de verificación
from sqlalchemy import text
from apscheduler.schedulers.background import BackgroundScheduler
import atexit

app = Flask(__name__)
app.config['SECRET_KEY'] = 'tu-clave-secreta-aqui'
# ===== BASE DE DATOS REMOTA PRINCIPAL =====
app.config['SQLALCHEMY_DATABASE_URI'] = (
    'mysql+pymysql://IvanUriel:iuOp20!!25'
    '@ad17solutions.dscloud.me:3307/AD17_Almacen?charset=utf8mb4'
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
    'connect_args': {'connect_timeout': 10},
}

# Configuración para archivos
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

# Departamentos estáticos (cámbialos si necesitas)
# Deja esto UNA sola vez en el archivo (quita las otras variantes)
DEPARTMENTS_STATIC = [
    "Producción", "Mantenimiento", "Ingeniería",
    "Compras", "Logística", "Administración",
]

@app.context_processor
def inject_globals():
    try:
        active_depts = Department.query.filter_by(is_active=True).order_by(Department.name).all()
        departments = active_depts if active_depts else DEPARTMENTS_STATIC
    except Exception:
        departments = DEPARTMENTS_STATIC

    # Cargar categorías y unidades también
    try:
        categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()
    except:
        categories = []
        
    try:
        units = Unit.query.filter_by(is_active=True).order_by(Unit.name).all()
    except:
        units = []

    return {
        'moment': datetime,
        'now': datetime.utcnow(),
        'departments': departments,
        'categories': categories,
        'units': units
    }

# Asegurarse de que el directorio existe
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS




@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.context_processor
def inject_datetime():
    return {
        'moment': datetime,
        'now': datetime.utcnow()
    }

# Función para verificar si la base de datos necesita migración
# Reemplaza estas funciones en tu app.py

def check_database_migration_needed():
    """Verifica si la base de datos necesita migración"""
    try:
        # Primero verificar si las tablas existen
        db.create_all()

        # ===== MIGRACIÓN AUTOMÁTICA DE COLUMNAS REQUEST =====
        from sqlalchemy import text, inspect
        inspector = inspect(db.engine)

        if 'request' in inspector.get_table_names():
            columns = [col['name'] for col in inspector.get_columns('request')]

            with db.engine.connect() as conn:
                if 'acquisition_deadline' not in columns:
                    conn.execute(text('ALTER TABLE request ADD COLUMN acquisition_deadline DATE'))
                    print("  ✅ Columna 'acquisition_deadline' agregada")
                if 'production_start_date' not in columns:
                    conn.execute(text('ALTER TABLE request ADD COLUMN production_start_date DATE'))
                    print("  ✅ Columna 'production_start_date' agregada")
                if 'has_returns' not in columns:
                    conn.execute(text('ALTER TABLE request ADD COLUMN has_returns BOOLEAN DEFAULT 0'))
                    print("  ✅ Columna 'has_returns' agregada")
                conn.commit()
        # ===== FIN MIGRACIÓN REQUEST =====

        # ===== MIGRACIÓN AUTOMÁTICA DE COLUMNAS PROJECT =====
        if 'project' in inspector.get_table_names():
            columns = [col['name'] for col in inspector.get_columns('project')]

            with db.engine.connect() as conn:
                if 'client' not in columns:
                    conn.execute(text('ALTER TABLE project ADD COLUMN client VARCHAR(150)'))
                    print("  ✅ Columna 'client' agregada a project")
                conn.commit()
        # ===== FIN MIGRACIÓN PROJECT =====

        # ===== MIGRACIÓN AUTOMÁTICA DE COLUMNAS MATERIAL =====
        if 'material' in inspector.get_table_names():
            columns = [col['name'] for col in inspector.get_columns('material')]

            with db.engine.connect() as conn:
                if 'fabric_width' not in columns:
                    conn.execute(text('ALTER TABLE material ADD COLUMN fabric_width FLOAT'))
                    print("  ✅ Columna 'fabric_width' agregada a material")
                if 'is_recycled' not in columns:
                    conn.execute(text('ALTER TABLE material ADD COLUMN is_recycled BOOLEAN DEFAULT 0'))
                    print("  ✅ Columna 'is_recycled' agregada a material")
                if 'is_pre_recycled' not in columns:
                    conn.execute(text('ALTER TABLE material ADD COLUMN is_pre_recycled BOOLEAN DEFAULT 0'))
                    print("  ✅ Columna 'is_pre_recycled' agregada a material")
                if 'recycled_from_id' not in columns:
                    conn.execute(text('ALTER TABLE material ADD COLUMN recycled_from_id INTEGER'))
                    print("  ✅ Columna 'recycled_from_id' agregada a material")
                if 'category_id' not in columns:
                    conn.execute(text('ALTER TABLE material ADD COLUMN category_id INTEGER'))
                    print("  ✅ Columna 'category_id' agregada a material")
                if 'unit_id' not in columns:
                    conn.execute(text('ALTER TABLE material ADD COLUMN unit_id INTEGER'))
                    print("  ✅ Columna 'unit_id' agregada a material")
                conn.commit()
        # ===== FIN MIGRACIÓN MATERIAL =====

        # Verificar columnas requeridas en stock_movement (MySQL-compatible)
        try:
            result = db.session.execute(db.text(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = 'AD17_Almacen' AND TABLE_NAME = 'stock_movement'"
            )).fetchall()
            column_names = [row[0] for row in result]
            required_columns = ['idm', 'rollos', 'fp_code', 'fecha', 'hora', 'personal', 'area', 'updated_at']
            missing_columns = [col for col in required_columns if col not in column_names]
            if missing_columns:
                print(f"⚠️  Columnas faltantes en stock_movement: {', '.join(missing_columns)}")
                return True
        except Exception:
            pass  # Si la tabla no existe aún, db.create_all() la creará

        return False  # No necesita migración

    except Exception as e:
        print(f"⚠️  Error al verificar migración: {e}")
        return False  # En MySQL continuar de todas formas

def _mysql_column_exists(conn, table, column, schema='AD17_Almacen'):
    """Verifica si una columna existe en MySQL usando INFORMATION_SCHEMA"""
    result = conn.execute(text(
        "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
        "WHERE TABLE_SCHEMA = :schema AND TABLE_NAME = :table AND COLUMN_NAME = :column"
    ), {"schema": schema, "table": table, "column": column}).scalar()
    return result > 0

def safe_init_db():
    """Inicializa la base de datos de forma segura con MySQL"""
    try:
        # Crear todas las tablas que no existan
        db.create_all()

        from sqlalchemy import text, inspect
        inspector = inspect(db.engine)
        tables = inspector.get_table_names()

        # ===== MIGRACIÓN AUTOMÁTICA DE COLUMNAS (MySQL-safe) =====
        with db.engine.connect() as conn:

            if 'request' in tables:
                if not _mysql_column_exists(conn, 'request', 'acquisition_deadline'):
                    conn.execute(text('ALTER TABLE `request` ADD COLUMN acquisition_deadline DATE'))
                    print("  ✅ Columna 'acquisition_deadline' agregada")
                if not _mysql_column_exists(conn, 'request', 'production_start_date'):
                    conn.execute(text('ALTER TABLE `request` ADD COLUMN production_start_date DATE'))
                    print("  ✅ Columna 'production_start_date' agregada")
                if not _mysql_column_exists(conn, 'request', 'assembly_start_date'):
                    conn.execute(text('ALTER TABLE `request` ADD COLUMN assembly_start_date DATE'))
                    print("  ✅ Columna 'assembly_start_date' agregada")
                if not _mysql_column_exists(conn, 'request', 'assembly_end_date'):
                    conn.execute(text('ALTER TABLE `request` ADD COLUMN assembly_end_date DATE'))
                    print("  ✅ Columna 'assembly_end_date' agregada")
                if not _mysql_column_exists(conn, 'request', 'has_returns'):
                    conn.execute(text('ALTER TABLE `request` ADD COLUMN has_returns TINYINT(1) DEFAULT 0'))
                    print("  ✅ Columna 'has_returns' agregada")
                if not _mysql_column_exists(conn, 'request', 'cancellation_requested'):
                    conn.execute(text('ALTER TABLE `request` ADD COLUMN cancellation_requested TINYINT(1) DEFAULT 0'))
                    print("  ✅ Columna 'cancellation_requested' agregada")
                if not _mysql_column_exists(conn, 'request', 'cancellation_requested_by'):
                    conn.execute(text('ALTER TABLE `request` ADD COLUMN cancellation_requested_by INT'))
                    print("  ✅ Columna 'cancellation_requested_by' agregada")
                if not _mysql_column_exists(conn, 'request', 'cancellation_requested_at'):
                    conn.execute(text('ALTER TABLE `request` ADD COLUMN cancellation_requested_at DATETIME'))
                    print("  ✅ Columna 'cancellation_requested_at' agregada")

            if 'project' in tables:
                if not _mysql_column_exists(conn, 'project', 'client'):
                    conn.execute(text('ALTER TABLE `project` ADD COLUMN client VARCHAR(150)'))
                    print("  ✅ Columna 'client' agregada a project")
                if not _mysql_column_exists(conn, 'project', 'analysis_date'):
                    conn.execute(text('ALTER TABLE `project` ADD COLUMN analysis_date DATE'))
                    print("  ✅ Columna 'analysis_date' agregada a project")

            if 'material' in tables:
                if not _mysql_column_exists(conn, 'material', 'fabric_width'):
                    conn.execute(text('ALTER TABLE `material` ADD COLUMN fabric_width FLOAT'))
                    print("  ✅ Columna 'fabric_width' agregada a material")
                if not _mysql_column_exists(conn, 'material', 'is_recycled'):
                    conn.execute(text('ALTER TABLE `material` ADD COLUMN is_recycled TINYINT(1) DEFAULT 0'))
                    print("  ✅ Columna 'is_recycled' agregada a material")
                if not _mysql_column_exists(conn, 'material', 'is_pre_recycled'):
                    conn.execute(text('ALTER TABLE `material` ADD COLUMN is_pre_recycled TINYINT(1) DEFAULT 0'))
                    print("  ✅ Columna 'is_pre_recycled' agregada a material")
                if not _mysql_column_exists(conn, 'material', 'recycled_from_id'):
                    conn.execute(text('ALTER TABLE `material` ADD COLUMN recycled_from_id INT'))
                    print("  ✅ Columna 'recycled_from_id' agregada a material")
                if not _mysql_column_exists(conn, 'material', 'category_id'):
                    conn.execute(text('ALTER TABLE `material` ADD COLUMN category_id INT'))
                    print("  ✅ Columna 'category_id' agregada a material")
                if not _mysql_column_exists(conn, 'material', 'unit_id'):
                    conn.execute(text('ALTER TABLE `material` ADD COLUMN unit_id INT'))
                    print("  ✅ Columna 'unit_id' agregada a material")

            if 'stock_movement' in tables:
                if not _mysql_column_exists(conn, 'stock_movement', 'idm'):
                    conn.execute(text('ALTER TABLE `stock_movement` ADD COLUMN idm VARCHAR(50)'))
                    print("  ✅ Columna 'idm' agregada a stock_movement")
                if not _mysql_column_exists(conn, 'stock_movement', 'rollos'):
                    conn.execute(text('ALTER TABLE `stock_movement` ADD COLUMN rollos INT DEFAULT 0'))
                    print("  ✅ Columna 'rollos' agregada a stock_movement")
                if not _mysql_column_exists(conn, 'stock_movement', 'personal'):
                    conn.execute(text('ALTER TABLE `stock_movement` ADD COLUMN personal VARCHAR(100)'))
                    print("  ✅ Columna 'personal' agregada a stock_movement")
                if not _mysql_column_exists(conn, 'stock_movement', 'area'):
                    conn.execute(text('ALTER TABLE `stock_movement` ADD COLUMN area VARCHAR(100)'))
                    print("  ✅ Columna 'area' agregada a stock_movement")
                if not _mysql_column_exists(conn, 'stock_movement', 'fp_code'):
                    conn.execute(text('ALTER TABLE `stock_movement` ADD COLUMN fp_code VARCHAR(100)'))
                    print("  ✅ Columna 'fp_code' agregada a stock_movement")
                if not _mysql_column_exists(conn, 'stock_movement', 'fecha'):
                    conn.execute(text('ALTER TABLE `stock_movement` ADD COLUMN fecha DATE'))
                    print("  ✅ Columna 'fecha' agregada a stock_movement")
                if not _mysql_column_exists(conn, 'stock_movement', 'hora'):
                    conn.execute(text('ALTER TABLE `stock_movement` ADD COLUMN hora TIME'))
                    print("  ✅ Columna 'hora' agregada a stock_movement")
                if not _mysql_column_exists(conn, 'stock_movement', 'updated_at'):
                    conn.execute(text('ALTER TABLE `stock_movement` ADD COLUMN updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP'))
                    print("  ✅ Columna 'updated_at' agregada a stock_movement")

            if 'request_item' in tables:
                if not _mysql_column_exists(conn, 'request_item', 'item_status'):
                    conn.execute(text("ALTER TABLE `request_item` ADD COLUMN item_status VARCHAR(50) DEFAULT 'pendiente'"))
                    print("  ✅ Columna 'item_status' agregada a request_item")
                if not _mysql_column_exists(conn, 'request_item', 'quantity_to_purchase'):
                    conn.execute(text('ALTER TABLE `request_item` ADD COLUMN quantity_to_purchase FLOAT DEFAULT 0'))
                    print("  ✅ Columna 'quantity_to_purchase' agregada a request_item")
                if not _mysql_column_exists(conn, 'request_item', 'quantity_supplied'):
                    conn.execute(text('ALTER TABLE `request_item` ADD COLUMN quantity_supplied FLOAT DEFAULT 0'))
                    print("  ✅ Columna 'quantity_supplied' agregada a request_item")
                if not _mysql_column_exists(conn, 'request_item', 'item_notes'):
                    conn.execute(text('ALTER TABLE `request_item` ADD COLUMN item_notes TEXT'))
                    print("  ✅ Columna 'item_notes' agregada a request_item")
                if not _mysql_column_exists(conn, 'request_item', 'actual_return_date'):
                    conn.execute(text('ALTER TABLE `request_item` ADD COLUMN actual_return_date DATE'))
                    print("  ✅ Columna 'actual_return_date' agregada a request_item")

            if 'user' in tables:
                if not _mysql_column_exists(conn, 'user', 'is_verified'):
                    conn.execute(text('ALTER TABLE `user` ADD COLUMN is_verified TINYINT(1) DEFAULT 0'))
                    print("  ✅ Columna 'is_verified' agregada a user")
                if not _mysql_column_exists(conn, 'user', 'verified_at'):
                    conn.execute(text('ALTER TABLE `user` ADD COLUMN verified_at DATETIME'))
                    print("  ✅ Columna 'verified_at' agregada a user")
                if not _mysql_column_exists(conn, 'user', 'is_leader'):
                    conn.execute(text('ALTER TABLE `user` ADD COLUMN is_leader TINYINT(1) DEFAULT 0'))
                    print("  ✅ Columna 'is_leader' agregada a user")

            conn.commit()

        # ===== FIN MIGRACIÓN AUTOMÁTICA =====

        print("✅ Base de datos remota verificada correctamente")
        return True

    except Exception as e:
        print(f"❌ Error al inicializar base de datos: {e}")
        import traceback
        traceback.print_exc()
        return False



class RemoteDatabase:
    """Clase para manejar la conexión a la base de datos MySQL remota"""
    def __init__(self):
        self.connection_params = {
            "host": "ad17solutions.dscloud.me",
            "port": 3307,
            "user": "IvanUriel",
            "password": "iuOp20!!25",
            "charset": 'utf8mb4',
            "cursorclass": pymysql.cursors.DictCursor
        }

    @contextmanager
    def get_connection(self, database="AD17_Materiales"):
        """Context manager para conexiones a la base de datos"""
        connection = None
        try:
            params = self.connection_params.copy()
            params['database'] = database
            connection = pymysql.connect(**params)
            yield connection
        except pymysql.Error as e:
            app.logger.error(f"Error de conexión MySQL: {e}")
            if connection:
                connection.rollback()
            raise
        finally:
            if connection:
                connection.close()

    def get_empleados_activos(self):
            """Obtener lista de empleados activos ordenados alfabéticamente"""
            query = """
            SELECT id, nomPropio as nombre
            FROM AD17_RH.empleados_activos
            ORDER BY nomPropio ASC;
            """
            try:
                with self.get_connection(database="AD17_RH") as conn:
                    with conn.cursor() as cursor:
                        cursor.execute(query)
                        rows = cursor.fetchall()
                        # Normalizar datos
                        for r in rows:
                            r['id'] = int(r['id']) if r.get('id') is not None else None
                            r['nombre'] = (r.get('nombre') or '').strip()
                        return rows
            except Exception as e:
                app.logger.error(f"Error obteniendo empleados activos: {e}")
                return []

    def get_empleado_by_id(self, employee_id):
        """Obtener información de un empleado por su ID"""
        query = """
        SELECT id, nomPropio as nombre
        FROM AD17_RH.empleados_activos
        WHERE id = %s;
        """
        try:
            with self.get_connection(database="AD17_RH") as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query, (employee_id,))
                    row = cursor.fetchone()
                    if row:
                        row['nombre'] = (row.get('nombre') or '').strip()
                    return row
        except Exception as e:
            app.logger.error(f"Error obteniendo empleado {employee_id}: {e}")
            return None

    def get_materiales_habilitados(self):
        """Obtener materiales habilitados con categorías y unidades correctas"""
        query = """
        SELECT i.id,
               d.nombre AS material,
               u.unidad AS uni_simbolo,
               u.singular AS uni_singular,
               u.plural AS uni_plural,
               u.tipovar AS tipovar,
               c.categoria AS categoria,
               d.descripcion AS descripcion,
               i.timestamp AS alta,
               d.timestamp AS modificacion
        FROM AD17_Materiales.ID AS i
        LEFT JOIN (
            SELECT * FROM AD17_Materiales.Datos
            WHERE regID IN (
                SELECT MAX(regID)
                FROM AD17_Materiales.Datos
                GROUP BY matID
            )
        ) AS d ON d.matID = i.id
        LEFT JOIN AD17_Materiales.Categoria AS c ON c.regID = d.categoria
        LEFT JOIN AD17_General.Unidades AS u ON u.regID = d.unidad
        WHERE i.habilitado = true
        ORDER BY material ASC;
        """
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query)
                    rows = cursor.fetchall()

                    # Normalización y logging mejorado
                    for r in rows:
                        r['id'] = str(r['id']).strip() if r.get('id') is not None else None
                        r['material'] = (r.get('material') or '').strip()
                        r['categoria'] = (r.get('categoria') or '').strip()
                        r['descripcion'] = (r.get('descripcion') or '').strip()
                        r['uni_simbolo'] = (r.get('uni_simbolo') or '').strip()

                        # Debug logging para los primeros 3 registros
                        if rows.index(r) < 3:
                            app.logger.debug(f"Material {r['id']}: cat={r['categoria']}, unit={r['uni_simbolo']}")

                    return rows
        except Exception as e:
            app.logger.error(f"Error obteniendo materiales: {e}")
            import traceback
            app.logger.error(traceback.format_exc())
            return []
    def get_unidades_habilitadas(self):
        query = """
        SELECT regID as id, unidad, singular, plural, tipovar
        FROM AD17_General.Unidades
        WHERE habilitado LIKE true;
        """
        try:
            with self.get_connection(database="AD17_General") as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query)
                    return cursor.fetchall()
        except Exception as e:
            app.logger.error(f"Error obteniendo unidades: {e}")
            return []

    def get_categorias_habilitadas(self):
        query = """
        SELECT regID AS id, categoria
        FROM AD17_Materiales.Categoria
        WHERE habilitado LIKE true;
        """
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query)
                    rows = cursor.fetchall()
                    for r in rows:
                        r['id'] = str(r['id']).strip() if r.get('id') is not None else None
                        r['categoria'] = (r.get('categoria') or '').strip()
                    return rows
        except Exception as e:
            app.logger.error(f"Error obteniendo categorías: {e}")
            return


    # Agregar estos métodos dentro de la clase RemoteDatabase

    def get_remote_categories_for_select(self):
        """Obtener categorías remotas para formularios con logging"""
        try:
            categories = self.get_categorias_habilitadas()
            result = [{'id': cat['id'], 'name': cat['categoria']} for cat in categories]
            app.logger.debug(f"📂 Categorías remotas obtenidas: {len(result)}")
            return result
        except Exception as e:
            app.logger.error(f"Error obteniendo categorías remotas: {e}")
            return []

    def get_remote_units_for_select(self):
        """Obtener unidades remotas para formularios con logging"""
        try:
            units = self.get_unidades_habilitadas()
            result = [{
                'id': unit['id'],
                'symbol': unit['unidad'],
                'singular': unit['singular'],
                'plural': unit['plural']
            } for unit in units]
            app.logger.debug(f"📏 Unidades remotas obtenidas: {len(result)}")
            return result
        except Exception as e:
            app.logger.error(f"Error obteniendo unidades remotas: {e}")
            return []

    def get_registro_materiales_minmax(self):
        """Obtener valores de mínimo y máximo de la tabla registro_materiales"""
        query = """
        SELECT id, minimo, maximo
        FROM AD17_Materiales.registro_materiales;
        """
        try:
            with self.get_connection(database="AD17_Materiales") as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query)
                    rows = cursor.fetchall()
                    # Normalizar datos
                    for r in rows:
                        r['id'] = str(r['id']).strip() if r.get('id') is not None else None
                        r['minimo'] = float(r['minimo']) if r.get('minimo') is not None else 0
                        r['maximo'] = float(r['maximo']) if r.get('maximo') is not None else 0
                    app.logger.info(f"📊 Obtenidos {len(rows)} registros de min/max desde registro_materiales")
                    return rows
        except Exception as e:
            app.logger.error(f"Error obteniendo registro_materiales: {e}")
            import traceback
            app.logger.error(traceback.format_exc())
            return []

# Instanciar la clase después de definirla
remote_db = RemoteDatabase()


def _remote_categories_set():
    cats = remote_db.get_categorias_habilitadas()
    return { (c['categoria'] or '').strip() for c in cats if c.get('categoria') }

def _remote_materials_map_by_code():
    # Mapa code -> dict remoto
    mats = remote_db.get_materiales_habilitados()
    return { str(m['id']).strip(): m for m in mats if m.get('id') is not None }

def _hydrate_local_from_remote(local_material, remote_row):
    """Sobrescribe campos 'autoridad' desde remoto en el material local"""
    local_material.name = remote_row.get('material') or local_material.name
    local_material.description = remote_row.get('descripcion') or ''

    # ✅ Manejo robusto de unidad
    unit = (remote_row.get('uni_simbolo') or '').strip()
    if unit:
        local_material.unit = unit
    elif not local_material.unit:
        local_material.unit = 'unidad'

    # ✅ Manejo robusto de categoría
    cat = (remote_row.get('categoria') or '').strip()
    if cat:
        local_material.category = cat

    # Actualizar timestamp
    if hasattr(local_material, 'updated_at'):
        local_material.updated_at = datetime.utcnow()
def sync_materials_from_remote():
    """Sincronizar materiales desde la base de datos remota con logging mejorado"""
    try:
        app.logger.info("🔄 Iniciando sincronización de materiales...")

        # Obtener materiales remotos
        remote_materials = remote_db.get_materiales_habilitados()
        app.logger.info(f"📦 Obtenidos {len(remote_materials)} materiales de la base remota")

        # Obtener categorías válidas
        allowed_categories = _remote_categories_set()
        app.logger.info(f"📂 Categorías válidas: {allowed_categories}")

        synced_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        for idx, r in enumerate(remote_materials):
            try:
                code = str(r['id']).strip()
                cat  = (r.get('categoria') or '').strip()
                unit = (r.get('uni_simbolo') or '').strip()
                name = (r.get('material') or '').strip()

                # Log de los primeros 5 para debug
                if idx < 5:
                    app.logger.debug(f"[{idx+1}] Procesando: {code} | Cat: '{cat}' | Unit: '{unit}' | Nombre: '{name}'")

                # Validar categoría
                if not cat or cat not in allowed_categories:
                    if idx < 5:
                        app.logger.warning(f"⚠️  Material {code} sin categoría válida (tiene: '{cat}')")
                    skipped_count += 1
                    continue

                # Validar unidad (usar default si no viene)
                if not unit:
                    unit = 'unidad'
                    if idx < 5:
                        app.logger.warning(f"⚠️  Material {code} sin unidad, usando default")

                # Buscar si existe localmente
                existing = Material.query.filter_by(code=code).first()

                if existing:
                    # Actualizar material existente
                    _hydrate_local_from_remote(existing, r)
                    updated_count += 1

                    if idx < 5:
                        app.logger.debug(f"✅ Actualizado: {code}")
                else:
                    # Crear nuevo material
                    new_material = Material(
                        code=code,
                        name=name or f"Material {code}",
                        description=r.get('descripcion') or '',
                        unit=unit,
                        category=cat,
                        current_stock=0,
                        min_stock=0,
                        max_stock=100,
                        unit_cost=0,
                        is_consumible=False,
                        can_recycle=False,
                        can_reuse=True
                    )
                    db.session.add(new_material)
                    synced_count += 1

                    if idx < 5:
                        app.logger.debug(f"🆕 Creado: {code}")

            except Exception as e:
                error_msg = f"Error procesando material {r.get('id', 'unknown')}: {str(e)}"
                app.logger.error(error_msg)
                errors.append(error_msg)
                continue

        # Commit de todos los cambios
        db.session.commit()

        result_msg = f"✅ Sincronización completada: {synced_count} nuevos, {updated_count} actualizados, {skipped_count} omitidos"
        app.logger.info(result_msg)

        if errors:
            app.logger.warning(f"⚠️  Se encontraron {len(errors)} errores durante la sincronización")

        return {
            'success': True,
            'synced': synced_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'total': len(remote_materials),
            'errors': errors[:10]  # Primeros 10 errores
        }

    except Exception as e:
        db.session.rollback()
        error_msg = f'Error en sincronización: {str(e)}'
        app.logger.error(error_msg)
        import traceback
        app.logger.error(traceback.format_exc())
        return {'success': False, 'error': error_msg}

def get_remote_categories_for_select():
    """Wrapper global para mantener compatibilidad"""
    return remote_db.get_remote_categories_for_select()

def get_remote_units_for_select():
    """Wrapper global para mantener compatibilidad"""
    return remote_db.get_remote_units_for_select()

def sync_minmax_from_remote():
    """Sincronizar valores mínimo y máximo desde registro_materiales remoto"""
    try:
        app.logger.info("🔄 Iniciando sincronización de mínimos y máximos...")

        # Obtener datos de min/max desde la tabla remota
        remote_minmax = remote_db.get_registro_materiales_minmax()

        if not remote_minmax:
            app.logger.warning("⚠️ No se obtuvieron registros de min/max")
            return {
                'success': False,
                'error': 'No se pudieron obtener los registros de min/max'
            }

        updated_count = 0
        not_found_count = 0
        errors = []

        for record in remote_minmax:
            try:
                mat_id = record.get('id')
                minimo = record.get('minimo', 0)
                maximo = record.get('maximo', 0)

                if not mat_id:
                    continue

                # Buscar el material local por código
                local_material = Material.query.filter_by(code=mat_id).first()

                if local_material:
                    # Actualizar min/max
                    local_material.min_stock = minimo
                    local_material.max_stock = maximo
                    updated_count += 1

                    if updated_count <= 5:
                        app.logger.debug(f"✅ Actualizado {mat_id}: min={minimo}, max={maximo}")
                else:
                    not_found_count += 1
                    if not_found_count <= 5:
                        app.logger.debug(f"⚠️ Material {mat_id} no encontrado localmente")

            except Exception as e:
                error_msg = f"Error procesando material {record.get('id', 'unknown')}: {str(e)}"
                app.logger.error(error_msg)
                errors.append(error_msg)
                continue

        # Commit de todos los cambios
        db.session.commit()

        result_msg = f"✅ Sincronización min/max completada: {updated_count} actualizados, {not_found_count} no encontrados"
        app.logger.info(result_msg)

        return {
            'success': True,
            'updated': updated_count,
            'not_found': not_found_count,
            'total': len(remote_minmax),
            'errors': errors[:10]
        }

    except Exception as e:
        db.session.rollback()
        error_msg = f'Error en sincronización min/max: {str(e)}'
        app.logger.error(error_msg)
        import traceback
        app.logger.error(traceback.format_exc())
        return {'success': False, 'error': error_msg}


def write_minmax_to_remote():
    """Escribir valores min/max locales a las tablas Minimos y Maximos de la BD remota"""
    try:
        app.logger.info("📤 Iniciando escritura de min/max a BD remota...")

        # Obtener materiales locales con min/max definidos
        local_materials = Material.query.filter(
            (Material.min_stock > 0) | (Material.max_stock > 0)
        ).all()

        if not local_materials:
            app.logger.info("ℹ️ No hay materiales con min/max definidos para sincronizar")
            return {'success': True, 'updated': 0, 'message': 'No hay datos para sincronizar'}

        updated_min = 0
        updated_max = 0
        errors = []

        connection_params = {
            "host": "ad17solutions.dscloud.me",
            "port": 3307,
            "user": "IvanUriel",
            "password": "iuOp20!!25",
            "database": "AD17_Materiales",
            "charset": 'utf8mb4',
            "cursorclass": pymysql.cursors.DictCursor
        }

        conn = pymysql.connect(**connection_params)
        try:
            with conn.cursor() as cursor:
                now = datetime.now(timezone.utc)

                for mat in local_materials:
                    try:
                        mat_id = int(mat.code) if mat.code.isdigit() else mat.code

                        # Escribir en tabla Minimos si hay valor
                        if mat.min_stock > 0:
                            # Verificar si ya existe un registro para este material
                            cursor.execute("SELECT regID FROM Minimos WHERE matID = %s ORDER BY regID DESC LIMIT 1", (mat_id,))
                            existing_min = cursor.fetchone()

                            if existing_min:
                                # Actualizar el registro existente
                                cursor.execute(
                                    "UPDATE Minimos SET minimo = %s, timestamp = %s WHERE regID = %s",
                                    (mat.min_stock, now, existing_min['regID'])
                                )
                            else:
                                # Insertar nuevo registro
                                cursor.execute(
                                    "INSERT INTO Minimos (matID, minimo, timestamp) VALUES (%s, %s, %s)",
                                    (mat_id, mat.min_stock, now)
                                )
                            updated_min += 1

                        # Escribir en tabla Maximos si hay valor
                        if mat.max_stock > 0:
                            # Verificar si ya existe un registro para este material
                            cursor.execute("SELECT regID FROM Maximos WHERE matID = %s ORDER BY regID DESC LIMIT 1", (mat_id,))
                            existing_max = cursor.fetchone()

                            if existing_max:
                                # Actualizar el registro existente
                                cursor.execute(
                                    "UPDATE Maximos SET maximo = %s, timestamp = %s WHERE regID = %s",
                                    (mat.max_stock, now, existing_max['regID'])
                                )
                            else:
                                # Insertar nuevo registro
                                cursor.execute(
                                    "INSERT INTO Maximos (matID, maximo, timestamp) VALUES (%s, %s, %s)",
                                    (mat_id, mat.max_stock, now)
                                )
                            updated_max += 1

                        if (updated_min + updated_max) <= 10:
                            app.logger.debug(f"✅ Escrito {mat.code}: min={mat.min_stock}, max={mat.max_stock}")

                    except Exception as e:
                        error_msg = f"Error escribiendo material {mat.code}: {str(e)}"
                        app.logger.error(error_msg)
                        errors.append(error_msg)
                        continue

                conn.commit()

        finally:
            conn.close()

        result_msg = f"✅ Escritura min/max: {updated_min} mínimos, {updated_max} máximos actualizados"
        app.logger.info(result_msg)

        return {
            'success': True,
            'updated': updated_min + updated_max,
            'updated_min': updated_min,
            'updated_max': updated_max,
            'total': len(local_materials),
            'errors': errors[:10]
        }

    except Exception as e:
        error_msg = f'Error escribiendo a BD remota: {str(e)}'
        app.logger.error(error_msg)
        import traceback
        app.logger.error(traceback.format_exc())
        return {'success': False, 'error': error_msg}


def write_materials_to_remote():
    """
    Escribir materiales locales a la BD remota.
    NOTA: La tabla Datos usa IDs numéricos para unidad/categoría.
    Solo actualizamos nombre y descripción de materiales existentes.
    Para crear nuevos materiales, se requiere acceso a las tablas de catálogo.
    """
    try:
        app.logger.info("📤 Iniciando escritura de materiales a BD remota...")

        # Obtener todos los materiales locales
        local_materials = Material.query.all()

        if not local_materials:
            app.logger.info("ℹ️ No hay materiales para sincronizar")
            return {'success': True, 'created': 0, 'updated': 0}

        updated_count = 0
        skipped_count = 0
        errors = []

        connection_params = {
            "host": "ad17solutions.dscloud.me",
            "port": 3307,
            "user": "IvanUriel",
            "password": "iuOp20!!25",
            "database": "AD17_Materiales",
            "charset": 'utf8mb4',
            "cursorclass": pymysql.cursors.DictCursor
        }

        conn = pymysql.connect(**connection_params)
        try:
            with conn.cursor() as cursor:
                now = datetime.now(timezone.utc)

                for mat in local_materials:
                    try:
                        mat_id = int(mat.code) if mat.code.isdigit() else None

                        if not mat_id:
                            skipped_count += 1
                            continue

                        # Verificar si el material existe en tabla Datos
                        cursor.execute(
                            "SELECT regID FROM Datos WHERE matID = %s ORDER BY regID DESC LIMIT 1",
                            (mat_id,)
                        )
                        existing = cursor.fetchone()

                        if existing:
                            # Solo actualizar nombre y descripción (campos de texto)
                            cursor.execute(
                                """UPDATE Datos
                                   SET nombre = %s,
                                       descripcion = %s,
                                       reutilizable = %s,
                                       timestamp = %s
                                   WHERE regID = %s""",
                                (
                                    mat.name,
                                    mat.description or '',
                                    1 if mat.can_reuse else 0,
                                    now,
                                    existing['regID']
                                )
                            )
                            if cursor.rowcount > 0:
                                updated_count += 1
                        else:
                            # Material no existe en remoto, lo omitimos
                            # (crear nuevos requiere IDs de unidad/categoría)
                            skipped_count += 1

                        if updated_count <= 5:
                            app.logger.debug(f"✅ Material {mat.code}: {mat.name}")

                    except Exception as e:
                        error_msg = f"Error escribiendo material {mat.code}: {str(e)}"
                        app.logger.error(error_msg)
                        errors.append(error_msg)
                        continue

                conn.commit()

        finally:
            conn.close()

        result_msg = f"✅ Escritura materiales: {updated_count} actualizados, {skipped_count} omitidos"
        app.logger.info(result_msg)

        return {
            'success': True,
            'created': 0,
            'updated': updated_count,
            'skipped': skipped_count,
            'total': len(local_materials),
            'errors': errors[:10]
        }

    except Exception as e:
        error_msg = f'Error escribiendo materiales a BD remota: {str(e)}'
        app.logger.error(error_msg)
        import traceback
        app.logger.error(traceback.format_exc())
        return {'success': False, 'error': error_msg}


def write_stocks_to_remote():
    """
    Escribir stocks locales a la tabla Stocks de la BD remota.
    Esta tabla se crea con create_stocks_table.py
    """
    try:
        app.logger.info("📤 Iniciando escritura de stocks a BD remota...")

        # Obtener todos los materiales locales con stock
        local_materials = Material.query.all()

        if not local_materials:
            return {'success': True, 'synced': 0, 'message': 'No hay materiales para sincronizar'}

        connection_params = {
            "host": "ad17solutions.dscloud.me",
            "port": 3307,
            "user": "IvanUriel",
            "password": "iuOp20!!25",
            "database": "AD17_Materiales",
            "charset": 'utf8mb4',
            "cursorclass": pymysql.cursors.DictCursor
        }

        synced = 0
        updated = 0
        errors = []

        conn = pymysql.connect(**connection_params)
        try:
            with conn.cursor() as cursor:
                # Verificar si la tabla Stocks existe
                cursor.execute("SHOW TABLES LIKE 'Stocks'")
                if not cursor.fetchone():
                    return {
                        'success': False,
                        'error': 'Tabla Stocks no existe. Ejecutar: python create_stocks_table.py create'
                    }

                now = datetime.now(timezone.utc)

                for mat in local_materials:
                    try:
                        # Determinar estado del stock
                        if mat.min_stock > 0 and mat.current_stock <= mat.min_stock:
                            status = 'bajo'
                        elif mat.max_stock > 0 and mat.current_stock >= mat.max_stock:
                            status = 'alto'
                        else:
                            status = 'ok'

                        # Calcular valor total
                        total_value = mat.current_stock * (mat.unit_cost or 0)

                        # Verificar si ya existe registro para este material
                        cursor.execute(
                            "SELECT regID FROM Stocks WHERE material_code = %s LIMIT 1",
                            (mat.code,)
                        )
                        existing = cursor.fetchone()

                        if existing:
                            # Actualizar registro existente
                            cursor.execute("""
                                UPDATE Stocks SET
                                    material_name = %s,
                                    category = %s,
                                    unit = %s,
                                    current_stock = %s,
                                    min_stock = %s,
                                    max_stock = %s,
                                    unit_cost = %s,
                                    total_value = %s,
                                    last_movement = %s,
                                    status = %s,
                                    synced_at = %s
                                WHERE regID = %s
                            """, (
                                mat.name,
                                mat.category,
                                mat.unit,
                                mat.current_stock,
                                mat.min_stock,
                                mat.max_stock,
                                mat.unit_cost or 0,
                                total_value,
                                mat.last_movement,
                                status,
                                now,
                                existing['regID']
                            ))
                            updated += 1
                        else:
                            # Insertar nuevo registro
                            cursor.execute("""
                                INSERT INTO Stocks (
                                    material_code, material_name, category, unit,
                                    current_stock, min_stock, max_stock, unit_cost,
                                    total_value, last_movement, status, synced_at
                                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, (
                                mat.code,
                                mat.name,
                                mat.category,
                                mat.unit,
                                mat.current_stock,
                                mat.min_stock,
                                mat.max_stock,
                                mat.unit_cost or 0,
                                total_value,
                                mat.last_movement,
                                status,
                                now
                            ))
                            synced += 1

                    except Exception as e:
                        errors.append(f"Error en {mat.code}: {str(e)}")
                        continue

                conn.commit()

        finally:
            conn.close()

        result_msg = f"✅ Stocks: {synced} nuevos, {updated} actualizados"
        app.logger.info(result_msg)

        return {
            'success': True,
            'synced': synced,
            'updated': updated,
            'total': len(local_materials),
            'errors': errors[:10]
        }

    except Exception as e:
        error_msg = f'Error escribiendo stocks a BD remota: {str(e)}'
        app.logger.error(error_msg)
        import traceback
        app.logger.error(traceback.format_exc())
        return {'success': False, 'error': error_msg}


# ============================================================
# SINCRONIZACIÓN AUTOMÁTICA PROGRAMADA
# ============================================================
SYNC_INTERVAL_MINUTES = 5  # Intervalo de sincronización en minutos

def scheduled_sync():
    """Tarea programada para sincronización bidireccional COMPLETA"""
    with app.app_context():
        try:
            now = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
            app.logger.info(f"🔄 [{now}] Ejecutando sincronización bidireccional...")

            # ========== LECTURA: BD Remota → Local ==========
            app.logger.info("   📥 LEYENDO desde BD remota...")

            # 1. Sync materiales desde remoto
            result_materials = sync_materials_from_remote()
            if result_materials['success']:
                app.logger.info(f"      Materiales: {result_materials['synced']} nuevos, {result_materials['updated']} actualizados")

            # 2. Sync min/max desde remoto
            result_minmax = sync_minmax_from_remote()
            if result_minmax['success']:
                app.logger.info(f"      Min/Max: {result_minmax['updated']} leídos")

            # ========== ESCRITURA: Local → BD Remota ==========
            app.logger.info("   📤 ESCRIBIENDO a BD remota...")

            # 3. Escribir materiales a remoto
            result_write_mat = write_materials_to_remote()
            if result_write_mat['success']:
                app.logger.info(f"      Materiales: {result_write_mat.get('created', 0)} nuevos, {result_write_mat.get('updated', 0)} actualizados")

            # 4. Escribir min/max a remoto
            result_write_minmax = write_minmax_to_remote()
            if result_write_minmax['success']:
                app.logger.info(f"      Min/Max: {result_write_minmax['updated']} escritos")

            app.logger.info(f"✅ [{now}] Sincronización bidireccional completada")

        except Exception as e:
            app.logger.error(f"❌ Error en sincronización automática: {str(e)}")


# Inicializar scheduler (se activará en el bloque principal)
scheduler = BackgroundScheduler()

def start_scheduler():
    """Inicia el scheduler de sincronización automática"""
    if not scheduler.running:
        scheduler.add_job(
            func=scheduled_sync,
            trigger='interval',
            minutes=SYNC_INTERVAL_MINUTES,
            id='sync_job',
            name='Sincronización bidireccional automática',
            replace_existing=True
        )
        scheduler.start()
        app.logger.info(f"⏰ Scheduler iniciado - Sincronización cada {SYNC_INTERVAL_MINUTES} minutos")

        # Ejecutar sincronización inicial
        scheduled_sync()

def stop_scheduler():
    """Detiene el scheduler al cerrar la aplicación"""
    if scheduler.running:
        scheduler.shutdown()
        app.logger.info("⏰ Scheduler detenido")

# Registrar función de limpieza al salir
atexit.register(stop_scheduler)

# --- Mapa de permisos por rol ---
ROLE_PERMISSIONS = {
    'requisitador': {
        'view_inventory', 'create_requisition', 'list_requisitions',
        'view_project_deliveries', 'view_movements_readonly'
    },
    'almacenista': {
        'view_inventory', 'create_requisition', 'list_requisitions',
        'view_project_deliveries', 'view_movements_readonly',
        'register_movements', 'view_material_usage_charts'
    },
    'admin': {
        '*', 'view_inventory', 'create_requisition', 'list_requisitions',
        'view_project_deliveries', 'view_movements_readonly',
        'register_movements', 'view_material_usage_charts',
        'view_costs', 'edit_records', 'edit_materials', 'delete_movement'
    }
}

def has_perm(user, perm):
    if not user.is_authenticated:
        return False
    perms = ROLE_PERMISSIONS.get(user.role, set())
    return '*' in perms or perm in perms

def role_required(*roles):
    def deco(f):
        @wraps(f)
        def wrapper(*a, **kw):
            if not current_user.is_authenticated or current_user.role not in roles:
                flash('Permisos insuficientes')
                return redirect(url_for('dashboard'))
            return f(*a, **kw)
        return wrapper
    return deco

def permission_required(perm):
    def deco(f):
        @wraps(f)
        def wrapper(*a, **kw):
            if not current_user.is_authenticated or not has_perm(current_user, perm):
                flash('Permisos insuficientes')
                return redirect(url_for('dashboard'))
            return f(*a, **kw)
        return wrapper
    return deco


def seed_verification_codes_fixed():
    """
    Inserta códigos de verificación fijos y permanentes
    que nunca se marcan como usados.
    """
    codes = [
        ("REQ-AB12CD", "requisitador"),
        ("ALM-34EF56", "almacenista"),
        ("ADM-7890AA", "admin")
    ]

    for code, role in codes:
        existing = db.session.execute(
            text("SELECT 1 FROM verification_code WHERE code = :c"),
            {"c": code}
        ).first()
        if not existing:
            db.session.execute(
                text("""
                    INSERT INTO verification_code (code, role, expires_at, used_by, is_active, created_at)
                    VALUES (:code, :role, NULL, NULL, 1, :created)
                """),
                {"code": code, "role": role, "created": datetime.utcnow()}
            )
    db.session.commit()
    print("✅ Códigos de verificación fijos insertados o actualizados.")
@app.template_filter('money_if_allowed')
def money_if_allowed(value):
    try:
        if current_user.is_authenticated and has_perm(current_user, 'view_costs'):
            return f"${value:,.2f}"
        return "N/D"
    except Exception:
        return "N/D"

@app.context_processor
def inject_permissions():
    return dict(has_perm=lambda p: has_perm(current_user, p))


@app.route('/api/sync/materials', methods=['POST'])
@login_required
def api_sync_materials():
    """API para sincronizar materiales desde la base remota"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Permisos insuficientes'})

    result = sync_materials_from_remote()

    if result['success']:
        return jsonify({
            'success': True,
            'message': f"Sincronización completada: {result['synced']} nuevos, {result['updated']} actualizados",
            'details': result
        })
    else:
        return jsonify({
            'success': False,
            'message': f"Error en sincronización: {result['error']}"
        })

@app.route('/api/sync/minmax', methods=['POST'])
@login_required
def api_sync_minmax():
    """API para sincronizar valores min/max desde registro_materiales"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Permisos insuficientes'})

    result = sync_minmax_from_remote()

    if result['success']:
        return jsonify({
            'success': True,
            'message': f"Sincronización min/max completada: {result['updated']} actualizados, {result['not_found']} no encontrados",
            'details': result
        })
    else:
        return jsonify({
            'success': False,
            'message': f"Error en sincronización: {result.get('error', 'Unknown error')}"
        })


@app.route('/api/sync/write-remote', methods=['POST'])
@login_required
def api_write_to_remote():
    """API para escribir valores min/max locales a la BD remota"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Permisos insuficientes'})

    result = write_minmax_to_remote()

    if result['success']:
        return jsonify({
            'success': True,
            'message': f"Escritura completada: {result['updated']} materiales escritos a BD remota",
            'details': result
        })
    else:
        return jsonify({
            'success': False,
            'message': f"Error en escritura: {result.get('error', 'Unknown error')}"
        })


@app.route('/api/sync/full', methods=['POST'])
@login_required
def api_full_sync():
    """API para forzar sincronización bidireccional completa"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Permisos insuficientes'})

    results = {}

    # ========== LECTURA: BD Remota → Local ==========
    results['materials_read'] = sync_materials_from_remote()
    results['minmax_read'] = sync_minmax_from_remote()

    # ========== ESCRITURA: Local → BD Remota ==========
    results['materials_write'] = write_materials_to_remote()
    results['minmax_write'] = write_minmax_to_remote()
    results['stocks_write'] = write_stocks_to_remote()

    return jsonify({
        'success': True,
        'message': 'Sincronización bidireccional completa ejecutada',
        'details': results
    })


@app.route('/api/sync/stocks', methods=['POST'])
@login_required
def api_sync_stocks():
    """API para sincronizar stocks a BD remota"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Permisos insuficientes'})

    result = write_stocks_to_remote()

    if result['success']:
        return jsonify({
            'success': True,
            'message': f"Sincronización completada: {result.get('synced', 0)} nuevos, {result.get('updated', 0)} actualizados",
            'details': result
        })
    else:
        return jsonify({
            'success': False,
            'message': result.get('error', 'Error desconocido')
        })

@app.route('/api/remote/categories')
@login_required
def get_remote_categories():
    """Obtener categorías de la base remota"""
    categories = get_remote_categories_for_select()
    return jsonify({'categories': categories})

@app.route('/api/remote/units')
@login_required
def get_remote_units():
    """Obtener unidades de la base remota"""
    units = get_remote_units_for_select()
    return jsonify({'units': units})


@app.route('/api/sync/categories', methods=['POST'])
@login_required
def api_sync_categories():
    """Sincronizar categorías desde BD remota a tabla local"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Permisos insuficientes'})

    try:
        # Obtener categorías de BD remota
        remote_categories = remote_db.get_categorias_habilitadas()

        if not remote_categories:
            return jsonify({'success': False, 'message': 'No se encontraron categorías en BD remota'})

        synced = 0
        created = 0

        for cat in remote_categories:
            remote_id = int(cat.get('id', 0)) if cat.get('id') else None
            name = cat.get('categoria', '').strip()

            if not name:
                continue

            # Buscar por remote_id o crear
            existing = Category.query.filter_by(remote_id=remote_id).first() if remote_id else None

            if existing:
                existing.name = name
                existing.synced_at = datetime.utcnow()
                synced += 1
            else:
                # Verificar si es categoría de telas
                is_fabric = 'tela' in name.lower() or 'textil' in name.lower()

                new_cat = Category(
                    remote_id=remote_id,
                    name=name,
                    is_fabric=is_fabric,
                    synced_at=datetime.utcnow()
                )
                db.session.add(new_cat)
                created += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Sincronización completada: {created} creadas, {synced} actualizadas',
            'created': created,
            'synced': synced
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/api/sync/units', methods=['POST'])
@login_required
def api_sync_units():
    """Sincronizar unidades desde BD remota a tabla local"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Permisos insuficientes'})

    try:
        # Obtener unidades de BD remota
        remote_units = remote_db.get_unidades_habilitadas()

        if not remote_units:
            return jsonify({'success': False, 'message': 'No se encontraron unidades en BD remota'})

        synced = 0
        created = 0

        for unit in remote_units:
            remote_id = int(unit.get('id', 0)) if unit.get('id') else None
            name = unit.get('unidad', '').strip() or unit.get('singular', '').strip()
            abbr = unit.get('singular', '').strip()

            if not name:
                continue

            # Buscar por remote_id o crear
            existing = Unit.query.filter_by(remote_id=remote_id).first() if remote_id else None

            if existing:
                existing.name = name
                existing.abbreviation = abbr
                existing.synced_at = datetime.utcnow()
                synced += 1
            else:
                new_unit = Unit(
                    remote_id=remote_id,
                    name=name,
                    abbreviation=abbr,
                    synced_at=datetime.utcnow()
                )
                db.session.add(new_unit)
                created += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Sincronización completada: {created} creadas, {synced} actualizadas',
            'created': created,
            'synced': synced
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/api/materials/next-code')
@login_required
def get_next_material_code():
    """Genera el siguiente código de material automáticamente"""
    try:
        # Buscar el último código MAT-XXXX
        last_material = Material.query.filter(
            Material.code.like('MAT-%')
        ).order_by(Material.id.desc()).first()

        if last_material and last_material.code.startswith('MAT-'):
            try:
                last_num = int(last_material.code.replace('MAT-', ''))
                next_num = last_num + 1
            except ValueError:
                next_num = 1
        else:
            # Contar todos los materiales + 1
            next_num = Material.query.count() + 1

        next_code = f"MAT-{next_num:04d}"

        # Verificar que no exista
        while Material.query.filter_by(code=next_code).first():
            next_num += 1
            next_code = f"MAT-{next_num:04d}"

        return jsonify({'success': True, 'code': next_code})

    except Exception as e:
        return jsonify({'success': False, 'code': 'MAT-0001', 'error': str(e)})


@app.route('/api/materials/check-duplicate')
@login_required
def check_material_duplicate():
    """Busca materiales con nombres similares para evitar duplicados"""
    name = request.args.get('name', '').strip()

    if len(name) < 2:
        return jsonify({'materials': []})

    try:
        # Buscar materiales con nombre similar
        similar = Material.query.filter(
            Material.name.ilike(f'%{name}%')
        ).limit(5).all()

        results = [{
            'id': m.id,
            'code': m.code,
            'name': m.name,
            'category': m.category
        } for m in similar]

        return jsonify({'materials': results})

    except Exception as e:
        return jsonify({'materials': [], 'error': str(e)})

@app.route('/api/remote/materials/search')
@login_required
def search_remote_materials():
    """Buscar materiales en la base remota"""
    query = request.args.get('q', '')

    if len(query) < 2:
        return jsonify({'materials': []})

    try:
        all_materials = remote_db.get_materiales_habilitados()

        # Filtrar materiales que coincidan con la búsqueda
        filtered_materials = []
        for mat in all_materials:
            if (query.lower() in str(mat['material']).lower() or
                query.lower() in str(mat['id']).lower()):
                filtered_materials.append({
                    'id': mat['id'],
                    'name': mat['material'],
                    'unit': mat['uni_simbolo'],
                    'category': mat['categoria'],
                    'description': mat['descripcion']
                })

        return jsonify({'materials': filtered_materials[:20]})  # Limitar a 20 resultados

    except Exception as e:
        return jsonify({'error': str(e), 'materials': []})
# Función para validar datos de Excel
def validate_excel_row(row_data, row_number):
    """Valida una fila de datos de Excel"""
    errors = []

    # Validar campos requeridos
    required_fields = ['Material', 'Movimiento', 'Cantidad']
    for field in required_fields:
        if pd.isna(row_data.get(field)) or str(row_data.get(field)).strip() == '':
            errors.append(f"Campo requerido '{field}' está vacío")

    # Validar tipo de movimiento
    if not pd.isna(row_data.get('Movimiento')):
        movement_type = str(row_data['Movimiento']).lower().strip()
        if movement_type not in ['entrada', 'salida', 'retorno']:
            errors.append(f"Tipo de movimiento inválido: '{row_data['Movimiento']}'")

    # Validar cantidad
    try:
        quantity = float(row_data.get('Cantidad', 0))
        if quantity <= 0:
            errors.append("La cantidad debe ser mayor a 0")
    except (ValueError, TypeError):
        errors.append("La cantidad debe ser un número válido")

    # Validar rollos si se proporciona
    if not pd.isna(row_data.get('Rollos')):
        try:
            rollos = int(row_data['Rollos'])
            if rollos < 0:
                errors.append("Los rollos no pueden ser negativos")
        except (ValueError, TypeError):
            errors.append("Los rollos deben ser un número entero")

    return errors

# Rutas principales
@app.route('/')
@login_required
def dashboard():
    # Estadísticas generales
    total_materials = Material.query.count()
    low_stock_materials = Material.query.filter(Material.current_stock <= Material.min_stock).count()
    
    # Personalizar KPIs y Alertas según rol
    alerts = []
    
    if current_user.role == 'requisitador':
        # Solo sus propias requisiciones pendientes
        pending_requests = Request.query.filter_by(
            user_id=current_user.id, 
            status='pendiente'
        ).count()
        
        # Alertas de sus requisiciones (ej. cambios recientes o estado actual)
        # Por ahora mostramos las últimas 5
        my_recent_requests = Request.query.filter_by(user_id=current_user.id).order_by(Request.created_at.desc()).limit(5).all()
        for req in my_recent_requests:
            alert_type = 'info'
            if req.status == 'pendiente': alert_type = 'warning'
            elif req.status == 'aprobada': alert_type = 'success'
            elif req.status == 'rechazada': alert_type = 'danger'
            elif req.status == 'entregada': alert_type = 'primary'
            
            alerts.append({
                'type': alert_type,
                'message': f'Tu Requisición #{req.id} está actualmente: {req.status.upper()}'
            })
            
    else:
        # Admin / Almacenista / Líder
        pending_requests = Request.query.filter_by(status='pendiente').count()

        # Alertas de Stock Bajo (solo para quienes gestionan stock)
        low_stock = Material.query.filter(Material.current_stock <= Material.min_stock).all()
        for material in low_stock:
            alerts.append({
                'type': 'warning',
                'message': f'Stock bajo: {material.name} ({material.current_stock} {material.unit})'
            })

    # Materiales sin movimiento en 6 meses
    six_months_ago = datetime.utcnow() - timedelta(days=180)
    no_movement_materials = Material.query.filter(
        (Material.last_movement < six_months_ago) | (Material.last_movement.is_(None))
    ).count()

    return render_template('dashboard.html',
                         total_materials=total_materials,
                         low_stock_materials=low_stock_materials,
                         pending_requests=pending_requests,
                         no_movement_materials=no_movement_materials,
                         alerts=alerts)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username      = request.form.get('username', '').strip()
        password      = request.form.get('password', '').strip()
        selected_role = request.form.get('selected_role', '').strip().lower()  # viene del modal (opcional)

        if not username or not password:
            flash('Usuario y contraseña son requeridos.', 'danger')
            return redirect(url_for('login'))

        user = User.query.filter_by(username=username).first()
        if not user or not check_password_hash(user.password_hash, password):
            flash('Usuario o contraseña inválidos.', 'danger')
            return redirect(url_for('login'))

        # Si el modal trae un rol y no coincide, lo dejamos entrar igual pero avisamos.
        if selected_role and selected_role != (user.role or '').lower():
            flash(
                f'Ingresaste como {user.role}. (El perfil elegido en la tarjeta era "{selected_role}")',
                'info'
            )

        login_user(user)
        flash(f'¡Bienvenido, {user.username}!', 'success')
        return redirect(url_for('dashboard'))

    # GET: mostrar login
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/materials')
@login_required
def materials():
    """Vista de materiales con filtros mejorados"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    category = request.args.get('category', '', type=str)
    material_type = request.args.get('type', 'materials', type=str)

    # Filtros de características
    consumibles_filter = request.args.get('consumibles', '', type=str)
    reciclables_filter = request.args.get('reciclables', '', type=str)
    reutilizables_filter = request.args.get('reutilizables', '', type=str)
    reciclados_filter = request.args.get('reciclados', '', type=str)
    error_stock_filter = request.args.get('error_stock', '', type=str)

    query = Material.query

    # Filtro de búsqueda
    if search:
        query = query.filter(
            (Material.name.contains(search)) |
            (Material.code.contains(search))
        )

    # Filtro de categoría
    if category:
        query = query.filter_by(category=category)

    # Filtros de tipo de material
    if consumibles_filter == 'true':
        query = query.filter_by(is_consumible=True)

    if reciclables_filter == 'true':
        query = query.filter_by(can_recycle=True)

    if reutilizables_filter == 'true':
        query = query.filter_by(can_reuse=True)

    if reciclados_filter == 'true':
        query = query.filter_by(is_recycled=True)

    # Filtro de error de stock
    if error_stock_filter == 'true':
        query = query.filter(
            db.or_(
                db.and_(
                    Material.min_stock > 0,
                    Material.current_stock <= Material.min_stock
                ),
                db.and_(
                    Material.max_stock > 0,
                    Material.current_stock >= Material.max_stock
                )
            )
        )

    materials_paginated = query.order_by(Material.name).paginate(page=page, per_page=20, error_out=False)

    # Contar consumibles para badge
    consumibles_count = Material.query.filter_by(is_consumible=True).count()

    # Obtener categorías
    try:
        remote_categories = get_remote_categories_for_select()
        categories = [cat['name'] for cat in remote_categories]
    except Exception as e:
        app.logger.error(f"Error obteniendo categorías remotas: {e}")
        categories = [c[0] for c in db.session.query(Material.category).distinct().all() if c[0]]

    return render_template('materials.html',
                         materials=materials_paginated,
                         categories=categories,
                         search=search,
                         category=category,
                         material_type=material_type,
                         consumibles_filter=consumibles_filter,
                         reciclables_filter=reciclables_filter,
                         reutilizables_filter=reutilizables_filter,
                         reciclados_filter=reciclados_filter,
                         error_stock_filter=error_stock_filter,
                         consumibles_count=consumibles_count)

@app.route('/materials/add', methods=['GET', 'POST'])
@login_required
def add_material():
    if current_user.role == 'requisitador':
        flash('No tienes permisos para agregar materiales.', 'danger')
        return redirect(url_for('materials'))

    if request.method == 'POST':
        # Auto-generate material code
        last_material = Material.query.order_by(Material.id.desc()).first()
        if last_material:
            # Extract number from last code (e.g., MAT-0001 -> 1)
            try:
                last_num = int(last_material.code.split('-')[-1])
                new_code = f"MAT-{str(last_num + 1).zfill(4)}"
            except:
                new_code = f"MAT-{str(last_material.id + 1).zfill(4)}"
        else:
            new_code = "MAT-0001"

        # Verify code is unique
        while Material.query.filter_by(code=new_code).first():
            try:
                num = int(new_code.split('-')[-1])
                new_code = f"MAT-{str(num + 1).zfill(4)}"
            except:
                new_code = f"MAT-{str(Material.query.count() + 1).zfill(4)}"

        # Obtener category_id y unit_id si están usando tablas sincronizadas
        category_id = None
        unit_id = None
        category_name = request.form.get('category', '')
        unit_name = request.form.get('unit', '')

        # Buscar categoría local
        if category_name:
            cat = Category.query.filter_by(name=category_name).first()
            if cat:
                category_id = cat.id

        # Buscar unidad local
        if unit_name:
            unit = Unit.query.filter_by(name=unit_name).first()
            if unit:
                unit_id = unit.id

        # Determine if it's fabric based on category name
        is_fabric = category_name.lower() in ['tela', 'telas', 'fabric', 'fabrics'] if category_name else False

        material = Material(
            code=new_code,
            name=request.form['name'],
            description=request.form.get('description', ''),
            unit=unit_name,
            category=category_name,
            category_id=category_id,
            unit_id=unit_id,
            current_stock=float(request.form.get('current_stock', 0) or 0),
            min_stock=float(request.form.get('min_stock', 0) or 0),
            max_stock=float(request.form.get('max_stock', 0) or 0),
            unit_cost=float(request.form.get('unit_cost', 0) or 0),
            # Opciones de tela (is_fabric_roll se determina por la categoría)
            is_fabric_roll=is_fabric,
            fabric_width=float(request.form.get('fabric_width', 0) or 0) if request.form.get('fabric_width') else None,
            # Opciones de reciclaje (mutuamente exclusivas)
            can_recycle=bool(request.form.get('can_recycle')),
            can_reuse=bool(request.form.get('can_reuse')),
            is_recycled=bool(request.form.get('is_recycled')),
            is_pre_recycled=False,  # Removed per requirements
            recycled_from_id=int(request.form.get('recycled_from_id')) if request.form.get('recycled_from_id') else None,
            is_consumible=bool(request.form.get('is_consumible'))
        )

        db.session.add(material)
        db.session.commit()
        flash(f'Material "{material.name}" agregado exitosamente con código {new_code}')
        return redirect(url_for('materials'))

    # GET request - pasar categorías y unidades (locales o remotas)
    # Primero intentar cargar de tablas locales sincronizadas
    local_categories = Category.query.filter_by(is_active=True).all()
    local_units = Unit.query.filter_by(is_active=True).all()

    # Si no hay datos locales, usar remotos
    if not local_categories:
        remote_categories = get_remote_categories_for_select()
    else:
        remote_categories = [{'id': c.id, 'name': c.name, 'is_fabric': c.is_fabric} for c in local_categories]

    if not local_units:
        remote_units = get_remote_units_for_select()
    else:
        remote_units = [{'id': u.id, 'name': u.name, 'abbreviation': u.abbreviation} for u in local_units]

    # Obtener solo materiales reciclables para el dropdown "Es Reciclado"
    recyclable_materials = Material.query.filter_by(can_recycle=True).order_by(Material.name).all()

    return render_template('add_material.html',
                         remote_categories=remote_categories,
                         remote_units=remote_units,
                         recyclable_materials=recyclable_materials)

# Actualizar la ruta de requests para soportar filtros adicionales
@app.route('/requests')
@login_required
def requests():
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', '', type=str)
    department_filter = request.args.get('department', '', type=str)
    incident_filter = request.args.get('incident_filter', '', type=str)
    search_filter = request.args.get('search', '', type=str)

    client_filter = request.args.get('client', '', type=str)
    sort_by = request.args.get('sort', 'created_at_desc', type=str)

    query = Request.query

    # Obtener lista de clientes únicos para el filtro
    clients = db.session.query(Project.client).filter(Project.client.isnot(None), Project.client != '').distinct().order_by(Project.client).all()
    unique_clients = [c[0] for c in clients]

    # Aplicar filtros de permisos
    if not current_user.role == 'admin':
        query = query.filter_by(user_id=current_user.id)

    # Aplicar filtros
    if status_filter:
        query = query.filter_by(status=status_filter)

    if department_filter:
        query = query.filter_by(department=department_filter)
        
    if client_filter:
        query = query.filter(Request.project.has(Project.client == client_filter))

    if incident_filter == 'incident':
        query = query.filter_by(is_incident=True)
    elif incident_filter == 'normal':
        query = query.filter_by(is_incident=False)

    if search_filter:
        query = query.filter(
            db.or_(
                Request.request_number.contains(search_filter),
                Request.project.has(Project.name.contains(search_filter)),
                Request.project.has(Project.fp_code.contains(search_filter)),
                Request.user.has(User.username.contains(search_filter))
            )
        )

    # Ordenamiento
    if sort_by == 'created_at_asc':
        query = query.order_by(Request.created_at.asc())
    elif sort_by == 'deadline_asc':
        # Nulls last logic might be needed but for simplicity:
        query = query.order_by(Request.acquisition_deadline.asc())
    elif sort_by == 'deadline_desc':
        query = query.order_by(Request.acquisition_deadline.desc())
    elif sort_by == 'status':
        query = query.order_by(Request.status)
    else: # created_at_desc
        query = query.order_by(Request.created_at.desc())

    requests = query.paginate(page=page, per_page=20, error_out=False)

    # Datos para la vista "Por Proyectos" y "Todos los Proyectos"
    # Proyectos Activos: Aquellos con al menos una requisición NO completada ni cancelada
    active_projects = Project.query.join(Request).filter(
        Request.status.notin_(['completada', 'cancelada'])
    ).distinct().order_by(Project.fp_code.desc()).all()

    # Todos los proyectos (para la nueva sección)
    all_projects = Project.query.order_by(Project.fp_code.desc()).all()

    return render_template('requests.html',
                         requests=requests,
                         status_filter=status_filter,
                         department_filter=department_filter,
                         client_filter=client_filter,
                         incident_filter=incident_filter,
                         search_filter=search_filter,
                         sort_by=sort_by,
                         unique_clients=unique_clients,
                         active_projects=active_projects,
                         all_projects=all_projects)
# Agregar estas rutas a tu app.py existente
# Asegúrate de tener estos imports al inicio de tu app.py:
# from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, current_app
@app.route('/search_fp', methods=['GET'])
@login_required
def search_fp():
    """Ruta para buscar proyectos por FP - corregida para retornar formato correcto"""
    fp = request.args.get('fp', '').strip()
    app.logger.debug("search_fp: FP recibido: '%s'", fp)

    if not fp:
        app.logger.debug("search_fp: FP vacío")
        return jsonify({"success": False, "error": "No se proporcionó un FP"}), 400

    try:
        # Primero buscar en la base de datos local
        local_projects = Project.query.filter(
            Project.fp_code.like(f'{fp}%')
        ).all()

        local_results = []
        for project in local_projects:
            local_results.append({
                'id': project.id,
                'fp_code': project.fp_code,
                'nombre': project.name,
                'cliente': 'Cliente Local',
                'source': 'local'
            })

        if local_results:
            app.logger.debug("search_fp: Encontrados %d proyectos locales", len(local_results))
            return jsonify({"success": True, "data": local_results})

        # Si no hay resultados locales, buscar en la base remota
        app.logger.debug("search_fp: Conectando a la base de datos remota...")

        with remote_db.get_connection(database="AD17_Proyectos") as conn:
            with conn.cursor() as cursor:
                query = """
                    SELECT i.regID AS fp,
                           d.nombre AS proyecto,
                           c.nombre AS cliente
                    FROM (SELECT * FROM AD17_Proyectos.ID WHERE regID LIKE %s) AS i
                    LEFT JOIN (SELECT * FROM AD17_Proyectos.Datos
                               WHERE regID IN (SELECT MAX(regID) FROM AD17_Proyectos.Datos GROUP BY fp)) AS d
                      ON d.fp LIKE i.regID
                    LEFT JOIN (SELECT * FROM AD17_Clientes.Datos
                               WHERE regID IN (SELECT MAX(regID) FROM AD17_Clientes.Datos GROUP BY cliID)) AS c
                      ON c.cliID LIKE d.cliente
                    ORDER BY fp DESC;
                """
                param = fp + '%'
                app.logger.debug("search_fp: Ejecutando query con parámetro: '%s'", param)
                cursor.execute(query, (param,))
                remote_results = cursor.fetchall()
                app.logger.debug("search_fp: Query ejecutada. Resultados: %s", remote_results)

        # Procesar resultados remotos
        final_results = []
        for remote_project in remote_results:
            fp_value = str(remote_project['fp'])
            project_name = remote_project['proyecto'] or f"Proyecto {fp_value}"
            client_name = remote_project['cliente'] or "Cliente desconocido"

            # Verificar si ya existe localmente
            existing_project = Project.query.filter_by(fp_code=fp_value).first()

            if not existing_project:
                try:
                    # Crear el proyecto localmente
                    from datetime import date, timedelta
                    today = date.today()

                    new_project = Project(
                        fp_code=fp_value,
                        name=project_name,
                        delivery_date=today + timedelta(days=365),
                        production_start=today,
                        assembly_date=today + timedelta(days=300),
                        status='activo'
                    )
                    db.session.add(new_project)
                    db.session.commit()
                    project_id = new_project.id
                    app.logger.debug("search_fp: Creado proyecto local con ID %d", project_id)
                except Exception as e:
                    db.session.rollback()
                    app.logger.error("search_fp: Error creando proyecto: %s", str(e))
                    continue
            else:
                project_id = existing_project.id
                project_name = existing_project.name
                app.logger.debug("search_fp: Proyecto ya existía con ID %d", project_id)

            final_results.append({
                'id': project_id,
                'fp_code': fp_value,
                'nombre': project_name,
                'cliente': client_name,
                'source': 'remote'
            })

        app.logger.debug("search_fp: Retornando %d datos.", len(final_results))
        return jsonify({"success": True, "data": final_results})

    except Exception as e:
        app.logger.error("Error en search_fp: %s", str(e))
        return jsonify({"success": False, "error": "Error interno en la búsqueda"}), 500

@app.route('/api/project_details/<fp_code>', methods=['GET'])
@login_required
def get_project_details(fp_code):
    """Obtiene los detalles de un proyecto desde la base de datos remota para el tooltip"""
    if not fp_code:
        return jsonify({"success": False, "error": "No se proporcionó un FP"}), 400

    try:
        with remote_db.get_connection(database="AD17_Proyectos") as conn:
            with conn.cursor() as cursor:
                query = """
                    SELECT d.nombre AS proyecto,
                           c.nombre AS cliente,
                           v.nombre AS vendedora,
                           l.nombre AS lider
                    FROM AD17_Proyectos.ID AS i
                    LEFT JOIN (SELECT * FROM AD17_Proyectos.Datos 
                               WHERE regID IN (SELECT MAX(regID) FROM AD17_Proyectos.Datos GROUP BY fp)) AS d 
                      ON d.fp = i.regID
                    LEFT JOIN (SELECT * FROM AD17_Clientes.Datos 
                               WHERE regID IN (SELECT MAX(regID) FROM AD17_Clientes.Datos GROUP BY cliID)) AS c 
                      ON c.cliID = d.cliente
                    LEFT JOIN (SELECT * FROM AD17_Vendedores.Datos 
                               WHERE regID IN (SELECT MAX(regID) FROM AD17_Vendedores.Datos GROUP BY vendID)) AS v 
                      ON v.vendID = d.vendedora
                    LEFT JOIN (SELECT * FROM AD17_Lideres.Datos 
                               WHERE regID IN (SELECT MAX(regID) FROM AD17_Lideres.Datos GROUP BY lideID)) AS l 
                      ON l.lideID = d.lider
                    WHERE i.regID = %s;
                """
                cursor.execute(query, (fp_code,))
                result = cursor.fetchone()

        if result:
            return jsonify({
                "success": True,
                "data": {
                    "nombre": result['proyecto'] or "No especificado",
                    "cliente": result['cliente'] or "No especificado",
                    "vendedora": result['vendedora'] or "No especificado",
                    "lider": result['lider'] or "No especificado"
                }
            })
        else:
            # Fallback a base de datos local si no se encuentra remoto
            local_project = Project.query.filter_by(fp_code=fp_code).first()
            if local_project:
                return jsonify({
                    "success": True,
                    "data": {
                        "nombre": local_project.name,
                        "cliente": local_project.client or "No especificado",
                        "vendedora": "No disponible",
                        "lider": "No disponible"
                    }
                })
            return jsonify({"success": False, "error": "Proyecto no encontrado"}), 404
            
    except Exception as e:
        app.logger.error(f"Error al obtener detalles del proyecto {fp_code}: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
@app.route('/requests/new', methods=['GET', 'POST'])
@login_required
def new_request():
    """Ruta para crear nueva requisición - VERSIÓN CORREGIDA"""
    if request.method == 'POST':
        print("\n" + "="*60)
        print("🚀 NUEVA REQUISICIÓN - INICIO")
        print("="*60)

        try:
            # ===== 1. VALIDAR DATOS RECIBIDOS =====
            print("\n📥 DATOS RECIBIDOS:")
            print(f"  - Content-Type: {request.content_type}")
            print(f"  - Form keys: {list(request.form.keys())}")

            # Obtener todos los campos
            project_id = request.form.get('project_id')
            department = request.form.get('department', '').strip()
            materials_json = request.form.get('materials_json')

            print(f"  - project_id: {project_id}")
            print(f"  - department: {department}")
            print(f"  - materials_json length: {len(materials_json) if materials_json else 0}")

            # ===== 2. VALIDAR PROJECT_ID =====
            if not project_id or project_id == '' or project_id == 'null':
                error_msg = f'Debe seleccionar un proyecto válido. Project ID recibido: "{project_id}"'
                print(f"❌ ERROR: {error_msg}")
                return jsonify({
                    'success': False,
                    'message': error_msg
                }), 400

            # Convertir a entero
            try:
                project_id = int(project_id)
                if project_id <= 0:
                    error_msg = f'ID de proyecto inválido: {project_id}'
                    print(f"❌ ERROR: {error_msg}")
                    return jsonify({
                        'success': False,
                        'message': error_msg
                    }), 400
            except (ValueError, TypeError) as e:
                error_msg = f'ID de proyecto no es un número válido: {project_id}'
                print(f"❌ ERROR: {error_msg}")
                return jsonify({
                    'success': False,
                    'message': error_msg
                }), 400

            print(f"✅ Project ID válido: {project_id}")

            # ===== 3. BUSCAR PROYECTO =====
            project = Project.query.get(project_id)
            if not project:
                error_msg = f'Proyecto con ID {project_id} no encontrado'
                print(f"❌ ERROR: {error_msg}")
                return jsonify({
                    'success': False,
                    'message': error_msg
                }), 404

            print(f"✅ Proyecto encontrado: {project.name} (FP: {project.fp_code})")

            # ===== 4. VALIDAR DEPARTMENT =====
            if not department:
                error_msg = 'Debe seleccionar un departamento'
                print(f"❌ ERROR: {error_msg}")
                return jsonify({
                    'success': False,
                    'message': error_msg
                }), 400

            print(f"✅ Departamento válido: {department}")

            # ===== 5. VALIDAR MATERIALS_JSON =====
            if not materials_json:
                error_msg = 'No se recibieron materiales'
                print(f"❌ ERROR: {error_msg}")
                return jsonify({
                    'success': False,
                    'message': error_msg
                }), 400

            # Parsear JSON de materiales
            import json
            try:
                materials_list = json.loads(materials_json)
                if not materials_list or not isinstance(materials_list, list):
                    error_msg = 'Lista de materiales vacía o inválida'
                    print(f"❌ ERROR: {error_msg}")
                    return jsonify({
                        'success': False,
                        'message': error_msg
                    }), 400

                print(f"✅ Materiales parseados: {len(materials_list)} items")
                for idx, mat in enumerate(materials_list, 1):
                    print(f"   {idx}. {mat.get('code')} - {mat.get('name')} x {mat.get('quantity')}")

            except json.JSONDecodeError as e:
                error_msg = f'Error al parsear JSON de materiales: {str(e)}'
                print(f"❌ ERROR: {error_msg}")
                return jsonify({
                    'success': False,
                    'message': error_msg
                }), 400

            # ===== 6. CREAR REQUISICIÓN =====
            print("\n📝 CREANDO REQUISICIÓN...")

            # Generar número de requisición en formato RQ-{FP}-{CONTADOR}
            fp_code = project.fp_code or str(project.id)
            # Contar requisiciones existentes para este proyecto
            existing_count = Request.query.filter_by(project_id=project.id).count()
            req_number = f"RQ-{fp_code}-{existing_count + 1:02d}"
            print(f"  - Número generado: {req_number}")

            # Procesar fechas
            acquisition_deadline = None
            production_start = None
            assembly_start = None
            assembly_end = None

            # Fecha límite de adquisición
            if request.form.get('acquisition_deadline'):
                try:
                    acquisition_deadline = datetime.strptime(request.form['acquisition_deadline'], '%Y-%m-%d').date()
                    print(f"  - Fecha límite adquisición: {acquisition_deadline}")
                except ValueError as e:
                    print(f"⚠️ Advertencia: Fecha límite adquisición inválida: {e}")

            # Fecha inicio producción
            if request.form.get('production_start_date'):
                try:
                    production_start = datetime.strptime(request.form['production_start_date'], '%Y-%m-%d').date()
                    print(f"  - Fecha inicio producción: {production_start}")
                except ValueError as e:
                    print(f"⚠️ Advertencia: Fecha inicio producción inválida: {e}")

            # Fechas de montaje
            if request.form.get('assembly_start_date'):
                try:
                    assembly_start = datetime.strptime(request.form['assembly_start_date'], '%Y-%m-%d').date()
                    print(f"  - Fecha inicio montaje: {assembly_start}")
                except ValueError as e:
                    print(f"⚠️ Advertencia: Fecha inicio montaje inválida: {e}")

            if request.form.get('assembly_end_date'):
                try:
                    assembly_end = datetime.strptime(request.form['assembly_end_date'], '%Y-%m-%d').date()
                    print(f"  - Fecha fin montaje: {assembly_end}")
                except ValueError as e:
                    print(f"⚠️ Advertencia: Fecha fin montaje inválida: {e}")

            # Crear objeto Request con nuevos campos
            new_req = Request(
                request_number=req_number,
                project_id=project.id,
                user_id=current_user.id,
                department=department,
                # Campos para compatibilidad con BD existente (NOT NULL constraints)
                area=department,
                request_type='interno',  # Valor por defecto ya que se eliminó del UI
                is_incident=False,
                incident_id=None,
                # Nuevos campos
                acquisition_deadline=acquisition_deadline,
                production_start_date=production_start,
                assembly_start_date=assembly_start,
                assembly_end_date=assembly_end,
                has_returns=False,  # Se actualizará si hay materiales con will_return
                notes=request.form.get('notes', '')
            )

            db.session.add(new_req)
            db.session.flush()  # Obtener el ID

            print(f"✅ Requisición creada con ID: {new_req.id}")

            # ===== 8. PROCESAR MATERIALES =====
            print(f"\n📦 PROCESANDO {len(materials_list)} MATERIALES...")

            items_created = 0
            new_materials_created = 0

            for idx, material_data in enumerate(materials_list, 1):
                print(f"\n  [{idx}/{len(materials_list)}] Procesando: {material_data.get('name')}")

                try:
                    # Determinar si es material nuevo o existente
                    material_id = None
                    is_new = material_data.get('is_new_material', False)

                    # Auto-generar código si es nuevo y no viene informado
                    if is_new and not material_data.get('code'):
                        # Generar un código temporal o definitivo
                        # Opción simple: MAT-{timestamp} para evitar colisiones
                        import time
                        timestamp_code = int(time.time())
                        material_data['code'] = f"MAT-{timestamp_code}"
                        print(f"    ℹ️ Código auto-generado: {material_data['code']}")

                    if not is_new:
                        material_id = material_data.get('material_id')
                        if material_id:
                            print(f"    → Material existente ID: {material_id}")

                    # Procesar fecha de retorno
                    return_date = None
                    if material_data.get('return_expected_date'):
                        try:
                            return_date = datetime.strptime(
                                material_data['return_expected_date'],
                                '%Y-%m-%d'
                            ).date()
                            print(f"    → Fecha retorno: {return_date}")
                        except ValueError:
                            print(f"    ⚠️ Fecha retorno inválida, se omitirá")

                    # Crear RequestItem
                    item = RequestItem(
                        request_id=new_req.id,
                        material_id=material_id,
                        new_material_code=material_data.get('code') if is_new else None,
                        new_material_name=material_data.get('name') if is_new else None,
                        new_material_unit=material_data.get('unit') if is_new else None,
                        new_material_category=material_data.get('category') if is_new else None,
                        is_new_material=is_new,
                        quantity_requested=float(material_data['quantity']),
                        item_type=material_data.get('item_type', 'nuevo'),
                        will_return=material_data.get('will_return', False),
                        return_expected_date=return_date,
                        notes=material_data.get('notes', '')
                    )

                    db.session.add(item)
                    items_created += 1
                    print(f"    ✅ Item de requisición creado")

                    # Si es material nuevo, agregarlo al catálogo
                    if is_new:
                        code = material_data.get('code')
                        existing = Material.query.filter_by(code=code).first()

                        if not existing:
                            new_material = Material(
                                code=code,
                                name=material_data.get('name'),
                                unit=material_data.get('unit'),
                                category=material_data.get('category'),
                                description=f"Material creado desde requisición {req_number}",
                                current_stock=0,
                                min_stock=0,
                                max_stock=0,
                                unit_cost=0
                            )
                            db.session.add(new_material)
                            new_materials_created += 1
                            print(f"    ✅ Material nuevo agregado al catálogo: {code}")
                        else:
                            print(f"    ℹ️ Material {code} ya existe en catálogo")

                except Exception as item_error:
                    print(f"    ❌ ERROR procesando material: {str(item_error)}")
                    raise  # Re-lanzar para que se capture en el try principal

            # ===== 8. COMMIT =====
            print(f"\n💾 GUARDANDO EN BASE DE DATOS...")

            # Actualizar has_returns si algún material tiene will_return = True
            has_any_returns = any(mat.get('will_return', False) for mat in materials_list)
            if has_any_returns:
                new_req.has_returns = True
                print(f"  - Requisición tiene devoluciones pendientes")

            db.session.commit()

            print(f"✅ REQUISICIÓN GUARDADA EXITOSAMENTE")
            print(f"  - Número: {req_number}")
            print(f"  - Items creados: {items_created}")
            print(f"  - Nuevos materiales: {new_materials_created}")
            print("="*60 + "\n")

            # ===== 10. RESPUESTA =====
            return jsonify({
                'success': True,
                'message': 'Requisición creada exitosamente',
                'request_number': req_number,
                'redirect_url': url_for('requests')
            })

        except Exception as e:
            db.session.rollback()
            error_msg = f'Error al crear requisición: {str(e)}'
            print(f"\n❌ ERROR GENERAL:")
            print(f"  {error_msg}")

            import traceback
            print("\n📋 TRACEBACK:")
            traceback.print_exc()
            print("="*60 + "\n")

            return jsonify({
                'success': False,
                'message': error_msg
            }), 500

    # GET request - mostrar formulario
    return render_template('new_request.html')

# 2. NUEVA RUTA: Obtener departamentos únicos que han hecho requisiciones para un FP
@app.route('/api/stock/departments-by-fp')
@login_required
def get_departments_by_fp():
    """Retorna los departamentos que han hecho requisiciones para un FP específico"""
    fp_code = request.args.get('fp_code')

    if not fp_code:
        return jsonify({'success': False, 'departments': []})

    # Buscar el proyecto
    project = Project.query.filter_by(fp_code=fp_code).first()
    if not project:
        return jsonify({'success': False, 'departments': []})

    # Obtener departamentos únicos de las requisiciones de este proyecto
    departments = db.session.query(Request.department).filter(
        Request.project_id == project.id,
        Request.department.isnot(None),
        Request.department != ''
    ).distinct().all()

    # Convertir a lista simple
    dept_list = [dept[0] for dept in departments if dept[0]]

    return jsonify({
        'success': True,
        'departments': dept_list
    })



# También necesitas asegurarte de que tengas esta ruta para buscar materiales
@app.route('/api/materials/search')
@login_required
def search_materials():
    """Buscar materiales para el autocompletado"""
    query = request.args.get('q', '')
    if len(query) < 2:
        return jsonify({'materials': []})

    materials = Material.query.filter(
        db.or_(
            Material.name.contains(query),
            Material.code.contains(query)
        )
    ).limit(25).all()

    materials_data = []
    for material in materials:
        materials_data.append({
            'id': material.id,
            'code': material.code,
            'name': material.name,
            'unit': material.unit,
            'category': material.category,
            'current_stock': material.current_stock,
            'min_stock': material.min_stock,
            'can_recycle': material.can_recycle,
            'can_reuse': material.can_reuse
        })

    return jsonify({'materials': materials_data})
@app.route('/api/requests/<int:request_id>/details')
@login_required
def get_request_details(request_id):
    """Obtener detalles completos de una requisición"""
    try:
        req = Request.query.get_or_404(request_id)

        # Verificar permisos
        if not (current_user.role == 'admin' or
                current_user.is_leader or
                req.user_id == current_user.id):
            return jsonify({'success': False, 'message': 'Sin permisos para ver esta requisición'})

        # Preparar datos de materiales
        items_data = []
        for item in req.items:
            item_data = {
                'id': item.id,
                'quantity_requested': item.quantity_requested,
                'quantity_delivered': item.quantity_delivered,
                'item_type': item.item_type,
                'will_return': item.will_return,
                'return_expected_date': item.return_expected_date.isoformat() if item.return_expected_date else None,
                'notes': item.notes,
                'is_new_material': item.is_new_material
            }

            if item.is_new_material:
                item_data.update({
                    'code': item.new_material_code,
                    'name': item.new_material_name,
                    'unit': item.new_material_unit,
                    'category': item.new_material_category
                })
            else:
                item_data.update({
                    'code': item.material.code,
                    'name': item.material.name,
                    'unit': item.material.unit,
                    'category': item.material.category,
                    'current_stock': item.material.current_stock
                })

            items_data.append(item_data)

        request_data = {
            'id': req.id,
            'request_number': req.request_number,
            'department': req.department,
            'area': req.area,
            'request_type': req.request_type,
            'is_incident': req.is_incident,
            'incident_id': req.incident_id,
            'client': req.project.client,
            'acquisition_deadline': req.acquisition_deadline.isoformat() if req.acquisition_deadline else None,
            'assembly_start_date': req.assembly_start_date.isoformat() if req.assembly_start_date else None,
            'assembly_end_date': req.assembly_end_date.isoformat() if req.assembly_end_date else None,
            'status': req.status,
            'notes': req.notes,
            'created_at': req.created_at.isoformat(),
            'approved_at': req.approved_at.isoformat() if req.approved_at else None,
            'user': {
                'username': req.user.username,
                'email': req.user.email
            },
            'project': {
                'name': req.project.name,
                'fp_code': req.project.fp_code,
                'delivery_date': req.project.delivery_date.isoformat(),
                'production_start': req.project.production_start.isoformat(),
                'assembly_date': req.project.assembly_date.isoformat()
            },
            'items': items_data
        }

        return jsonify({'success': True, 'request': request_data})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al obtener detalles: {str(e)}'})


@app.route('/api/requests/export')
@login_required
def export_requests():
    """Exportar lista de requisiciones"""
    try:
        # Obtener filtros
        status_filter = request.args.get('status')
        department_filter = request.args.get('department')
        incident_filter = request.args.get('incident_filter')
        search_filter = request.args.get('search')

        # Construir consulta
        query = Request.query

        # Aplicar filtros de permisos
        if not current_user.role == 'admin':
            query = query.filter_by(user_id=current_user.id)

        if status_filter:
            query = query.filter_by(status=status_filter)
        if department_filter:
            query = query.filter_by(department=department_filter)
        if incident_filter == 'incident':
            query = query.filter_by(is_incident=True)
        elif incident_filter == 'normal':
            query = query.filter_by(is_incident=False)
        if search_filter:
            query = query.filter(
                db.or_(
                    Request.request_number.contains(search_filter),
                    Request.project.has(Project.name.contains(search_filter)),
                    Request.project.has(Project.fp_code.contains(search_filter))
                )
            )

        requests = query.order_by(Request.created_at.desc()).all()

        # Preparar datos para exportar
        export_data = []
        for req in requests:
            # Contar materiales
            total_materials = len(req.items)
            new_materials = len([item for item in req.items if item.is_new_material])
            return_materials = len([item for item in req.items if item.will_return])

            export_data.append({
                'No. Requisicion': req.request_number,
                'Solicitante': req.user.username,
                'Email': req.user.email,
                'Departamento': req.department,
                'Area': req.area or '',
                'Proyecto': req.project.name,
                'Codigo FP': req.project.fp_code,
                'Tipo Requisicion': req.request_type,
                'Es Incidencia': 'SÍ' if req.is_incident else 'NO',
                'ID Incidencia': req.incident_id or '',
                'Estado': req.status,
                'Total Materiales': total_materials,
                'Materiales Nuevos': new_materials,
                'Materiales con Retorno': return_materials,
                'Inicio Montaje': req.assembly_start_date.strftime('%d/%m/%Y') if req.assembly_start_date else '',
                'Fin Montaje': req.assembly_end_date.strftime('%d/%m/%Y') if req.assembly_end_date else '',
                'Fecha Creacion': req.created_at.strftime('%d/%m/%Y %H:%M'),
                'Fecha Aprobacion': req.approved_at.strftime('%d/%m/%Y %H:%M') if req.approved_at else '',
                'Fecha Entrega Proyecto': req.project.delivery_date.strftime('%d/%m/%Y'),
                'Notas': req.notes or ''
            })

        # Crear archivo Excel
        df = pd.DataFrame(export_data)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'requisiciones_{timestamp}.xlsx'
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        df.to_excel(filepath, index=False, sheet_name='Requisiciones')

        return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al exportar: {str(e)}'})

@app.route('/api/requests/export-selected')
@login_required
def export_selected_requests():
    """Exportar requisiciones seleccionadas"""
    try:
        ids_param = request.args.get('ids', '')
        if not ids_param:
             return jsonify({'success': False, 'message': 'No se seleccionaron requisiciones'}), 400
        
        ids_list = [int(id) for id in ids_param.split(',') if id.isdigit()]
        
        # Consultar requisiciones seleccionadas
        requests = Request.query.filter(Request.id.in_(ids_list)).order_by(Request.created_at.desc()).all()

        # Preparar datos para exportar
        export_data = []
        for req in requests:
            # Contar materiales
            total_materials = len(req.items)
            new_materials = len([item for item in req.items if item.is_new_material])
            return_materials = len([item for item in req.items if item.will_return])

            row = {
                'No. Requisicion': req.request_number,
                'Solicitante': req.user.username,
                'Email': req.user.email,
                'Departamento': req.department,
                'Area': req.area or '',
                'Proyecto': req.project.name,
                'Cliente': req.project.client or '',
                'Codigo FP': req.project.fp_code,
                'Tipo Requisicion': req.request_type,
                'Es Incidencia': 'SÍ' if req.is_incident else 'NO',
                'ID Incidencia': req.incident_id or '',
                'Estado': req.status,
                'Total Materiales': total_materials,
                'Materiales Nuevos': new_materials,
                'Materiales con Retorno': return_materials,
                'Fecha Solicitud': req.created_at.strftime('%d/%m/%Y %H:%M'),
                'Fecha Limite Adquisicion': req.acquisition_deadline.strftime('%d/%m/%Y') if req.acquisition_deadline else '',
                'Inicio Produccion': req.request_production_start_date.strftime('%d/%m/%Y') if hasattr(req, 'request_production_start_date') and req.request_production_start_date else (req.production_start_date.strftime('%d/%m/%Y') if hasattr(req, 'production_start_date') and req.production_start_date else ''),
                'Inicio Montaje': req.assembly_start_date.strftime('%d/%m/%Y') if req.assembly_start_date else '',
                'Fin Montaje': req.assembly_end_date.strftime('%d/%m/%Y') if req.assembly_end_date else '',
                'Fecha Aprobacion': req.approved_at.strftime('%d/%m/%Y %H:%M') if req.approved_at else '',
                'Fecha Entrega Proyecto': req.project.delivery_date.strftime('%d/%m/%Y'),
                'Notas': req.notes or ''
            }
            export_data.append(row)

        # Crear archivo Excel
        df = pd.DataFrame(export_data)
        
        # Asegurar directorio de uploads
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            os.makedirs(app.config['UPLOAD_FOLDER'])
            
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'requisiciones_seleccion_{timestamp}.xlsx'
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        df.to_excel(filepath, index=False, sheet_name='Requisiciones')

        return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Error al exportar: {str(e)}'})


@app.route('/requests/<int:request_id>/print')
@login_required
def print_request(request_id):
    """Vista de impresión para una requisición"""
    req = Request.query.get_or_404(request_id)
    
    # Verificar permisos (lectura)
    if not (current_user.role == 'admin' or
            current_user.is_leader or
            req.user_id == current_user.id):
        flash('No tiene permisos para ver esta requisición.', 'danger')
        return redirect(url_for('home'))

    return render_template('print_request.html', request=req, now=datetime.now())


@app.route('/api/projects/<fp_code>/export-report')
@login_required
def export_project_report(fp_code):
    """Exportar reporte detallado de un proyecto"""
    try:
        project = Project.query.filter_by(fp_code=fp_code).first_or_404()

        # Obtener todas las requisiciones del proyecto
        project_requests = Request.query.filter_by(project_id=project.id).all()

        # Crear workbook con múltiples hojas
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Hoja 1: Resumen del proyecto
        ws_summary = wb.active
        ws_summary.title = "Resumen Proyecto"

        # Headers con estilo
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Información del proyecto
        ws_summary['A1'] = "REPORTE DE PROYECTO"
        ws_summary['A1'].font = Font(bold=True, size=16)

        ws_summary['A3'] = "Código FP:"
        ws_summary['B3'] = project.fp_code
        ws_summary['A4'] = "Nombre:"
        ws_summary['B4'] = project.name
        ws_summary['A5'] = "Fecha Entrega:"
        ws_summary['B5'] = project.delivery_date.strftime('%d/%m/%Y')
        ws_summary['A6'] = "Inicio Producción:"
        ws_summary['B6'] = project.production_start.strftime('%d/%m/%Y')
        ws_summary['A7'] = "Fecha Montaje:"
        ws_summary['B7'] = project.assembly_date.strftime('%d/%m/%Y')

        # Estadísticas
        total_requests = len(project_requests)
        pending_requests = len([r for r in project_requests if r.status == 'pendiente'])
        completed_requests = len([r for r in project_requests if r.status == 'completada'])
        incident_requests = len([r for r in project_requests if r.is_incident])

        ws_summary['A9'] = "ESTADÍSTICAS:"
        ws_summary['A9'].font = Font(bold=True)
        ws_summary['A10'] = "Total Requisiciones:"
        ws_summary['B10'] = total_requests
        ws_summary['A11'] = "Pendientes:"
        ws_summary['B11'] = pending_requests
        ws_summary['A12'] = "Completadas:"
        ws_summary['B12'] = completed_requests
        ws_summary['A13'] = "Por Incidencia:"
        ws_summary['B13'] = incident_requests

        # Hoja 2: Lista de requisiciones
        ws_requests = wb.create_sheet("Requisiciones")

        req_headers = [
            "No. Requisición", "Solicitante", "Departamento", "Área",
            "Tipo", "Estado", "Es Incidencia", "ID Incidencia",
            "Materiales", "Fecha Creación", "Notas"
        ]

        for col, header in enumerate(req_headers, 1):
            cell = ws_requests.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row, req in enumerate(project_requests, 2):
            ws_requests.cell(row=row, column=1, value=req.request_number)
            ws_requests.cell(row=row, column=2, value=req.user.username)
            ws_requests.cell(row=row, column=3, value=req.department)
            ws_requests.cell(row=row, column=4, value=req.area or '')
            ws_requests.cell(row=row, column=5, value=req.request_type)
            ws_requests.cell(row=row, column=6, value=req.status)
            ws_requests.cell(row=row, column=7, value='SÍ' if req.is_incident else 'NO')
            ws_requests.cell(row=row, column=8, value=req.incident_id or '')
            ws_requests.cell(row=row, column=9, value=len(req.items))
            ws_requests.cell(row=row, column=10, value=req.created_at.strftime('%d/%m/%Y'))
            ws_requests.cell(row=row, column=11, value=req.notes or '')

        # Hoja 3: Consolidado de materiales
        ws_materials = wb.create_sheet("Materiales Consolidado")

        # Consolidar materiales
        materials_consolidated = {}
        for req in project_requests:
            for item in req.items:
                if item.is_new_material:
                    key = item.new_material_code
                    name = item.new_material_name
                    unit = item.new_material_unit
                else:
                    key = item.material.code
                    name = item.material.name
                    unit = item.material.unit

                if key not in materials_consolidated:
                    materials_consolidated[key] = {
                        'name': name,
                        'unit': unit,
                        'total_requested': 0,
                        'total_delivered': 0,
                        'will_return': False,
                        'requests': []
                    }

                materials_consolidated[key]['total_requested'] += item.quantity_requested
                materials_consolidated[key]['total_delivered'] += item.quantity_delivered
                if item.will_return:
                    materials_consolidated[key]['will_return'] = True
                materials_consolidated[key]['requests'].append(req.request_number)

        mat_headers = [
            "Código", "Material", "Unidad", "Total Solicitado",
            "Total Entregado", "Pendiente", "¿Retorna?", "Requisiciones"
        ]

        for col, header in enumerate(mat_headers, 1):
            cell = ws_materials.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row, (code, data) in enumerate(materials_consolidated.items(), 2):
            pending = data['total_requested'] - data['total_delivered']
            ws_materials.cell(row=row, column=1, value=code)
            ws_materials.cell(row=row, column=2, value=data['name'])
            ws_materials.cell(row=row, column=3, value=data['unit'])
            ws_materials.cell(row=row, column=4, value=data['total_requested'])
            ws_materials.cell(row=row, column=5, value=data['total_delivered'])
            ws_materials.cell(row=row, column=6, value=pending)
            ws_materials.cell(row=row, column=7, value='SÍ' if data['will_return'] else 'NO')
            ws_materials.cell(row=row, column=8, value=', '.join(set(data['requests'])))

        # Ajustar ancho de columnas
        for ws in [ws_summary, ws_requests, ws_materials]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Guardar archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'proyecto_{fp_code}_{timestamp}.xlsx'
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        wb.save(filepath)

        return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al exportar reporte: {str(e)}'})

@app.route('/api/projects/<int:project_id>/generate-summary', methods=['POST'])
@login_required
def generate_project_summary(project_id):
    """Generar resumen actualizado de un proyecto"""
    try:
        project = Project.query.get_or_404(project_id)

        # Obtener o crear resumen
        summary = ProjectSummary.query.filter_by(project_id=project_id).first()
        if not summary:
            summary = ProjectSummary(project_id=project_id)
            db.session.add(summary)

        # Calcular estadísticas
        project_requests = Request.query.filter_by(project_id=project_id).all()

        total_requests = len(project_requests)
        total_materials = sum(len(req.items) for req in project_requests)

        # Calcular costo total estimado
        total_cost = 0
        for req in project_requests:
            for item in req.items:
                if item.material and item.material.unit_cost:
                    total_cost += item.quantity_requested * item.material.unit_cost

        # Actualizar resumen
        summary.total_requests = total_requests
        summary.total_materials = total_materials
        summary.total_cost = total_cost
        summary.last_updated = datetime.utcnow()

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Resumen del proyecto actualizado',
            'summary': {
                'total_requests': total_requests,
                'total_materials': total_materials,
                'total_cost': total_cost,
                'last_updated': summary.last_updated.isoformat()
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al generar resumen: {str(e)}'})



@app.route('/requests/<int:id>/edit')
@login_required
def edit_request(id):
    req = Request.query.get_or_404(id)
    materials = Material.query.all()
    # Usar 'requisition' en vez de 'request' para no sobrescribir el request de Flask
    return render_template('edit_request.html', requisition=req, materials=materials)


# ===== ENDPOINTS PARA MANEJO DE REQUISICIONES =====

@app.route('/api/request/<int:request_id>/request-cancellation', methods=['POST'])
@login_required
def request_cancellation(request_id):
    """Solicitar cancelación de una requisición"""
    req = Request.query.get_or_404(request_id)
    
    # Solo el creador puede solicitar cancelación (si no es admin)
    if current_user.role != 'admin' and req.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'No tiene permiso para solicitar cancelación'}), 403
        
    if req.status in ['cancelada', 'completada', 'en_entrega']:
        return jsonify({'success': False, 'message': 'No se puede cancelar una requisición en este estado'}), 400
        
    try:
        req.cancellation_requested = True
        req.cancellation_requested_at = datetime.utcnow()
        req.cancellation_requester_id = current_user.id
        db.session.commit()
        return jsonify({'success': True, 'message': 'Solicitud de cancelación enviada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/request/<int:request_id>/confirm-cancellation', methods=['POST'])
@login_required
def confirm_cancellation(request_id):
    """Confirmar cancelación de una requisición"""
    # Solo admin o almacenista puede confirmar
    if current_user.role not in ['admin', 'almacenista']:
        return jsonify({'success': False, 'message': 'No tiene permiso para confirmar cancelación'}), 403
        
    req = Request.query.get_or_404(request_id)
    
    try:
        req.status = 'cancelada'
        req.cancellation_requested = False # Ya se atendió
        
        # Cancelar también todos los items pendientes
        for item in req.items:
            if item.item_status not in ['abastecido', 'entregado']:
                item.item_status = 'cancelado'
                
        db.session.commit()
        return jsonify({'success': True, 'message': 'Requisición cancelada exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/request/<int:request_id>/update-dates', methods=['POST'])
@login_required
def update_request_dates(request_id):
    """Actualizar fechas de la requisición"""
    req = Request.query.get_or_404(request_id)
    
    # Validar permisos (admin o dueño si está pendiente)
    if current_user.role != 'admin' and (req.user_id != current_user.id or req.status != 'pendiente'):
        return jsonify({'success': False, 'message': 'No tiene permiso para editar fechas'}), 403
        
    try:
        data = request.form
        if 'acquisition_deadline' in data:
            val = data['acquisition_deadline']
            req.acquisition_deadline = datetime.strptime(val, '%Y-%m-%d').date() if val else None
        if 'production_start_date' in data:
            val = data['production_start_date']
            req.production_start_date = datetime.strptime(val, '%Y-%m-%d').date() if val else None
        if 'assembly_start_date' in data:
            val = data['assembly_start_date']
            req.assembly_start_date = datetime.strptime(val, '%Y-%m-%d').date() if val else None
        if 'assembly_end_date' in data:
            val = data['assembly_end_date']
            req.assembly_end_date = datetime.strptime(val, '%Y-%m-%d').date() if val else None
            
        db.session.commit()
        return jsonify({'success': True, 'message': 'Fechas actualizadas correctamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/request/<int:id>/update-notes', methods=['POST'])
@login_required
def api_update_request_notes(id):
    """Actualizar notas de una requisición"""
    try:
        req = Request.query.get_or_404(id)
        req.notes = request.form.get('notes', '').strip()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Notas actualizadas'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/request/<int:id>/request-cancellation', methods=['POST'])
@login_required
def api_request_cancellation(id):
    """Requisitador solicita cancelación de una requisición"""
    try:
        req = Request.query.get_or_404(id)

        # Verificar que el usuario es requisitador
        if current_user.role != 'requisitador':
            return jsonify({'success': False, 'message': 'Solo requisitadores pueden solicitar cancelación'})

        # Verificar que la requisición no está en estados finales
        if req.status in ['en_entrega', 'completada', 'cancelada']:
            return jsonify({'success': False, 'message': f'No se puede cancelar una requisición en estado "{req.status}"'})

        req.cancellation_requested = True
        req.cancellation_requested_by = current_user.id
        req.cancellation_requested_at = datetime.utcnow()
        db.session.commit()

        return jsonify({'success': True, 'message': 'Solicitud de cancelación enviada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/request/<int:id>/confirm-cancellation', methods=['POST'])
@login_required
def api_confirm_cancellation(id):
    """Admin/Almacenista confirma cancelación de una requisición"""
    try:
        req = Request.query.get_or_404(id)

        # Verificar permisos
        if current_user.role not in ['admin', 'almacenista']:
            return jsonify({'success': False, 'message': 'No tiene permisos para confirmar cancelación'})

        # Verificar que hay una solicitud pendiente
        if not req.cancellation_requested:
            return jsonify({'success': False, 'message': 'No hay solicitud de cancelación pendiente'})

        req.status = 'cancelada'
        # Cancelar todos los items
        for item in req.items:
            if item.item_status != 'cancelado':
                item.item_status = 'cancelado'

        db.session.commit()
        return jsonify({'success': True, 'message': 'Requisición cancelada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


def recalculate_request_status(req):
    """Recalcula el estado de una requisición basado en sus items"""
    if req.status == 'cancelada':
        return  # No cambiar si está cancelada

    items = [i for i in req.items if i.item_status != 'cancelado']

    if not items:
        return

    statuses = [i.item_status for i in items]

    # Verificar retornos pendientes
    today = datetime.utcnow().date()
    for item in items:
        if item.will_return and item.return_expected_date:
            if today > item.return_expected_date and not item.actual_return_date:
                item.item_status = 'pendiente_retorno'

    # Recargar estados después de verificar retornos
    statuses = [i.item_status for i in items]

    # Priority 1: pendiente_compra (highest priority per requirements)
    if 'pendiente_compra' in statuses:
        req.status = 'pendiente_compra'
    # Priority 2: pendiente_retorno
    elif 'pendiente_retorno' in statuses:
        req.status = 'pendiente_retorno'
    elif all(s == 'abastecido' for s in statuses):
        # Verificar si es día de entrega
        if req.acquisition_deadline and today >= req.acquisition_deadline:
            req.status = 'en_entrega'
        else:
            req.status = 'abastecido'
    elif any(s == 'pendiente' for s in statuses):
        req.status = 'pendiente'


# ===== ENDPOINTS PARA ITEMS DE REQUISICIÓN =====

@app.route('/api/request-item/<int:id>/update-notes', methods=['POST'])
@login_required
def api_update_item_notes(id):
    """Actualizar notas de un item"""
    try:
        item = RequestItem.query.get_or_404(id)
        item.item_notes = request.form.get('notes', '').strip()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Notas actualizadas'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/request-item/<int:id>/update-quantity', methods=['POST'])
@login_required
def api_update_item_quantity(id):
    """Modificar cantidad de un item (solo reducir)"""
    try:
        item = RequestItem.query.get_or_404(id)

        if item.item_status != 'pendiente':
            return jsonify({'success': False, 'message': 'Solo se puede modificar items pendientes'})

        new_qty = float(request.form.get('quantity', 0))
        if new_qty <= 0:
            return jsonify({'success': False, 'message': 'La cantidad debe ser mayor a 0'})
        if new_qty > item.quantity_requested:
            return jsonify({'success': False, 'message': 'Solo se puede reducir la cantidad, no aumentar'})

        item.quantity_requested = new_qty
        db.session.commit()
        return jsonify({'success': True, 'message': 'Cantidad actualizada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/request-item/<int:id>/analyze-stock', methods=['POST'])
@login_required
def api_analyze_item_stock(id):
    """Analizar item contra stock y definir cantidades"""
    try:
        item = RequestItem.query.get_or_404(id)

        if current_user.role not in ['admin', 'almacenista']:
            return jsonify({'success': False, 'message': 'Sin permisos para esta acción'})

        if item.item_status != 'pendiente':
            return jsonify({'success': False, 'message': 'Item ya fue procesado'})

        quantity_to_purchase = float(request.form.get('quantity_to_purchase', 0))

        # Validar
        requested = item.quantity_requested
        supplied = requested - quantity_to_purchase

        if quantity_to_purchase < 0:
            return jsonify({'success': False, 'message': 'Cantidad a comprar no puede ser negativa'})
        if quantity_to_purchase > requested:
            return jsonify({'success': False, 'message': 'Cantidad a comprar no puede exceder lo solicitado'})

        item.quantity_to_purchase = quantity_to_purchase
        item.quantity_supplied = supplied

        # Determinar estado
        if quantity_to_purchase > 0:
            item.item_status = 'pendiente_compra'
        else:
            item.item_status = 'abastecido'

        # Recalcular estado de requisición
        recalculate_request_status(item.request)

        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Análisis completado: {supplied} a abastecer, {quantity_to_purchase} a comprar',
            'new_status': item.item_status
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/request-item/<int:id>/cancel', methods=['POST'])
@login_required
def api_cancel_item(id):
    """Cancelar un item individual"""
    try:
        item = RequestItem.query.get_or_404(id)

        if item.item_status != 'pendiente':
            return jsonify({'success': False, 'message': 'Solo se pueden cancelar items pendientes'})

        item.item_status = 'cancelado'
        recalculate_request_status(item.request)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Material cancelado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/request-item/<int:id>/mark-supplied', methods=['POST'])
@login_required
def api_mark_item_supplied(id):
    """Marcar item como abastecido (después de que llegue la compra)"""
    try:
        item = RequestItem.query.get_or_404(id)

        if current_user.role not in ['admin', 'almacenista']:
            return jsonify({'success': False, 'message': 'Sin permisos para esta acción'})

        if item.item_status != 'pendiente_compra':
            return jsonify({'success': False, 'message': 'Esta acción solo aplica a items pendientes de compra'})

        item.item_status = 'abastecido'
        item.quantity_supplied = item.quantity_requested
        recalculate_request_status(item.request)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Material marcado como abastecido'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/request-item/<int:id>/details')
@login_required
def api_get_item_details(id):
    """Obtener detalles de un item para modales"""
    try:
        item = RequestItem.query.get_or_404(id)
        material = item.material

        return jsonify({
            'success': True,
            'item': {
                'id': item.id,
                'quantity_requested': item.quantity_requested,
                'quantity_supplied': item.quantity_supplied or 0,
                'quantity_to_purchase': item.quantity_to_purchase or 0,
                'item_status': item.item_status or 'pendiente',
                'item_notes': item.item_notes or '',
                'material_name': material.name if material else item.new_material_name,
                'material_code': material.code if material else item.new_material_code,
                'current_stock': material.current_stock if material else 0,
                'unit': material.unit if material else item.new_material_unit
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/project/<fp_code>')
@login_required
def get_project_info(fp_code):
    project = Project.query.filter_by(fp_code=fp_code).first()
    if project:
        return jsonify({
            'found': True,
            'name': project.name,
            'delivery_date': project.delivery_date.isoformat(),
            'production_start': project.production_start.isoformat(),
            'assembly_date': project.assembly_date.isoformat()
        })
    return jsonify({'found': False})

@app.route('/fabric-rolls')
@login_required
def fabric_rolls():
    rolls = FabricRoll.query.join(Material).all()
    # Obtener solo materiales cuya categoría sea 'Telas'
    materials = Material.query.filter(Material.category.ilike('%tela%')).all()
    return render_template('fabric_rolls.html', rolls=rolls, materials=materials)

@app.route('/stock-movements')
@login_required
def stock_movements():
    page = request.args.get('page', 1, type=int)

    # Filtros
    movement_type = request.args.get('type', '', type=str)
    material_filter = request.args.get('material', '', type=str)
    area_filter = request.args.get('area', '', type=str)
    date_from = request.args.get('date_from', '', type=str)
    date_to = request.args.get('date_to', '', type=str)

    query = StockMovement.query.join(Material)

    # Aplicar filtros
    if movement_type:
        query = query.filter(StockMovement.movement_type == movement_type)

    if material_filter:
        query = query.filter(
            (Material.name.contains(material_filter)) |
            (Material.code.contains(material_filter))
        )

    if area_filter:
        query = query.filter(StockMovement.area.contains(area_filter))

    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(StockMovement.created_at >= date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            # Agregar un día para incluir todo el día
            date_to_obj = datetime.combine(date_to_obj, datetime.max.time())
            query = query.filter(StockMovement.created_at <= date_to_obj)
        except ValueError:
            pass

    movements = query.order_by(StockMovement.created_at.desc()).paginate(
        page=page, per_page=30, error_out=False)

    # Buscar información de cliente localmente para los FP en pantalla
    fp_codes = set()
    for m in movements.items:
        if m.fp_code:
            fp_codes.add(m.fp_code)
    
    projects_dict = {}
    if fp_codes:
        projs = Project.query.filter(Project.fp_code.in_(fp_codes)).all()
        for p in projs:
            projects_dict[p.fp_code] = p.name or ''

    return render_template('stock_movements.html', movements=movements, projects_dict=projects_dict)


@app.route('/api/requests/approval', methods=['POST'])
@login_required
def approve_reject_request():
    """Aprobar o rechazar una requisición"""
    try:
        # Verificar permisos
        if not (current_user.role == 'admin' or current_user.is_leader):
            return jsonify({'success': False, 'message': 'Sin permisos para aprobar/rechazar requisiciones'})

        request_id = request.form.get('request_id')
        action = request.form.get('action')  # 'approve' o 'reject'
        reason = request.form.get('reason', '')
        comments = request.form.get('comments', '')

        req = Request.query.get_or_404(request_id)

        if req.status != 'pendiente':
            return jsonify({'success': False, 'message': 'Esta requisición ya fue procesada'})

        if action == 'approve':
            req.status = 'aprobada'
            req.approved_at = datetime.utcnow()
            req.approved_by = current_user.id

            # Si es requisición por incidencia, marcar como prioritaria
            if req.is_incident:
                req.status = 'en_proceso'  # Pasar directamente a proceso por ser incidencia

            message = 'Requisición aprobada exitosamente'

        elif action == 'reject':
            if not reason:
                return jsonify({'success': False, 'message': 'El motivo del rechazo es requerido'})

            req.status = 'cancelada'
            req.notes = f"{req.notes}\n\nRECHAZADA: {reason}" if req.notes else f"RECHAZADA: {reason}"

            message = 'Requisición rechazada'

        else:
            return jsonify({'success': False, 'message': 'Acción no válida'})

        # Agregar comentarios si existen
        if comments:
            req.notes = f"{req.notes}\n\nComentarios del aprobador: {comments}" if req.notes else f"Comentarios del aprobador: {comments}"

        db.session.commit()

        return jsonify({'success': True, 'message': message})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al procesar solicitud: {str(e)}'})



@app.route('/reports')
@login_required
def reports():
    if current_user.role == 'requisitador':
        flash('Sección no disponible para requisitadores.', 'warning')
        return redirect(url_for('dashboard'))

    if current_user.role == 'requisitador':
        flash('Sección no disponible para requisitadores.', 'warning')
        return redirect(url_for('dashboard'))

    # Materiales con stock bajo
    low_stock = Material.query.filter(Material.current_stock <= Material.min_stock).all()

    # Materiales sin movimiento en 6 meses
    six_months_ago = datetime.utcnow() - timedelta(days=180)
    no_movement = Material.query.filter(
        (Material.last_movement < six_months_ago) | (Material.last_movement.is_(None))
    ).all()

    # Estadísticas de devoluciones
    returns_stats = db.session.query(
        StockMovement.condition_on_return,
        db.func.count(StockMovement.id)
    ).filter(
        StockMovement.movement_type == 'retorno'
    ).group_by(StockMovement.condition_on_return).all()

    return render_template('reports.html',
                         low_stock=low_stock,
                         no_movement=no_movement,
                         returns_stats=returns_stats)
@app.route('/leader-dashboard')
@login_required
def leader_dashboard():
    if not (current_user.is_leader or current_user.role == 'admin'):
        flash('Acceso denegado')
        return redirect(url_for('dashboard'))

    dept_param = (request.args.get('dept') or '').strip()
    if current_user.role == 'admin':
        scope_all = (dept_param.upper() == 'ALL' or dept_param == '')
        department = None if scope_all else dept_param
    else:
        department = current_user.department
        scope_all = False

    today = datetime.utcnow().date()
    start_month = today.replace(day=1)
    next_month = (start_month + timedelta(days=32)).replace(day=1)

    q = Request.query.filter(Request.created_at >= start_month,
                             Request.created_at < next_month)
    if not scope_all:
        q = q.filter(Request.department == department)

    total_requests_month   = q.count()
    pending_approval_count = q.filter(Request.status == 'pendiente').count()
    completed_count        = q.filter(Request.status == 'completada').count()
    efficiency = round((completed_count / total_requests_month) * 100, 1) if total_requests_month else 0.0

    dept_requests = q.order_by(Request.created_at.desc()).all()
    pending_requests = q.filter(Request.status == 'pendiente').order_by(Request.created_at.asc()).limit(50).all()

    if scope_all:
        team_members = User.query.count()
        active_members = db.session.query(User.id).join(Request, Request.user_id == User.id)\
            .filter(Request.created_at >= start_month, Request.created_at < next_month)\
            .distinct().count()
    else:
        team_members = User.query.filter(User.department == department).count()
        active_members = db.session.query(User.id).join(Request, Request.user_id == User.id)\
            .filter(User.department == department,
                    Request.created_at >= start_month,
                    Request.created_at < next_month)\
            .distinct().count()
    productivity = round((active_members / team_members) * 100) if team_members else 0

    top_users_q = db.session.query(
        User.username.label('username'),
        func.count(Request.id).label('req_count')
    ).join(Request, Request.user_id == User.id)\
     .filter(Request.created_at >= start_month, Request.created_at < next_month)
    if not scope_all:
        top_users_q = top_users_q.filter(Request.department == department)
    top_users = top_users_q.group_by(User.id).order_by(func.count(Request.id).desc()).limit(5).all()

    top_materials_q = db.session.query(
        Material.code.label('code'),
        Material.name.label('name'),
        func.count(RequestItem.id).label('reqs')
    ).join(RequestItem, RequestItem.material_id == Material.id)\
     .join(Request, Request.id == RequestItem.request_id)\
     .filter(Request.created_at >= start_month, Request.created_at < next_month)
    if not scope_all:
        top_materials_q = top_materials_q.filter(Request.department == department)
    top_materials = top_materials_q.group_by(Material.id).order_by(func.count(RequestItem.id).desc()).limit(5).all()

    projects_q = Project.query
    if not scope_all:
        projects_q = projects_q.join(Request, Request.project_id == Project.id)\
            .filter(Request.created_at >= start_month,
                    Request.created_at < next_month,
                    Request.department == department)
    projects_active = projects_q.filter(Project.status == 'activo').order_by(Project.delivery_date.asc()).limit(5).all()

    # --- progreso de proyecto para el template (evita min/max en Jinja) ---
    project_progress = {}
    for p in projects_active:
        if p.delivery_date:
            days_to_delivery = (p.delivery_date - today).days   # >0 si falta; <0 si ya pasó
            raw = ((-days_to_delivery) / 90.0) * 100.0          # misma fórmula de tu template
            pct = int(max(0, min(100, raw)))                    # clamp 0..100
        else:
            pct = 0
        project_progress[p.id] = pct

    # Actividad reciente (combinada, real)
    recent_requests_q = q.order_by(Request.created_at.desc()).limit(10).all()
    recent_moves_q = []
    if hasattr(StockMovement, 'created_at'):
        recent_moves_q = StockMovement.query.order_by(StockMovement.created_at.desc()).limit(10).all()

    def _req_entry(r):
        return {'type': 'request', 'title': f"Requisición {r.status.capitalize()}",
                'subtitle': f"{r.request_number} - {(r.user.username if r.user else 'N/A')}",
                'when': r.created_at}
    def _mv_entry(m):
        try:
            when = datetime.combine(m.fecha, m.hora) if (m.fecha and m.hora) else datetime.utcnow()
        except Exception:
            when = datetime.utcnow()
        return {'type': 'movement', 'title': f"Movimiento: {m.movement_type.capitalize()}",
                'subtitle': f"{m.idm} - {m.personal or 'N/A'}", 'when': when}

    activity_recent_raw = [_req_entry(r) for r in recent_requests_q] + [_mv_entry(m) for m in recent_moves_q]
    activity_recent = sorted(activity_recent_raw, key=lambda x: x['when'], reverse=True)[:10]

    # Aprobadas vs rechazadas por semana (portable, sin floor())
    days_in_month = (next_month - start_month).days
    weeks_count = (days_in_month + 6) // 7
    week_labels = [f"Semana {i}" for i in range(1, weeks_count + 1)]
    approved_per_week = [0] * weeks_count
    rejected_per_week = [0] * weeks_count
    for status, dt in q.with_entities(Request.status, Request.created_at).all():
        idx = (dt.day - 1) // 7
        if 0 <= idx < weeks_count:
            if status == 'completada':
                approved_per_week[idx] += 1
            elif status == 'rechazada':
                rejected_per_week[idx] += 1

    status_colors = {
        'pendiente':  'warning',
        'aprobada':   'primary',
        'en_proceso': 'info',
        'completada': 'success',
        'rechazada':  'danger',
        'cancelada':  'secondary'
    }

    period_label = start_month.strftime('%B %Y')
    now = datetime.utcnow()

    return render_template(
        'leader_dashboard.html',
        total_requests_month=total_requests_month,
        pending_approval_count=pending_approval_count,
        completed_count=completed_count,
        efficiency=efficiency,
        dept_requests=dept_requests,
        pending_requests=pending_requests,
        team_members=team_members,
        active_members=active_members,
        productivity=productivity,
        top_users=top_users,
        top_materials=top_materials,
        projects_active=projects_active,
        project_progress=project_progress,     # <--- NUEVO
        activity_recent=activity_recent,
        week_labels=week_labels,
        approved_per_week=approved_per_week,
        rejected_per_week=rejected_per_week,
        period_label=period_label,
        scope_all=scope_all,
        department=(department or 'TODOS'),
        now=now,
        status_colors=status_colors
    )
# ============= NUEVAS RUTAS API =============

# Ruta para obtener lista de materiales (API)
@app.route('/api/materials/list')
@login_required
def get_materials_list():
    # Obtener materiales que no sean rollos de tela y cuya categoría no incluya 'tela'
    materials = Material.query.filter(
        Material.is_fabric_roll == False,
        ~Material.category.ilike('%tela%')
    ).all()
    materials_data = []

    for material in materials:
        materials_data.append({
            'id': material.id,
            'code': material.code,
            'name': material.name,
            'current_stock': material.current_stock,
            'unit': material.unit,
            'category': material.category
        })

    return jsonify({'materials': materials_data})

@app.route('/api/empleados/list')
@login_required
def get_empleados_activos_list():
    try:
        empleados = remote_db.get_empleados_activos()
        return jsonify({'success': True, 'empleados': empleados})
    except Exception as e:
        app.logger.error(f"Error al obtener empleados: {e}")
        return jsonify({'success': False, 'message': 'No se pudieron cargar los empleados'}), 500

# Ruta para obtener movimientos pendientes de retorno
@app.route('/api/stock/movements/pending-returns')
@login_required
def get_pending_returns():
    pending_movements = StockMovement.query.filter_by(
        movement_type='salida',
        returned=False
    ).join(Material).all()

    movements_data = []
    for movement in pending_movements:
        movements_data.append({
            'id': movement.id,
            'material_code': movement.material.code,
            'material_name': movement.material.name,
            'quantity': movement.quantity,
            'unit': movement.material.unit,
            'date': movement.created_at.strftime('%d/%m/%Y'),
            'user': movement.user.username,
            'area': movement.area or 'N/A'
        })

    return jsonify({'movements': movements_data})

# Ruta para importar movimientos desde Excel
@app.route('/api/stock/movements/import', methods=['POST'])
@login_required
def import_movements():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No se seleccionó archivo'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No se seleccionó archivo'})

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': 'Tipo de archivo no permitido'})

    try:
        # Guardar archivo temporalmente
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # Leer archivo Excel/CSV
        if filename.endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)

        # Validar columnas requeridas
        required_columns = ['IDM', 'Material', 'Movimiento', 'Cantidad', 'Rollos', 'FP', 'Fecha', 'Hora', 'Personal', 'Area']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            os.remove(filepath)  # Limpiar archivo temporal
            return jsonify({
                'success': False,
                'message': f'Columnas faltantes: {", ".join(missing_columns)}'
            })

        # Procesar datos
        imported_count = 0
        errors = []

        for index, row in df.iterrows():
            try:
                # Validar fila
                row_errors = validate_excel_row(row, index + 2)
                if row_errors:
                    errors.extend([f"Fila {index + 2}: {error}" for error in row_errors])
                    continue

                # Buscar material por código
                material = Material.query.filter_by(code=str(row['Material']).strip()).first()
                if not material:
                    errors.append(f"Fila {index + 2}: Material '{row['Material']}' no encontrado")
                    continue

                # Validar tipo de movimiento
                movement_type = str(row['Movimiento']).lower().strip()

                # Procesar fecha y hora
                try:
                    if pd.isna(row['Fecha']):
                        fecha = datetime.now().date()
                    else:
                        fecha = pd.to_datetime(row['Fecha']).date()

                    if pd.isna(row['Hora']):
                        hora = datetime.now().time()
                    else:
                        if isinstance(row['Hora'], str):
                            hora = datetime.strptime(row['Hora'], '%H:%M').time()
                        else:
                            hora = row['Hora']
                except Exception as e:
                    fecha = datetime.now().date()
                    hora = datetime.now().time()
                    errors.append(f"Fila {index + 2}: Error en fecha/hora, usando valores actuales")

                # Crear movimiento
                movement = StockMovement(
                    idm=str(row['IDM']).strip() if not pd.isna(row['IDM']) else None,
                    material_id=material.id,
                    movement_type=movement_type,
                    quantity=float(row['Cantidad']),
                    rollos=int(row['Rollos']) if not pd.isna(row['Rollos']) else 0,
                    fp_code=str(row['FP']).strip() if not pd.isna(row['FP']) else None,
                    fecha=fecha,
                    hora=hora,
                    personal=str(row['Personal']).strip() if not pd.isna(row['Personal']) else None,
                    area=str(row['Area']).strip() if not pd.isna(row['Area']) else None,
                    user_id=current_user.id,
                    notes=f"Importado desde Excel: {filename}"
                )

                db.session.add(movement)

                # Actualizar stock del material
                if movement_type == 'entrada':
                    material.current_stock += movement.quantity
                elif movement_type == 'salida':
                    material.current_stock -= movement.quantity
                # Para retorno, no se actualiza automáticamente el stock

                material.last_movement = datetime.utcnow()
                imported_count += 1

            except Exception as e:
                errors.append(f"Fila {index + 2}: Error al procesar - {str(e)}")
                continue

        # Guardar cambios
        db.session.commit()

        # Limpiar archivo temporal
        os.remove(filepath)

        return jsonify({
            'success': True,
            'message': f'Importación completada: {imported_count} movimientos procesados',
            'imported_count': imported_count,
            'errors': errors[:10]  # Limitar a 10 errores para no saturar la respuesta
        })

    except Exception as e:
        if 'filepath' in locals() and os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'success': False, 'message': f'Error al procesar archivo: {str(e)}'})

# Ruta para descargar plantilla de Excel
@app.route('/api/stock/movements/template')
@login_required
def download_movements_template():
    # Crear DataFrame con las columnas requeridas
    template_data = {
        'IDM': ['IDM001', 'IDM002'],
        'Material': ['MAT-001', 'MAT-002'],
        'Movimiento': ['entrada', 'salida'],
        'Cantidad': [100, 50],
        'Rollos': [1, 2],
        'FP': ['FP-2025-001', 'FP-2025-002'],
        'Fecha': ['2025-06-12', '2025-06-12'],
        'Hora': ['10:30', '14:15'],
        'Personal': ['Juan Pérez', 'María García'],
        'Area': ['Producción', 'Almacén']
    }

    df = pd.DataFrame(template_data)

    # Crear archivo temporal
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'plantilla_movimientos_{timestamp}.xlsx'
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    # Guardar Excel con formato
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Movimientos', index=False)

        # Obtener la hoja de trabajo
        worksheet = writer.sheets['Movimientos']

        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    return send_file(filepath, as_attachment=True, download_name=filename)


# Agregar estas rutas a tu app.py después de la ruta add_material

@app.route('/materials/<int:material_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_material(material_id):
    if current_user.role == 'requisitador':
        flash('No tienes permisos para editar materiales.', 'danger')
        return redirect(url_for('materials'))

    """Editar un material existente"""
    material = Material.query.get_or_404(material_id)

    if request.method == 'POST':
        try:
            # Verificar si es una acción de eliminación
            if request.form.get('action') == 'delete':
                # Solo admins pueden eliminar
                if current_user.role != 'admin':
                    return jsonify({
                        'success': False,
                        'message': 'Solo los administradores pueden eliminar materiales'
                    })
                # Solo si stock es 0
                if material.current_stock > 0:
                    return jsonify({
                        'success': False,
                        'message': 'No se puede eliminar un material con stock. El stock actual debe ser 0.'
                    })
                return delete_material(material_id)

            # Validar stocks (solo si ambos son mayores que 0)
            min_stock = float(request.form.get('min_stock', 0) or 0)
            max_stock = float(request.form.get('max_stock', 0) or 0)
            if min_stock > 0 and max_stock > 0 and min_stock >= max_stock:
                return jsonify({
                    'success': False,
                    'message': 'El stock máximo debe ser mayor al stock mínimo'
                })

            # Actualizar campos básicos
            material.name = request.form['name']
            material.description = request.form.get('description', '')

            # Manejar categoría
            category = request.form.get('category', '')
            material.category = category

            # Determinar si es tela basado en categoría
            is_fabric = category.lower() in ['tela', 'telas', 'fabric', 'fabrics'] if category else False
            material.is_fabric_roll = is_fabric

            # Manejar unidad
            unit = request.form.get('unit', '')
            material.unit = unit

            # Ancho de tela
            if is_fabric and request.form.get('fabric_width'):
                material.fabric_width = float(request.form.get('fabric_width') or 0)

            # Actualizar stocks y costos
            material.min_stock = min_stock
            material.max_stock = max_stock

            # Solo admins pueden cambiar costos
            if current_user.role == 'admin':
                material.unit_cost = float(request.form.get('unit_cost') or 0)

            # Actualizar opciones de reciclaje (mutuamente exclusivas)
            material.can_recycle = bool(request.form.get('can_recycle'))
            material.can_reuse = bool(request.form.get('can_reuse'))
            material.is_recycled = bool(request.form.get('is_recycled'))

            # Material origen si es reciclado
            if material.is_recycled and request.form.get('recycled_from_id'):
                material.recycled_from_id = int(request.form.get('recycled_from_id'))
            else:
                material.recycled_from_id = None

            material.is_consumible = bool(request.form.get('is_consumible'))

            # Actualizar timestamp de modificación
            material.updated_at = datetime.utcnow()

            db.session.commit()

            # Registrar el cambio en el log (opcional)
            log_material_change(material_id, current_user.id, 'Material actualizado')

            return jsonify({
                'success': True,
                'message': 'Material actualizado exitosamente',
                'redirect_url': url_for('materials')
            })

        except ValueError as e:
            return jsonify({
                'success': False,
                'message': f'Error en los datos numéricos: {str(e)}'
            })
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error actualizando material {material_id}: {e}")
            return jsonify({
                'success': False,
                'message': f'Error al actualizar material: {str(e)}'
            })

    # GET request - mostrar formulario de edición

    # Obtener categorías y unidades remotas
    remote_categories = get_remote_categories_for_select()
    remote_units = get_remote_units_for_select()

    # Obtener categorías locales como fallback
    local_categories = db.session.query(Material.category).distinct().filter(
        Material.category.isnot(None)
    ).all()
    local_categories = [cat[0] for cat in local_categories]

    # Obtener materiales reciclables para dropdown "Es Reciclado"
    recyclable_materials = Material.query.filter(
        Material.can_recycle == True,
        Material.id != material_id  # Excluir el material actual
    ).order_by(Material.name).all()

    return render_template('edit_material.html',
                         material=material,
                         remote_categories=remote_categories,
                         remote_units=remote_units,
                         local_categories=local_categories,
                         recyclable_materials=recyclable_materials)

@app.route('/material/delete/<int:material_id>', methods=['POST'])
@login_required
def delete_material(material_id):
    if current_user.role == 'requisitador':
        return jsonify({'success': False, 'error': 'No tienes permisos para eliminar materiales.'}), 403

    """Eliminar un material"""
    try:
        material = Material.query.get_or_404(material_id)

        # Verificar si el material tiene movimientos de stock
        movements_count = StockMovement.query.filter_by(material_id=material_id).count()

        # Verificar si está en requisiciones
        requests_count = RequestItem.query.filter_by(material_id=material_id).count()

        # Verificar si tiene stock actual
        has_stock = material.current_stock > 0

        # Permitir eliminación solo si no tiene dependencias críticas
        if movements_count > 0 or requests_count > 0 or has_stock:
            warning_messages = []
            if has_stock:
                warning_messages.append(f"Stock actual: {material.current_stock} {material.unit}")
            if movements_count > 0:
                warning_messages.append(f"{movements_count} movimientos de stock")
            if requests_count > 0:
                warning_messages.append(f"{requests_count} requisiciones asociadas")

            return jsonify({
                'success': False,
                'message': f'No se puede eliminar el material. Tiene: {", ".join(warning_messages)}',
                'can_delete': False
            })

        # Si no hay dependencias, proceder con la eliminación
        material_name = material.name
        material_code = material.code

        # Eliminar rollos de tela asociados si existen
        if material.is_fabric_roll:
            FabricRoll.query.filter_by(material_id=material_id).delete()

        # Eliminar el material
        db.session.delete(material)
        db.session.commit()

        # Registrar eliminación en el log
        log_material_change(material_id, current_user.id, f'Material eliminado: {material_name} ({material_code})')

        return jsonify({
            'success': True,
            'message': f'Material {material_name} eliminado exitosamente',
            'redirect_url': url_for('materials')
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error eliminando material {material_id}: {e}")
        return jsonify({
            'success': False,
            'message': f'Error al eliminar material: {str(e)}'
        })

@app.route('/materials/<int:material_id>/add-stock')
@login_required
def add_stock_page(material_id):
    if current_user.role == 'requisitador':
        flash('No tienes permisos para agregar stock.', 'danger')
        return redirect(url_for('materials'))

    """Página para agregar stock a un material"""
    material = Material.query.get_or_404(material_id)
    return render_template('add_stock.html', material=material)


@app.route('/materials/<int:material_id>/add-stock', methods=['POST'])
@login_required
def add_stock_api(material_id):
    if current_user.role == 'requisitador':
        return jsonify({'success': False, 'message': 'No tienes permisos para agregar stock.'}), 403

    """Procesar entrada de stock a un material"""
    try:
        material = Material.query.get_or_404(material_id)

        quantity = float(request.form.get('quantity', 0))
        movement_type = request.form.get('movement_type', 'entrada')
        fp_code = request.form.get('fp_code', '').strip()
        notes = request.form.get('notes', '').strip()
        proveedor = request.form.get('proveedor', '').strip()
        documento = request.form.get('documento', '').strip()

        if quantity <= 0:
            flash('La cantidad debe ser mayor a 0', 'error')
            return redirect(url_for('add_stock_page', material_id=material_id))

        # Construir notas completas
        full_notes = []
        if proveedor:
            full_notes.append(f"Proveedor: {proveedor}")
        if documento:
            full_notes.append(f"Doc: {documento}")
        if notes:
            full_notes.append(notes)

        # Crear movimiento de stock
        movement = StockMovement(
            material_id=material_id,
            movement_type=movement_type,
            quantity=quantity,
            fp_code=fp_code if fp_code else None,
            user_id=current_user.id,
            notes=' | '.join(full_notes) if full_notes else None,
            created_at=datetime.utcnow()
        )

        # Actualizar stock del material
        material.current_stock += quantity
        material.last_movement = datetime.utcnow()

        db.session.add(movement)
        db.session.commit()

        flash(f'Entrada registrada: +{quantity} {material.unit} de {material.name}', 'success')
        return redirect(url_for('materials'))

    except Exception as e:
        db.session.rollback()
        flash(f'Error al registrar entrada: {str(e)}', 'error')
        return redirect(url_for('add_stock_page', material_id=material_id))

@app.route('/api/materials/<int:material_id>/details')
@login_required
def get_material_details(material_id):
    """API para obtener detalles de un material - VERSIÓN SEGURA SIN UPDATED_AT"""
    try:
        material = Material.query.get_or_404(material_id)

        # Calcular estadísticas
        total_movements = StockMovement.query.filter_by(material_id=material_id).count()
        last_movement = StockMovement.query.filter_by(material_id=material_id)\
                                           .order_by(StockMovement.created_at.desc()).first()

        # Movimientos de los últimos 30 días
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        recent_entries = StockMovement.query.filter(
            StockMovement.material_id == material_id,
            StockMovement.movement_type == 'entrada',
            StockMovement.created_at >= thirty_days_ago
        ).count()

        recent_exits = StockMovement.query.filter(
            StockMovement.material_id == material_id,
            StockMovement.movement_type == 'salida',
            StockMovement.created_at >= thirty_days_ago
        ).count()

        # Calcular valor total
        total_value = (material.current_stock * (material.unit_cost or 0))

        # Estado del stock - CONSIDERAR MIN/MAX = 0
        if material.min_stock > 0 and material.current_stock <= material.min_stock:
            stock_status = {'status': 'critical', 'label': 'Crítico', 'class': 'danger'}
        elif material.max_stock > 0 and material.current_stock >= material.max_stock:
            stock_status = {'status': 'high', 'label': 'Alto', 'class': 'warning'}
        else:
            stock_status = {'status': 'normal', 'label': 'Normal', 'class': 'success'}

        # ✅ Preparar datos CON MANEJO SEGURO de atributos que pueden no existir
        material_data = {
            'id': material.id,
            'code': material.code,
            'name': material.name,
            'description': getattr(material, 'description', None) or '',
            'category': getattr(material, 'category', None) or '',
            'unit': material.unit,
            'current_stock': material.current_stock,
            'min_stock': material.min_stock,
            'max_stock': material.max_stock,
            'unit_cost': material.unit_cost or 0,
            'total_value': round(total_value, 2),
            'is_fabric_roll': getattr(material, 'is_fabric_roll', False),
            'can_recycle': getattr(material, 'can_recycle', False),
            'can_reuse': getattr(material, 'can_reuse', False),
            'is_consumible': getattr(material, 'is_consumible', False),
            'stock_status': stock_status,
            'last_movement': last_movement.created_at.isoformat() if last_movement else None,
            'total_movements': total_movements,
            'recent_entries': recent_entries,
            'recent_exits': recent_exits,
            'created_at': material.created_at.isoformat() if hasattr(material, 'created_at') and material.created_at else None,
            # ✅ NO incluir updated_at si no existe
        }

        return jsonify({
            'success': True,
            'material': material_data
        })

    except Exception as e:
        app.logger.error(f"Error en get_material_details: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': f'Error al obtener detalles: {str(e)}'
        }), 500

def log_material_change(material_id, user_id, action_description):
    """Registrar cambios en materiales para auditoría"""
    try:
        # Crear un movimiento especial para auditoría
        log_entry = StockMovement(
            material_id=material_id,
            movement_type='sistema',
            quantity=0,
            reference_type='auditoria',
            user_id=user_id,
            notes=action_description
        )
        db.session.add(log_entry)
        db.session.commit()
    except Exception as e:
        app.logger.error(f"Error registrando cambio de material: {e}")
        # No fallar si no se puede registrar el log

@app.route('/api/materials/<int:material_id>/history')
@login_required
def get_material_history(material_id):
    """Obtener historial de cambios de un material"""
    try:
        material = Material.query.get_or_404(material_id)

        # Obtener movimientos de stock
        movements = StockMovement.query.filter_by(material_id=material_id).order_by(
            StockMovement.created_at.desc()
        ).limit(50).all()

        history_data = []
        for movement in movements:
            history_data.append({
                'id': movement.id,
                'movement_type': movement.movement_type,
                'quantity': movement.quantity,
                'date': movement.created_at.strftime('%d/%m/%Y %H:%M'),
                'user': movement.user.username,
                'notes': movement.notes or '',
                'reference_type': movement.reference_type,
                'idm': movement.idm,
                'fp_code': movement.fp_code
            })

        return jsonify({
            'success': True,
            'material': {
                'name': material.name,
                'code': material.code
            },
            'history': history_data
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al obtener historial: {str(e)}'
        })

@app.route('/api/materials/<int:material_id>/validate-delete')
@login_required
def validate_material_deletion(material_id):
    """Validar si un material puede ser eliminado"""
    try:
        material = Material.query.get_or_404(material_id)

        # Verificar dependencias
        movements_count = StockMovement.query.filter_by(material_id=material_id).count()
        requests_count = RequestItem.query.filter_by(material_id=material_id).count()
        has_stock = material.current_stock > 0

        warnings = []
        can_delete = True

        if has_stock:
            warnings.append({
                'type': 'stock',
                'message': f'Tiene stock actual de {material.current_stock} {material.unit}',
                'severity': 'high'
            })
            can_delete = False

        if movements_count > 0:
            warnings.append({
                'type': 'movements',
                'message': f'Tiene {movements_count} movimientos de stock registrados',
                'severity': 'medium'
            })

        if requests_count > 0:
            warnings.append({
                'type': 'requests',
                'message': f'Está asociado a {requests_count} requisiciones',
                'severity': 'medium'
            })

        # Si solo tiene movimientos o requisiciones pero no stock, permitir eliminación con advertencia
        if not has_stock and (movements_count > 0 or requests_count > 0):
            can_delete = True

        return jsonify({
            'success': True,
            'can_delete': can_delete,
            'warnings': warnings,
            'material': {
                'name': material.name,
                'code': material.code
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al validar eliminación: {str(e)}'
        })

# Ruta para exportar movimientos actuales
@app.route('/api/stock/movements/export')
@login_required
def export_movements():
    # Obtener parámetros de filtro
    start_date = request.args.get('date_from')
    end_date = request.args.get('date_to')
    movement_type = request.args.get('type')
    material_filter = request.args.get('material')
    area_filter = request.args.get('area')

    # Construir consulta
    query = StockMovement.query.join(Material)

    if start_date:
        query = query.filter(StockMovement.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(StockMovement.created_at <= datetime.strptime(end_date, '%Y-%m-%d'))
    if movement_type:
        query = query.filter(StockMovement.movement_type == movement_type)
    if material_filter:
        query = query.filter(Material.name.contains(material_filter))
    if area_filter:
        query = query.filter(StockMovement.area.contains(area_filter))

    movements = query.order_by(StockMovement.created_at.desc()).all()

    # Obtener nombres de clientes localmente
    fp_codes = set()
    for m in movements:
        if m.fp_code:
            fp_codes.add(m.fp_code)
            
    projects_dict = {}
    if fp_codes:
        projs = Project.query.filter(Project.fp_code.in_(fp_codes)).all()
        for p in projs:
            projects_dict[p.fp_code] = p.name or ''

    # Preparar datos para exportar
    export_data = []
    for movement in movements:
        cliente = projects_dict.get(movement.fp_code, '') if movement.fp_code else ''
        export_data.append({
            'ID': movement.id,
            'IDM': movement.idm or '',
            'Material': movement.material.code,
            'Movimiento': movement.movement_type,
            'Cantidad': movement.quantity,
            'Rollos': movement.rollos or 0,
            'FP': movement.fp_code or '',
            'Cliente': cliente,
            'Fecha': movement.fecha.strftime('%Y-%m-%d') if movement.fecha else movement.created_at.strftime('%Y-%m-%d'),
            'Hora': movement.hora.strftime('%H:%M') if movement.hora else movement.created_at.strftime('%H:%M'),
            'Entregado a': movement.personal or movement.user.username,
            'Area': movement.area or '',
            'Fecha Creacion': movement.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'Fecha Modificacion': movement.updated_at.strftime('%Y-%m-%d %H:%M:%S') if movement.updated_at else ''
        })

    # Crear archivo Excel
    df = pd.DataFrame(export_data)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'movimientos_stock_{timestamp}.xlsx'
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    df.to_excel(filepath, index=False, sheet_name='Movimientos')

    return send_file(filepath, as_attachment=True, download_name=filename)


# Ruta para registrar entrada rápida
@app.route('/api/stock/entry', methods=['POST'])
@login_required
@permission_required('register_movements')
def register_entry():
    if current_user.role == 'requisitador':
        return jsonify({'success': False, 'message': 'No tienes permisos para registrar entradas.'}), 403

    try:
        data = request.form
        material_id = data.get('material_id')
        quantity = float(data.get('quantity'))

        material = Material.query.get(material_id)
        if not material:
            return jsonify({'success': False, 'message': 'Material no encontrado.'})

        # Generar IDM único
        last_id = StockMovement.query.count()
        idm = f"ENT-{datetime.utcnow().strftime('%y%m%d')}-{last_id + 1:04d}"

        movement = StockMovement(
            idm=idm,
            material_id=material.id,
            movement_type='entrada',
            quantity=quantity,
            rollos=int(data.get('rollos', 0)),
            area=data.get('area'),
            unit_cost=material.unit_cost, # Costo automático del catálogo
            fecha=datetime.utcnow().date(), # Fecha automática
            hora=datetime.utcnow().time(), # Hora automática
            personal=current_user.username, # Personal loggeado
            user_id=current_user.id,
            notes=data.get('notes'),
            reference_type='manual'
        )
        db.session.add(movement)

        material.current_stock += quantity
        material.last_movement = datetime.utcnow()
        db.session.commit()

        return jsonify({'success': True, 'message': 'Entrada registrada exitosamente.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/projects/list')
@login_required
def get_projects_list():
    projects = Project.query.order_by(Project.fp_code).all()
    project_list = [{'id': p.id, 'fp_code': p.fp_code, 'name': p.name} for p in projects]
    return jsonify({'projects': project_list})

@app.route('/api/stock/requisitioned-materials')
@login_required
def get_requisitioned_materials():
    """
    Obtiene materiales requisitados para un FP y departamento específico.
    Ahora usa 'department' en lugar de 'area'
    """
    fp_code = request.args.get('fp_code')
    department = request.args.get('department')  # ✅ Cambiado de 'area' a 'department'

    if not fp_code or not department:
        return jsonify({'success': False, 'materials': []})

    project = Project.query.filter_by(fp_code=fp_code).first()
    if not project:
        return jsonify({'success': False, 'materials': []})

    # Consulta usando 'department' en lugar de 'area'
    items = db.session.query(
        Material.id,
        Material.code,
        Material.name,
        Material.unit,
        Material.current_stock,
        func.sum(RequestItem.quantity_requested).label('total_requested'),
        func.sum(RequestItem.quantity_delivered).label('total_delivered')
    ).join(RequestItem, Material.id == RequestItem.material_id)\
     .join(Request, RequestItem.request_id == Request.id)\
     .filter(
         Request.project_id == project.id,
         Request.department == department  # ✅ Cambiado de area a department
     )\
     .group_by(Material.id, Material.code, Material.name, Material.unit, Material.current_stock)\
     .all()

    materials_data = []
    for item in items:
        pending_quantity = item.total_requested - (item.total_delivered or 0)
        if pending_quantity > 0:
            materials_data.append({
                'id': item.id,
                'code': item.code,
                'name': item.name,
                'unit': item.unit,
                'stock': item.current_stock,
                'pending': pending_quantity
            })

    return jsonify({'success': True, 'materials': materials_data})



@app.route('/api/stock/exit', methods=['POST'])
@login_required
@permission_required('register_movements')
def register_exit():
    if current_user.role == 'requisitador':
        return jsonify({'success': False, 'message': 'No tienes permisos para registrar salidas.'}), 403

    """
    Registra una salida de material.
    Ahora usa 'department' en lugar de 'area'
    """
    try:
        data = request.form
        material_id = data.get('material_id')
        quantity = float(data.get('quantity'))
        fp_code = data.get('fp_code')
        department = data.get('department')

        print(f"🔍 Registro de salida:")
        print(f"   Material ID: {material_id}")
        print(f"   Cantidad: {quantity}")
        print(f"   FP: {fp_code}")
        print(f"   Departamento: {department}")

        material = Material.query.get(material_id)
        if not material:
            return jsonify({'success': False, 'message': 'Material no encontrado.'})

        if quantity > material.current_stock:
            return jsonify({
                'success': False,
                'message': f'Stock insuficiente. Disponible: {material.current_stock} {material.unit}'
            })

        # Generar IDM único para la salida
        last_id = StockMovement.query.count()
        idm = f"SAL-{datetime.utcnow().strftime('%y%m%d')}-{last_id + 1:04d}"

        movement = StockMovement(
            idm=idm,
            material_id=material.id,
            movement_type='salida',
            quantity=quantity,
            fp_code=fp_code,
            area=department,  # ✅ Guardar el department en el campo area
            fecha=datetime.utcnow().date(),
            hora=datetime.utcnow().time(),
            personal=current_user.username,
            user_id=current_user.id,
            notes=f"Solicitante: {data.get('requester', 'N/A')}. {data.get('notes', '')}",
            reference_type='requisicion'
        )

        db.session.add(movement)

        # Actualizar stock del material
        material.current_stock -= quantity
        material.last_movement = datetime.utcnow()

        # Actualizar quantity_delivered en RequestItem
        project = Project.query.filter_by(fp_code=fp_code).first()
        if project:
            req_item = db.session.query(RequestItem)\
                .join(Request)\
                .filter(
                    Request.project_id == project.id,
                    Request.department == department,
                    RequestItem.material_id == material.id
                )\
                .first()

            if req_item:
                req_item.quantity_delivered = (req_item.quantity_delivered or 0) + quantity
                print(f"✅ Actualizado RequestItem - Cantidad entregada: {req_item.quantity_delivered}")

        db.session.commit()
        print(f"✅ Salida registrada exitosamente: {idm}")

        return jsonify({
            'success': True,
            'message': 'Salida registrada exitosamente.',
            'idm': idm
        })

    except Exception as e:
        db.session.rollback()
        print(f"❌ Error en register_exit: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
@app.route('/api/stock/pending-return-by-fp')
@login_required
def get_pending_return_by_fp():
    fp_code = request.args.get('fp_code')
    if not fp_code:
        return jsonify({'materials': []})

    # Cantidades de salida para el proyecto
    salidas = db.session.query(
        StockMovement.material_id,
        func.sum(StockMovement.quantity).label('total_out')
    ).filter_by(fp_code=fp_code, movement_type='salida').group_by(StockMovement.material_id).subquery()

    # Cantidades de retorno para el proyecto
    retornos = db.session.query(
        StockMovement.material_id,
        func.sum(StockMovement.quantity).label('total_in')
    ).filter_by(fp_code=fp_code, movement_type='retorno').group_by(StockMovement.material_id).subquery()

    # Unir resultados para calcular pendientes
    pending_materials = db.session.query(
        Material.id, Material.code, Material.name, Material.unit,
        salidas.c.total_out,
        retornos.c.total_in
    ).join(salidas, Material.id == salidas.c.material_id)\
     .outerjoin(retornos, Material.id == retornos.c.material_id)\
     .all()

    result = []
    for mat in pending_materials:
        pending_qty = mat.total_out - (mat.total_in or 0)
        if pending_qty > 0.001: # Usar una pequeña tolerancia
            result.append({
                'id': mat.id,
                'code': mat.code,
                'name': mat.name,
                'unit': mat.unit,
                'pending_quantity': pending_qty
            })

    return jsonify({'materials': result})

# Agregar estas rutas a tu archivo principal de rutas (app.py o routes.py)

@app.route('/api/employees/active')
@login_required
def get_active_employees():
    """Obtener lista de empleados activos para el dropdown"""
    try:
        # ✅ Usar la clase RemoteDatabase en lugar de db.session
        employees = remote_db.get_empleados_activos()

        if not employees:
            return jsonify({
                'success': False,
                'message': 'No se encontraron empleados activos',
                'employees': []
            })

        # Formatear respuesta
        employees_data = []
        for emp in employees:
            if emp.get('id') and emp.get('nombre'):
                employees_data.append({
                    'id': emp['id'],
                    'nombre': emp['nombre']
                })

        return jsonify({
            'success': True,
            'employees': employees_data
        })

    except Exception as e:
        app.logger.error(f"Error obteniendo empleados activos: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al obtener empleados: {str(e)}',
            'employees': []
        })


@app.route('/api/user/fullname')
@login_required
def get_current_user_fullname():
    """Obtener nombre completo del usuario actual desde la base de datos remota"""
    try:
        # ✅ Buscar el empleado en la base remota usando el ID del usuario actual
        # Asumiendo que current_user.id corresponde al ID en empleados_activos
        empleado = remote_db.get_empleado_by_id(current_user.id)

        if empleado and empleado.get('nombre'):
            fullname = empleado['nombre']
        else:
            # Fallback: si no se encuentra, usar el username
            app.logger.warning(f"No se encontró empleado con ID {current_user.id}, usando username")
            fullname = current_user.username

        return jsonify({
            'success': True,
            'fullname': fullname
        })

    except Exception as e:
        app.logger.error(f"Error obteniendo nombre completo: {str(e)}")
        # En caso de error, devolver el username como fallback
        return jsonify({
            'success': True,
            'fullname': current_user.username
        })

@app.route('/api/stock/exit-multiple', methods=['POST'])
@login_required
@permission_required('register_movements')
def register_exit_multiple():
    if current_user.role == 'requisitador':
        return jsonify({'success': False, 'message': 'No tienes permisos para registrar salidas.'}), 403

    """
    Registra múltiples salidas de material en una sola operación.
    Todos los materiales comparten la misma información de proyecto, departamento y notas.
    """
    try:
        data = request.get_json()

        fp_code = data.get('fp_code')
        department = data.get('department')
        requester_id = data.get('requester_id')
        requester_name = data.get('requester_name')
        deliverer_name = data.get('deliverer_name')
        notes = data.get('notes', '')
        materials = data.get('materials', [])  # Lista de {material_id, quantity}

        if not materials or len(materials) == 0:
            return jsonify({
                'success': False,
                'message': 'Debe agregar al menos un material'
            })

        # Validar que el proyecto existe
        project = Project.query.filter_by(fp_code=fp_code).first()
        if not project:
            return jsonify({
                'success': False,
                'message': 'Proyecto no encontrado'
            })

        movements_created = []
        errors = []

        # Procesar cada material
        for item in materials:
            try:
                material_id = item.get('material_id')
                quantity = float(item.get('quantity'))

                material = Material.query.get(material_id)
                if not material:
                    errors.append(f"Material ID {material_id} no encontrado")
                    continue

                # Validar stock suficiente
                if quantity > material.current_stock:
                    errors.append(
                        f"{material.code}: Stock insuficiente. "
                        f"Disponible: {material.current_stock} {material.unit}"
                    )
                    continue

                # Generar IDM único
                last_id = StockMovement.query.count()
                idm = f"SAL-{datetime.utcnow().strftime('%y%m%d')}-{last_id + 1:04d}"

                # Crear movimiento
                movement = StockMovement(
                    idm=idm,
                    material_id=material.id,
                    movement_type='salida',
                    quantity=quantity,
                    fp_code=fp_code,
                    area=department,
                    fecha=datetime.utcnow().date(),
                    hora=datetime.utcnow().time(),
                    personal=deliverer_name,
                    user_id=current_user.id,
                    notes=f"Solicitante: {requester_name}. {notes}",
                    reference_type='requisicion'
                )

                db.session.add(movement)

                # Actualizar stock
                material.current_stock -= quantity
                material.last_movement = datetime.utcnow()

                # Actualizar quantity_delivered en RequestItem
                req_item = db.session.query(RequestItem)\
                    .join(Request)\
                    .filter(
                        Request.project_id == project.id,
                        Request.department == department,
                        RequestItem.material_id == material.id
                    )\
                    .first()

                if req_item:
                    req_item.quantity_delivered = (req_item.quantity_delivered or 0) + quantity

                movements_created.append({
                    'idm': idm,
                    'material': material.code,
                    'quantity': quantity
                })

            except Exception as e:
                errors.append(f"Error procesando material {item.get('material_id')}: {str(e)}")
                continue

        # Commit si hay al menos un movimiento exitoso
        if movements_created:
            db.session.commit()

            message = f"Se registraron {len(movements_created)} salida(s) exitosamente."
            if errors:
                message += f" {len(errors)} error(es): {'; '.join(errors[:3])}"

            return jsonify({
                'success': True,
                'message': message,
                'movements': movements_created,
                'errors': errors
            })
        else:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': 'No se pudo registrar ninguna salida',
                'errors': errors
            })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error en register_exit_multiple: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@app.route('/api/stock/material-pending-quantity')
@login_required
def get_material_pending_quantity():
    """
    Obtiene la cantidad pendiente de entregar para un material específico
    en un proyecto y departamento
    """
    try:
        material_id = request.args.get('material_id')
        fp_code = request.args.get('fp_code')
        department = request.args.get('department')

        if not all([material_id, fp_code, department]):
            return jsonify({
                'success': False,
                'message': 'Parámetros incompletos'
            })

        project = Project.query.filter_by(fp_code=fp_code).first()
        if not project:
            return jsonify({
                'success': False,
                'message': 'Proyecto no encontrado'
            })

        # Obtener cantidad requisitada y entregada
        req_item = db.session.query(RequestItem)\
            .join(Request)\
            .filter(
                Request.project_id == project.id,
                Request.department == department,
                RequestItem.material_id == material_id
            )\
            .first()

        if not req_item:
            return jsonify({
                'success': False,
                'message': 'No se encontró requisición para este material'
            })

        pending = req_item.quantity_requested - (req_item.quantity_delivered or 0)

        # Obtener info del material para la unidad
        material = Material.query.get(material_id)

        return jsonify({
            'success': True,
            'pending': pending,
            'unit': material.unit if material else '',
            'unit_symbol': material.unit if material else '',
            'requested': req_item.quantity_requested,
            'delivered': req_item.quantity_delivered or 0
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        })


# No olvides importar text de sqlalchemy

# Ruta para registrar retorno rápido
@app.route('/api/stock/quick-return', methods=['POST'])
@login_required
def register_quick_return():
    try:
        movement_id = request.form.get('movement_id')
        return_quantity = float(request.form.get('return_quantity'))
        condition = request.form.get('condition_on_return')
        notes = request.form.get('notes')

        # Obtener movimiento original
        original_movement = StockMovement.query.get(movement_id)
        if not original_movement:
            return jsonify({'success': False, 'message': 'Movimiento no encontrado'})

        if original_movement.returned:
            return jsonify({'success': False, 'message': 'Este movimiento ya fue devuelto'})

        if return_quantity > original_movement.quantity:
            return jsonify({'success': False, 'message': 'Cantidad de retorno excede la cantidad original'})

        # Actualizar movimiento original
        original_movement.returned = True
        original_movement.return_date = datetime.utcnow()
        original_movement.return_quantity = return_quantity
        original_movement.condition_on_return = condition
        original_movement.updated_at = datetime.utcnow()

        # Crear movimiento de retorno
        return_movement = StockMovement(
            material_id=original_movement.material_id,
            movement_type='retorno',
            quantity=return_quantity,
            reference_type='devolucion',
            reference_id=movement_id,
            user_id=current_user.id,
            notes=f"Retorno - Condición: {condition}. {notes if notes else ''}",
            rollos=original_movement.rollos,
            fp_code=original_movement.fp_code,
            area=original_movement.area,
            personal=original_movement.personal,
            fecha=datetime.now().date(),
            hora=datetime.now().time()
        )

        db.session.add(return_movement)

        # Actualizar stock según condición
        if condition in ['bueno', 'reutilizable']:
            # Devolver al stock normal
            original_movement.material.current_stock += return_quantity
        elif condition == 'reciclable':
            # Devolver al stock pero marcar como reciclable
            original_movement.material.current_stock += return_quantity
        # Si es 'desecho', no se devuelve al stock

        original_movement.material.last_movement = datetime.utcnow()

        db.session.commit()

        return jsonify({'success': True, 'message': 'Retorno registrado exitosamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al registrar retorno: {str(e)}'})

# Ruta para registrar retorno completo
@app.route('/api/stock/return', methods=['POST'])
@login_required
def register_return():
    if current_user.role == 'requisitador':
        return jsonify({'success': False, 'message': 'No tienes permisos para registrar retornos.'}), 403

    try:
        movement_id = request.form.get('movement_id')
        return_quantity = float(request.form.get('return_quantity'))
        condition = request.form.get('condition_on_return')
        notes = request.form.get('notes')

        # Lógica similar a quick_return pero más completa
        original_movement = StockMovement.query.get(movement_id)
        if not original_movement:
            return jsonify({'success': False, 'message': 'Movimiento no encontrado'})

        # Validaciones
        if original_movement.movement_type != 'salida':
            return jsonify({'success': False, 'message': 'Solo se pueden devolver salidas'})

        if original_movement.returned:
            return jsonify({'success': False, 'message': 'Este movimiento ya fue devuelto'})

        if return_quantity > original_movement.quantity:
            return jsonify({'success': False, 'message': 'Cantidad de retorno excede la cantidad original'})

        # Crear movimiento de retorno
        return_movement = StockMovement(
            idm=f"RET-{original_movement.idm}" if original_movement.idm else None,
            material_id=original_movement.material_id,
            movement_type='retorno',
            quantity=return_quantity,
            rollos=original_movement.rollos,
            fp_code=original_movement.fp_code,
            fecha=datetime.now().date(),
            hora=datetime.now().time(),
            personal=original_movement.personal,
            area=original_movement.area,
            reference_type='devolucion',
            reference_id=movement_id,
            user_id=current_user.id,
            notes=notes
        )

        db.session.add(return_movement)

        # Actualizar movimiento original
        original_movement.returned = True
        original_movement.return_date = datetime.utcnow()
        original_movement.return_quantity = return_quantity
        original_movement.condition_on_return = condition
        original_movement.updated_at = datetime.utcnow()

        # Actualizar stock del material
        material = original_movement.material
        if condition in ['bueno', 'reutilizable', 'reciclable']:
            material.current_stock += return_quantity

        material.last_movement = datetime.utcnow()

        db.session.commit()

        return jsonify({'success': True, 'message': 'Retorno registrado exitosamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al registrar retorno: {str(e)}'})




# Ruta para eliminar movimiento
@app.route('/api/stock/movements/<int:movement_id>', methods=['DELETE'])
@login_required
def delete_movement(movement_id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Permisos insuficientes'})

    try:
        movement = StockMovement.query.get(movement_id)
        if not movement:
            return jsonify({'success': False, 'message': 'Movimiento no encontrado'})

        # Revertir cambios en el stock
        material = movement.material
        if movement.movement_type == 'entrada':
            material.current_stock -= movement.quantity
        elif movement.movement_type == 'salida':
            material.current_stock += movement.quantity
        elif movement.movement_type == 'retorno':
            material.current_stock -= movement.quantity

        # Eliminar movimiento
        db.session.delete(movement)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Movimiento eliminado exitosamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar movimiento: {str(e)}'})

# Middleware para limpiar archivos temporales
@app.after_request
def cleanup_temp_files(response):
    """Limpia archivos temporales después de cada request"""
    try:
        upload_folder = app.config['UPLOAD_FOLDER']
        if os.path.exists(upload_folder):
            now = datetime.now()
            for filename in os.listdir(upload_folder):
                filepath = os.path.join(upload_folder, filename)
                if os.path.isfile(filepath):
                    file_time = datetime.fromtimestamp(os.path.getctime(filepath))
                    if (now - file_time).total_seconds() > 3600:  # 1 hora
                        try:
                            os.remove(filepath)
                        except:
                            pass
    except:
        pass

    return response

def init_db():
    """Inicializar la base de datos con datos de ejemplo"""
    try:
        # Verificar si la base de datos necesita migración
        if not safe_init_db():
            return False

        print("✅ Base de datos inicializada correctamente")

        # Crear usuario administrador por defecto
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                email='admin@empresa.com',
                password_hash=generate_password_hash('admin123'),
                role='admin',
                is_leader=True,
                department='Administración'
            )
            db.session.add(admin)

        # Crear usuarios adicionales de ejemplo
        if not User.query.filter_by(username='operador').first():
            operador = User(
                username='operador',
                email='operador@empresa.com',
                password_hash=generate_password_hash('operador123'),
                role='usuario',
                department='Producción'
            )
            db.session.add(operador)

            lider = User(
                username='lider',
                email='lider@empresa.com',
                password_hash=generate_password_hash('lider123'),
                role='usuario',
                is_leader=True,
                department='Producción'
            )
            db.session.add(lider)

        # Sembrar algunos códigos por rol si no existen
        if not VerificationCode.query.first():
            seeds = [
                ('REQ-AB12CD', 'requisitador'),
                ('ALM-34EF56', 'almacenista'),
                ('ADM-7890AA', 'admin'),
            ]
            for code, role in seeds:
                db.session.add(VerificationCode(code=code, role=role, expires_at=datetime.utcnow()+timedelta(days=90)))
            db.session.commit()


        # Crear proyecto de ejemplo
        if not Project.query.first():
            projects = [
                Project(
                    fp_code='FP-2025-001',
                    name='Proyecto Alpha',
                    delivery_date=datetime(2025, 12, 31).date(),
                    production_start=datetime(2025, 7, 1).date(),
                    assembly_date=datetime(2025, 11, 15).date()
                ),
                Project(
                    fp_code='FP-2025-002',
                    name='Proyecto Beta',
                    delivery_date=datetime(2025, 10, 15).date(),
                    production_start=datetime(2025, 8, 1).date(),
                    assembly_date=datetime(2025, 9, 30).date()
                )
            ]
            for project in projects:
                db.session.add(project)

        # Crear materiales de ejemplo
        if not Material.query.first():
            materials = [
                Material(
                    code='MAT-001',
                    name='Acero Inoxidable 304',
                    unit='kg',
                    category='Metales',
                    description='Acero inoxidable grado 304 para estructuras',
                    min_stock=100,
                    max_stock=1000,
                    unit_cost=15.50,
                    current_stock=0
                ),
                Material(
                    code='MAT-002',
                    name='Tela Algodón 100%',
                    unit='metros',
                    category='Textiles',
                    description='Tela de algodón 100% para vestimenta',
                    min_stock=500,
                    max_stock=2000,
                    unit_cost=8.75,
                    is_fabric_roll=True,
                    current_stock=0
                ),
                Material(
                    code='MAT-003',
                    name='Tornillos M6x20',
                    unit='piezas',
                    category='Ferretería',
                    description='Tornillos de acero M6x20mm',
                    min_stock=1000,
                    max_stock=5000,
                    unit_cost=0.25,
                    can_reuse=True,
                    current_stock=0
                ),
                Material(
                    code='MAT-004',
                    name='Pintura Acrílica Blanca',
                    unit='litros',
                    category='Químicos',
                    description='Pintura acrílica base agua color blanco',
                    min_stock=50,
                    max_stock=200,
                    unit_cost=25.00,
                    current_stock=0
                ),
                Material(
                    code='MAT-005',
                    name='Lámina Aluminio 2mm',
                    unit='metros²',
                    category='Metales',
                    description='Lámina de aluminio calibre 2mm',
                    min_stock=20,
                    max_stock=100,
                    unit_cost=45.00,
                    current_stock=0
                )
            ]
            for material in materials:
                db.session.add(material)

        # Crear algunos movimientos de ejemplo si no existen
        movements_exist = True
        try:
            StockMovement.query.first()
        except:
            movements_exist = False

        if not movements_exist:
            # Buscar materiales y usuario admin
            admin_user = User.query.filter_by(username='admin').first()
            operador_user = User.query.filter_by(username='operador').first()

            if admin_user:
                materials = Material.query.all()

                # Crear entradas iniciales para todos los materiales
                for i, material in enumerate(materials):
                    entrada_ejemplo = StockMovement(
                        idm=f'IDM{str(i+1).zfill(3)}',
                        material_id=material.id,
                        movement_type='entrada',
                        quantity=material.max_stock * 0.6,  # 60% del stock máximo
                        rollos=i+1 if material.is_fabric_roll else 0,
                        fp_code='FP-2025-001',
                        fecha=datetime.now().date(),
                        hora=datetime.now().time(),
                        personal='María Rodríguez',
                        area='Almacén',
                        unit_cost=material.unit_cost,
                        user_id=admin_user.id,
                        notes=f'Entrada inicial de inventario - {material.name}',
                        reference_type='inventario_inicial'
                    )
                    db.session.add(entrada_ejemplo)

                    # Actualizar stock del material
                    material.current_stock = material.max_stock * 0.6
                    material.last_movement = datetime.utcnow()

                # Crear algunas salidas de ejemplo
                if operador_user and len(materials) >= 3:
                    salida1 = StockMovement(
                        idm='IDM101',
                        material_id=materials[0].id,  # Acero
                        movement_type='salida',
                        quantity=50,
                        rollos=0,
                        fp_code='FP-2025-001',
                        fecha=datetime.now().date(),
                        hora=datetime.now().time(),
                        personal='Carlos López',
                        area='Producción',
                        user_id=operador_user.id,
                        notes='Salida para línea de producción A',
                        reference_type='requisicion'
                    )
                    db.session.add(salida1)
                    materials[0].current_stock -= 50

                    salida2 = StockMovement(
                        idm='IDM102',
                        material_id=materials[1].id,  # Tela
                        movement_type='salida',
                        quantity=25,
                        rollos=1,
                        fp_code='FP-2025-002',
                        fecha=datetime.now().date(),
                        hora=datetime.now().time(),
                        personal='Ana Martínez',
                        area='Confección',
                        user_id=operador_user.id,
                        notes='Salida para confección de uniformes',
                        reference_type='requisicion'
                    )
                    db.session.add(salida2)
                    materials[1].current_stock -= 25

                    # Crear un retorno de ejemplo
                    retorno1 = StockMovement(
                        idm='RET-IDM101',
                        material_id=materials[0].id,  # Acero
                        movement_type='retorno',
                        quantity=10,
                        rollos=0,
                        fp_code='FP-2025-001',
                        fecha=datetime.now().date(),
                        hora=datetime.now().time(),
                        personal='Carlos López',
                        area='Producción',
                        user_id=operador_user.id,
                        notes='Retorno de material sobrante - Estado: Buen estado',
                        reference_type='devolucion',
                        reference_id=1  # ID del movimiento de salida
                    )
                    db.session.add(retorno1)
                    materials[0].current_stock += 10

        # Crear algunos rollos de tela de ejemplo
        if not FabricRoll.query.first():
            tela_material = Material.query.filter_by(is_fabric_roll=True).first()
            if tela_material:
                rolls = [
                    FabricRoll(
                        material_id=tela_material.id,
                        roll_number='ROL-001',
                        total_length=100.0,
                        remaining_length=85.0,
                        width=150.0,
                        status='disponible'
                    ),
                    FabricRoll(
                        material_id=tela_material.id,
                        roll_number='ROL-002',
                        total_length=100.0,
                        remaining_length=20.0,
                        width=150.0,
                        status='critico'
                    ),
                    FabricRoll(
                        material_id=tela_material.id,
                        roll_number='ROL-003',
                        total_length=100.0,
                        remaining_length=0.0,
                        width=150.0,
                        status='agotado'
                    )
                ]
                for roll in rolls:
                    db.session.add(roll)

        try:
            db.session.commit()
            init_consumibles_data()

            print("🎉 Base de datos inicializada exitosamente con datos de ejemplo")
            return True
        except Exception as e:
            db.session.rollback()
            print(f"❌ Error al inicializar base de datos: {e}")
            return False

    except Exception as e:
        print(f"❌ Error crítico en inicialización: {e}")
        return False


def init_remote_sync():
    """Sincronización inicial con la base de datos remota"""
    try:
        print("🔄 Iniciando sincronización con base de datos remota...")

        # Sincronizar materiales
        result = sync_materials_from_remote()
        if result['success']:
            print(f"✅ Materiales sincronizados: {result['synced']} nuevos, {result['updated']} actualizados")
        else:
            print(f"❌ Error en sincronización de materiales: {result['error']}")

        # Mostrar estadísticas
        total_local = Material.query.count()
        print(f"📊 Total de materiales locales: {total_local}")

        return True

    except Exception as e:
        print(f"❌ Error en sincronización inicial: {e}")
        return False

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username      = request.form.get('username', '').strip()
        email         = request.form.get('email', '').strip().lower()
        password      = request.form.get('password', '').strip()
        selected_role = request.form.get('selected_role', '').strip().lower()
        code_input    = request.form.get('code', '').strip()

        # Validaciones base
        if not username or not email or not password or not code_input:
            flash('Todos los campos son obligatorios.', 'danger')
            return redirect(url_for('register'))

        # Reconstruir y validar código fijo
        full_code = _normalize_full_code(selected_role, code_input)
        role_from_code = VALID_CODES.get(full_code)
        if not role_from_code:
            flash('Código de verificación inválido.', 'danger')
            return redirect(url_for('register'))

        # Prevenir duplicados
        if User.query.filter((User.username == username) | (User.email == email)).first():
            flash('El usuario o correo ya existe.', 'warning')
            return redirect(url_for('register'))

        # Crear usuario (el rol mandatorio es el que deriva del CÓDIGO)
        new_user = User(
            username=username,
            email=email,
            password_hash=generate_password_hash(password),
            role=role_from_code,
            department="General",
            is_leader=True if role_from_code == 'admin' else False
        )
        db.session.add(new_user)
        db.session.commit()

        # Feedback
        if selected_role and selected_role != role_from_code:
            # Si el usuario eligió un rol pero el código es de otro, avisamos (se prioriza el código)
            flash(
                f'Usuario "{username}" creado como {role_from_code}. '
                f'(El código no corresponde al rol seleccionado).',
                'warning'
            )
        else:
            flash(f'Usuario "{username}" creado exitosamente como {role_from_code}.', 'success')

        return redirect(url_for('login'))

    # GET: mostrar formulario
    return render_template('register.html')
VALID_CODES = {
    "REQ-AB12CD": "requisitador",
    "ALM-34EF56": "almacenista",
    "ADM-7890AA": "admin",
}

# Mapa para reconstruir el código desde el rol elegido en el modal
ROLE_PREFIX = {"requisitador": "REQ", "almacenista": "ALM", "admin": "ADM"}

def _normalize_full_code(selected_role: str, code_input: str) -> str:
    """
    Tus templates de registro envían:
      - selected_role = 'requisitador' | 'almacenista' | 'admin'
      - code = 'XXXXXX' (solo sufijo)
    Si el usuario pegó el código completo, lo respetamos.
    Si solo dio sufijo, lo completamos con el prefijo del rol.
    """
    c = (code_input or "").strip().upper()
    if "-" in c and len(c) >= 4:
        return c  # ya viene completo (p.e. REQ-AB12CD)
    prefix = ROLE_PREFIX.get((selected_role or "").strip().lower(), "").upper()
    return f"{prefix}-{c}" if prefix and c else c

@app.route('/admin/verification-codes', methods=['GET','POST'])
@login_required
@role_required('admin')
def manage_verification_codes():
    if request.method == 'POST':
        role = request.form.get('role')
        qty  = int(request.form.get('qty', 1))
        days = int(request.form.get('days', 30))

        if role not in ('requisitador','almacenista','admin'):
            return jsonify({'success': False, 'message': 'Rol inválido'})

        new_codes = []
        for _ in range(qty):
            # Prefijo por rol y token cortito
            prefix = {'requisitador':'REQ','almacenista':'ALM','admin':'ADM'}[role]
            token  = secrets.token_hex(3).upper()  # 6 hex = 3 bytes
            code   = f"{prefix}-{token}"
            vc = VerificationCode(
                code=code,
                role=role,
                expires_at=datetime.utcnow()+timedelta(days=days)
            )
            db.session.add(vc)
            new_codes.append(code)
        db.session.commit()
        return jsonify({'success': True, 'codes': new_codes})

    codes = VerificationCode.query.order_by(VerificationCode.created_at.desc()).limit(50).all()
    return render_template('manage_codes.html', codes=codes)


# Función para migrar datos existentes (si es necesario)
def migrate_existing_data():
    """Migra datos existentes para completar las nuevas columnas"""
    with app.app_context():
        try:
            movements = StockMovement.query.filter(
                (StockMovement.fecha.is_(None)) |
                (StockMovement.hora.is_(None))
            ).all()

            for movement in movements:
                if not movement.fecha:
                    movement.fecha = movement.created_at.date()
                if not movement.hora:
                    movement.hora = movement.created_at.time()
                if not movement.personal:
                    movement.personal = movement.user.username
                if not movement.updated_at:
                    movement.updated_at = movement.created_at

            db.session.commit()
            print(f"✅ Migrados {len(movements)} movimientos existentes")
            return True
        except Exception as e:
            db.session.rollback()
            print(f"❌ Error en migración: {e}")
            return False

@app.route('/api/consumibles', methods=['POST'])
@login_required
def create_consumible():
    """Crear un nuevo consumible"""
    try:
        # Verificar si el código ya existe
        existing = Material.query.filter_by(code=request.form['code']).first()
        if existing:
            return jsonify({'success': False, 'message': 'Ya existe un material con ese código'})

        # Validar datos requeridos
        required_fields = ['code', 'name', 'category', 'unit', 'min_stock', 'max_stock']
        for field in required_fields:
            if not request.form.get(field):
                return jsonify({'success': False, 'message': f'El campo {field} es requerido'})

        # Validar que min_stock < max_stock
        min_stock = float(request.form['min_stock'])
        max_stock = float(request.form['max_stock'])
        if min_stock >= max_stock:
            return jsonify({'success': False, 'message': 'El stock máximo debe ser mayor al mínimo'})

        # Crear nuevo consumible
        consumible = Material(
            code=request.form['code'],
            name=request.form['name'],
            description=request.form.get('description', ''),
            unit=request.form['unit'],
            category=request.form['category'],
            current_stock=float(request.form.get('current_stock', 0)),
            min_stock=min_stock,
            max_stock=max_stock,
            unit_cost=float(request.form.get('unit_cost', 0)),
            is_consumible=True,  # Marcar como consumible
            can_recycle=False,   # Los consumibles generalmente no se reciclan
            can_reuse=False      # Los consumibles generalmente no se reutilizan
        )

        db.session.add(consumible)

        # Si tiene stock inicial, crear movimiento de entrada
        initial_stock = float(request.form.get('current_stock', 0))
        if initial_stock > 0:
            movement = StockMovement(
                material_id=consumible.id,
                movement_type='entrada',
                quantity=initial_stock,
                unit_cost=consumible.unit_cost,
                reference_type='stock_inicial',
                user_id=current_user.id,
                notes='Stock inicial al crear consumible'
            )
            db.session.add(movement)
            consumible.last_movement = datetime.utcnow()

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Consumible creado exitosamente',
            'consumible_id': consumible.id
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al crear consumible: {str(e)}'})

@app.route('/api/consumibles/add-stock', methods=['POST'])
@login_required
def add_consumible_stock():
    """Agregar stock a un consumible"""
    try:
        material_id = request.form.get('material_id')
        quantity = float(request.form.get('quantity', 0))
        unit_cost = float(request.form.get('unit_cost', 0))
        reason = request.form.get('reason')
        notes = request.form.get('notes', '')

        if not material_id or quantity <= 0:
            return jsonify({'success': False, 'message': 'Datos inválidos'})

        # Buscar el consumible
        consumible = Material.query.filter_by(id=material_id, is_consumible=True).first()
        if not consumible:
            return jsonify({'success': False, 'message': 'Consumible no encontrado'})

        # Actualizar stock
        consumible.current_stock += quantity
        consumible.last_movement = datetime.utcnow()

        # Actualizar costo unitario si se proporciona
        if unit_cost > 0:
            consumible.unit_cost = unit_cost

        # Crear movimiento de stock
        movement = StockMovement(
            material_id=material_id,
            movement_type='entrada',
            quantity=quantity,
            unit_cost=unit_cost if unit_cost > 0 else consumible.unit_cost,
            reference_type=reason,
            user_id=current_user.id,
            notes=f'Agregación de stock - {reason}: {notes}'
        )

        db.session.add(movement)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Stock agregado exitosamente. Nuevo stock: {consumible.current_stock} {consumible.unit}',
            'new_stock': consumible.current_stock
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al agregar stock: {str(e)}'})

@app.route('/api/consumibles/consume', methods=['POST'])
@login_required
def consume_stock():
    """Consumir stock de un consumible"""
    try:
        material_id = request.form.get('material_id')
        quantity = float(request.form.get('quantity', 0))
        department = request.form.get('department')
        reason = request.form.get('reason')
        notes = request.form.get('notes', '')

        if not material_id or quantity <= 0:
            return jsonify({'success': False, 'message': 'Datos inválidos'})

        # Buscar el consumible
        consumible = Material.query.filter_by(id=material_id, is_consumible=True).first()
        if not consumible:
            return jsonify({'success': False, 'message': 'Consumible no encontrado'})

        # Verificar stock disponible
        if consumible.current_stock < quantity:
            return jsonify({
                'success': False,
                'message': f'Stock insuficiente. Disponible: {consumible.current_stock} {consumible.unit}'
            })

        # Actualizar stock
        consumible.current_stock -= quantity
        consumible.last_movement = datetime.utcnow()

        # Crear movimiento de stock
        movement = StockMovement(
            material_id=material_id,
            movement_type='salida',
            quantity=quantity,
            unit_cost=consumible.unit_cost,
            reference_type='consumo',
            user_id=current_user.id,
            notes=f'Consumo - {department} - {reason}: {notes}'
        )

        db.session.add(movement)

        # Verificar si necesita alerta de stock bajo
        alert_message = ""
        if consumible.current_stock <= consumible.min_stock:
            alert_message = f" ⚠️ ALERTA: Stock crítico ({consumible.current_stock} {consumible.unit})"

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Consumo registrado exitosamente. Stock restante: {consumible.current_stock} {consumible.unit}{alert_message}',
            'new_stock': consumible.current_stock,
            'is_critical': consumible.current_stock <= consumible.min_stock
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al registrar consumo: {str(e)}'})

@app.route('/api/consumibles/<int:consumible_id>/urgent-order', methods=['POST'])
@login_required
def create_urgent_order(consumible_id):
    """Crear orden de compra urgente para un consumible"""
    try:
        consumible = Material.query.filter_by(id=consumible_id, is_consumible=True).first()
        if not consumible:
            return jsonify({'success': False, 'message': 'Consumible no encontrado'})

        # Verificar si ya existe una solicitud pendiente
        existing_request = PurchaseRequest.query.filter_by(
            material_id=consumible_id,
            status='pendiente'
        ).first()

        if existing_request:
            return jsonify({
                'success': False,
                'message': 'Ya existe una solicitud de compra pendiente para este consumible'
            })

        # Calcular cantidad a solicitar (hasta el stock máximo)
        quantity_needed = consumible.max_stock - consumible.current_stock

        # Crear solicitud de compra urgente
        from utils import generate_purchase_request_number
        request_number = generate_purchase_request_number()

        purchase_request = PurchaseRequest(
            request_number=request_number,
            material_id=consumible_id,
            quantity=quantity_needed,
            requested_by=current_user.id,
            status='urgente',  # Estado especial para órdenes urgentes
            notes=f'ORDEN URGENTE - Stock crítico: {consumible.current_stock} {consumible.unit}. Consumible: {consumible.name}'
        )

        db.session.add(purchase_request)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Orden de compra urgente {request_number} creada',
            'order_number': request_number,
            'quantity': quantity_needed
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al crear orden urgente: {str(e)}'})

@app.route('/api/consumibles/bulk-purchase-order', methods=['POST'])
@login_required
def create_bulk_purchase_order():
    """Crear orden de compra masiva para consumibles críticos"""
    try:
        # Buscar todos los consumibles con stock crítico
        critical_consumibles = Material.query.filter(
            Material.is_consumible == True,
            Material.current_stock <= Material.min_stock
        ).all()

        if not critical_consumibles:
            return jsonify({
                'success': False,
                'message': 'No hay consumibles críticos que requieran pedido'
            })

        # Crear solicitudes de compra para cada consumible crítico
        from utils import generate_purchase_request_number
        orders_created = []
        total_items = 0
        total_cost = 0

        for consumible in critical_consumibles:
            # Verificar que no tenga ya una solicitud pendiente
            existing = PurchaseRequest.query.filter_by(
                material_id=consumible.id,
                status__in=['pendiente', 'urgente']
            ).first()

            if existing:
                continue

            quantity_needed = consumible.max_stock - consumible.current_stock
            item_cost = quantity_needed * (consumible.unit_cost or 0)

            request_number = generate_purchase_request_number()

            purchase_request = PurchaseRequest(
                request_number=request_number,
                material_id=consumible.id,
                quantity=quantity_needed,
                requested_by=current_user.id,
                status='pendiente',
                purchase_cost=item_cost,
                notes=f'Orden masiva - Stock: {consumible.current_stock}/{consumible.max_stock} {consumible.unit}'
            )

            db.session.add(purchase_request)
            orders_created.append({
                'order_number': request_number,
                'consumible': consumible.name,
                'quantity': quantity_needed,
                'cost': item_cost
            })

            total_items += 1
            total_cost += item_cost

        if not orders_created:
            return jsonify({
                'success': False,
                'message': 'Todos los consumibles críticos ya tienen solicitudes pendientes'
            })

        db.session.commit()

        # Generar número de orden maestra
        master_order = f"ORD-MASIVA-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        return jsonify({
            'success': True,
            'message': f'Orden de compra masiva creada con {total_items} elementos',
            'master_order': master_order,
            'items_count': total_items,
            'total_cost': total_cost,
            'orders': orders_created,
            'order_url': f'/api/consumibles/purchase-order/{master_order}/pdf'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al crear orden masiva: {str(e)}'})

@app.route('/api/consumibles/export')
@login_required
def export_consumibles():
    """Exportar lista de consumibles"""
    try:
        consumibles = Material.query.filter_by(is_consumible=True).all()

        # Crear datos para exportar
        export_data = []
        for consumible in consumibles:
            # Calcular estado
            if consumible.current_stock <= consumible.min_stock:
                estado = "Crítico"
            elif consumible.current_stock <= (consumible.min_stock * 1.5):
                estado = "Bajo"
            elif consumible.current_stock >= consumible.max_stock:
                estado = "Completo"
            else:
                estado = "Normal"

            # Calcular próximo pedido estimado
            if consumible.current_stock <= consumible.min_stock:
                proximo_pedido = "Inmediato"
            elif consumible.current_stock <= (consumible.min_stock * 2):
                proximo_pedido = "Esta semana"
            else:
                proximo_pedido = "Próximo mes"

            export_data.append({
                'Código': consumible.code,
                'Nombre': consumible.name,
                'Categoría': consumible.category,
                'Descripción': consumible.description or '',
                'Stock Actual': consumible.current_stock,
                'Stock Mínimo': consumible.min_stock,
                'Stock Máximo': consumible.max_stock,
                'Unidad': consumible.unit,
                'Estado': estado,
                'Próximo Pedido': proximo_pedido,
                'Costo Unitario': consumible.unit_cost or 0,
                'Valor Total Stock': (consumible.current_stock * (consumible.unit_cost or 0)),
                'Último Movimiento': consumible.last_movement.strftime('%d/%m/%Y') if consumible.last_movement else 'Nunca',
                'Cantidad a Pedir': max(0, consumible.max_stock - consumible.current_stock),
                'Costo Reposición': max(0, consumible.max_stock - consumible.current_stock) * (consumible.unit_cost or 0)
            })

        # En una implementación real, aquí generarías un archivo Excel/CSV
        # Por ahora retornamos los datos en JSON
        return jsonify({
            'success': True,
            'data': export_data,
            'filename': f'consumibles_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al exportar: {str(e)}'})

@app.route('/api/consumibles/statistics')
@login_required
def consumibles_statistics():
    """Obtener estadísticas de consumibles"""
    try:
        consumibles = Material.query.filter_by(is_consumible=True).all()

        total_consumibles = len(consumibles)
        critical_count = 0
        low_count = 0
        normal_count = 0
        complete_count = 0
        total_value = 0
        monthly_consumption_cost = 0

        category_stats = {}
        urgent_orders = []

        for consumible in consumibles:
            # Contar por estado
            if consumible.current_stock <= consumible.min_stock:
                critical_count += 1
                urgent_orders.append({
                    'id': consumible.id,
                    'code': consumible.code,
                    'name': consumible.name,
                    'current_stock': consumible.current_stock,
                    'min_stock': consumible.min_stock,
                    'quantity_needed': consumible.max_stock - consumible.current_stock,
                    'cost_needed': (consumible.max_stock - consumible.current_stock) * (consumible.unit_cost or 0)
                })
            elif consumible.current_stock <= (consumible.min_stock * 1.5):
                low_count += 1
            elif consumible.current_stock >= consumible.max_stock:
                complete_count += 1
            else:
                normal_count += 1

            # Calcular valores
            total_value += consumible.current_stock * (consumible.unit_cost or 0)
            monthly_consumption_cost += (consumible.min_stock or 0) * (consumible.unit_cost or 0)

            # Estadísticas por categoría
            category = consumible.category
            if category not in category_stats:
                category_stats[category] = {
                    'count': 0,
                    'critical': 0,
                    'value': 0
                }

            category_stats[category]['count'] += 1
            category_stats[category]['value'] += consumible.current_stock * (consumible.unit_cost or 0)

            if consumible.current_stock <= consumible.min_stock:
                category_stats[category]['critical'] += 1

        return jsonify({
            'success': True,
            'statistics': {
                'total_consumibles': total_consumibles,
                'critical_count': critical_count,
                'low_count': low_count,
                'normal_count': normal_count,
                'complete_count': complete_count,
                'total_value': round(total_value, 2),
                'monthly_consumption_cost': round(monthly_consumption_cost, 2),
                'category_stats': category_stats,
                'urgent_orders': urgent_orders[:10]  # Primeros 10
            }
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al obtener estadísticas: {str(e)}'})


def init_consumibles_data():
    """Función para agregar datos de ejemplo de consumibles"""

    # Verificar si ya existen consumibles
    if Material.query.filter_by(is_consumible=True).first():
        return

    # Consumibles de ejemplo
    sample_consumibles = [
        {
            'code': 'CON-ELE-001',
            'name': 'Pilas AA Alcalinas',
            'description': 'Pilas AA para equipos electrónicos',
            'unit': 'piezas',
            'category': 'Electrónicos',
            'current_stock': 25,
            'min_stock': 10,
            'max_stock': 100,
            'unit_cost': 1.50
        },
        {
            'code': 'CON-LIM-001',
            'name': 'Papel Higiénico Industrial',
            'description': 'Rollos de papel higiénico para baños',
            'unit': 'rollos',
            'category': 'Limpieza',
            'current_stock': 5,
            'min_stock': 15,
            'max_stock': 50,
            'unit_cost': 2.25
        },
        {
            'code': 'CON-OFI-001',
            'name': 'Bolígrafos Azules',
            'description': 'Bolígrafos de tinta azul para oficina',
            'unit': 'piezas',
            'category': 'Oficina',
            'current_stock': 8,
            'min_stock': 20,
            'max_stock': 100,
            'unit_cost': 0.75
        },
        {
            'code': 'CON-MAN-001',
            'name': 'Cinta Aislante Negra',
            'description': 'Cinta aislante para mantenimiento eléctrico',
            'unit': 'rollos',
            'category': 'Mantenimiento',
            'current_stock': 3,
            'min_stock': 10,
            'max_stock': 30,
            'unit_cost': 3.50
        },
        {
            'code': 'CON-SEG-001',
            'name': 'Mascarillas N95',
            'description': 'Mascarillas de protección respiratoria',
            'unit': 'piezas',
            'category': 'Seguridad',
            'current_stock': 12,
            'min_stock': 50,
            'max_stock': 200,
            'unit_cost': 2.80
        }
    ]

    for consumible_data in sample_consumibles:
        consumible = Material(
            code=consumible_data['code'],
            name=consumible_data['name'],
            description=consumible_data['description'],
            unit=consumible_data['unit'],
            category=consumible_data['category'],
            current_stock=consumible_data['current_stock'],
            min_stock=consumible_data['min_stock'],
            max_stock=consumible_data['max_stock'],
            unit_cost=consumible_data['unit_cost'],
            is_consumible=True,
            can_recycle=False,
            can_reuse=False
        )

        db.session.add(consumible)

    db.session.commit()
    print("Datos de ejemplo de consumibles agregados")


@app.route('/admin/departments', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def manage_departments():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        code = (request.form.get('code') or '').strip() or None
        if not name:
            flash('El nombre es requerido')
        else:
            exists = Department.query.filter(
                db.or_(Department.name==name, Department.code==code if code else False)
            ).first()
            if exists:
                flash('Ya existe un departamento con ese nombre/código')
            else:
                db.session.add(Department(name=name, code=code))
                db.session.commit()
                flash('Departamento creado')
        return redirect(url_for('manage_departments'))

    depts = Department.query.order_by(Department.name).all()
    return render_template('departments.html', depts=depts)

@app.route('/admin/departments/<int:dept_id>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def toggle_department(dept_id):
    d = Department.query.get_or_404(dept_id)
    d.is_active = not d.is_active
    db.session.commit()
    return redirect(url_for('manage_departments'))

# ======== UTILIDADES ROLLOS DE TELA ========

def _roll_status(total, remaining):
    """Devuelve el estado del rollo en función de sus longitudes."""
    try:
        total = float(total or 0)
        remaining = float(remaining or 0)
    except:
        return 'disponible'
    if remaining <= 0:
        return 'agotado'
    if total > 0:
        ratio = remaining / total
    else:
        ratio = 1
    if ratio <= 0.20:
        return 'critico'
    if remaining < total:
        return 'en_uso'
    return 'disponible'


def get_default_width_for_material(material_id):
    """Ancho por defecto para un material de tela:
       - ancho del último rollo de ese material (si existe)
       - si no, 150 cm por defecto
    """
    last_roll = FabricRoll.query.filter_by(material_id=material_id)\
                                .order_by(FabricRoll.id.desc()).first()
    if last_roll and last_roll.width:
        return float(last_roll.width)
    return 150.0


# ======== API: DEFAULTS (ancho, número sugerido) ========

@app.route('/api/fabric-rolls/defaults')
@login_required
def api_fabric_roll_defaults():
    material_id = request.args.get('material_id', type=int)
    if not material_id:
        return jsonify({'success': False, 'message': 'material_id requerido'}), 400

    material = Material.query.get(material_id)
    if not material:
        return jsonify({'success': False, 'message': 'Material no encontrado'}), 404
    if not material.is_fabric_roll:
        return jsonify({'success': False, 'message': 'El material no es de tipo rollo de tela'}), 400

    default_width = get_default_width_for_material(material_id)

    # Sugerencia de número de rollo tipo ROL-YYYYMMDD-### (secuencial del día)
    today_prefix = f"ROL-{datetime.utcnow().strftime('%Y%m%d')}-"
    seq = FabricRoll.query.filter(FabricRoll.roll_number.like(f"{today_prefix}%")).count() + 1
    suggested_roll_number = f"{today_prefix}{seq:03d}"

    return jsonify({
        'success': True,
        'default_width': default_width,
        'suggested_roll_number': suggested_roll_number,
        'unit': material.unit
    })


# ======== API: CREAR ROLLO ========

@app.route('/api/fabric-rolls', methods=['POST'])
@login_required
@permission_required('register_movements')
def api_create_fabric_roll():
    try:
        material_id   = request.form.get('material_id', type=int)
        roll_number   = (request.form.get('roll_number') or '').strip()
        total_length  = request.form.get('total_length', type=float)
        notes         = (request.form.get('notes') or '').strip()

        # Validaciones básicas
        if not material_id or not roll_number or not total_length or total_length <= 0:
            return jsonify({'success': False, 'message': 'Datos inválidos (material, número y longitud son obligatorios)'}), 400

        material = Material.query.get(material_id)
        if not material:
            return jsonify({'success': False, 'message': 'Material no encontrado'}), 404
        if not material.is_fabric_roll:
            return jsonify({'success': False, 'message': 'El material no es de tipo rollo de tela'}), 400

        # roll_number único
        if FabricRoll.query.filter_by(roll_number=roll_number).first():
            return jsonify({'success': False, 'message': 'El número de rollo ya existe'}), 400

        # Ancho forzado por el sistema (ignora lo que venga del form)
        width = get_default_width_for_material(material_id)

        # Crear rollo
        roll = FabricRoll(
            material_id=material_id,
            roll_number=roll_number,
            total_length=total_length,
            remaining_length=total_length,
            width=width,
            status=_roll_status(total_length, total_length),
            notes=notes
        )
        db.session.add(roll)

        # Actualizar stock del material (metros)
        material.current_stock = (material.current_stock or 0) + total_length
        material.last_movement = datetime.utcnow()

        # Registrar movimiento de ENTRADA por alta de rollo
        mv = StockMovement(
            material_id=material.id,
            movement_type='entrada',
            quantity=total_length,
            rollos=1,
            reference_type='rollo_alta',
            user_id=current_user.id,
            personal=current_user.username,
            area='Almacén',
            fecha=datetime.utcnow().date(),
            hora=datetime.utcnow().time(),
            notes=f'Alta de rollo {roll_number} (ancho {width} cm)'
        )
        db.session.add(mv)

        db.session.commit()
        return jsonify({'success': True, 'message': 'Rollo creado', 'roll_id': roll.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al crear rollo: {str(e)}'}), 500


# ======== API: CORTE DE ROLLO ========

@app.route('/api/fabric-rolls/cut', methods=['POST'])
@login_required
@permission_required('register_movements')
def api_cut_fabric_roll():
    try:
        roll_id     = request.form.get('roll_id', type=int)
        cut_length  = request.form.get('cut_length', type=float)
        reason      = (request.form.get('reason') or '').strip()
        req_number  = (request.form.get('requisition_number') or '').strip()
        notes       = (request.form.get('notes') or '').strip()

        if not roll_id or not cut_length or cut_length <= 0:
            return jsonify({'success': False, 'message': 'Datos inválidos'}), 400

        roll = FabricRoll.query.get(roll_id)
        if not roll:
            return jsonify({'success': False, 'message': 'Rollo no encontrado'}), 404

        if cut_length > (roll.remaining_length or 0):
            return jsonify({'success': False, 'message': 'La longitud de corte excede la disponible'}), 400

        # Actualizar longitudes / estado
        roll.remaining_length = (roll.remaining_length or 0) - cut_length
        roll.status = _roll_status(roll.total_length, roll.remaining_length)

        # Actualizar stock del material
        material = roll.material
        material.current_stock = (material.current_stock or 0) - cut_length
        material.last_movement = datetime.utcnow()

        # Movimiento de SALIDA por corte de rollo
        mot = f'Corte rollo {roll.roll_number}'
        if reason:
            mot += f' · Motivo: {reason}'
        if req_number:
            mot += f' · Req: {req_number}'
        if notes:
            mot += f' · {notes}'

        mv = StockMovement(
            material_id=material.id,
            movement_type='salida',
            quantity=cut_length,
            rollos=0,
            reference_type='corte_rollo',
            user_id=current_user.id,
            personal=current_user.username,
            area='Producción',
            fecha=datetime.utcnow().date(),
            hora=datetime.utcnow().time(),
            notes=mot
        )
        db.session.add(mv)

        db.session.commit()
        return jsonify({'success': True, 'message': 'Corte registrado', 'remaining': roll.remaining_length, 'status': roll.status})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al cortar: {str(e)}'}), 500


# ======== API: ACTUALIZAR ROLLO ========

@app.route('/api/fabric-rolls/update', methods=['POST','PUT'])
@login_required
@permission_required('register_movements')
def api_update_fabric_roll():
    try:
        roll_id      = request.form.get('roll_id', type=int)
        if not roll_id:
            return jsonify({'success': False, 'message': 'roll_id requerido'}), 400

        roll = FabricRoll.query.get(roll_id)
        if not roll:
            return jsonify({'success': False, 'message': 'Rollo no encontrado'}), 404

        # Actualizables
        new_number = (request.form.get('roll_number') or '').strip()
        if new_number and new_number != roll.roll_number:
            if FabricRoll.query.filter(FabricRoll.roll_number == new_number, FabricRoll.id != roll.id).first():
                return jsonify({'success': False, 'message': 'El número de rollo ya existe'}), 400
            roll.roll_number = new_number

        # total_length lo dejamos bloqueado si ya hubo cortes
        if (request.form.get('total_length') is not None and
            float(request.form.get('total_length') or 0) >= (roll.remaining_length or 0) and
            (roll.remaining_length == roll.total_length)):  # sólo si no hubo cortes
            roll.total_length = float(request.form.get('total_length'))

        # El ancho puede ajustarse puntualmente (edición)
        if request.form.get('width') is not None:
            try:
                roll.width = float(request.form.get('width'))
            except:
                pass

        status = (request.form.get('status') or '').strip()
        if status in ('disponible','en_uso','critico','agotado'):
            roll.status = status
        notes = (request.form.get('notes') or '').strip()
        if notes is not None:
            roll.notes = notes

        db.session.commit()
        return jsonify({'success': True, 'message': 'Rollo actualizado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al actualizar rollo: {str(e)}'}), 500


# ======== API: ELIMINAR ROLLO ========

@app.route('/api/fabric-rolls/<int:roll_id>', methods=['DELETE'])
@login_required
@role_required('admin')
def api_delete_fabric_roll(roll_id):
    try:
        roll = FabricRoll.query.get(roll_id)
        if not roll:
            return jsonify({'success': False, 'message': 'Rollo no encontrado'}), 404

        # Ajustar stock del material (sale lo que quedaba)
        material = roll.material
        if roll.remaining_length and roll.remaining_length > 0:
            material.current_stock = (material.current_stock or 0) - roll.remaining_length
            material.last_movement = datetime.utcnow()
            # Registrar salida por ajuste
            mv = StockMovement(
                material_id=material.id,
                movement_type='salida',
                quantity=roll.remaining_length,
                rollos=0,
                reference_type='ajuste_eliminacion_rollo',
                user_id=current_user.id,
                personal=current_user.username,
                area='Almacén',
                fecha=datetime.utcnow().date(),
                hora=datetime.utcnow().time(),
                notes=f'Eliminación rollo {roll.roll_number}'
            )
            db.session.add(mv)

        db.session.delete(roll)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Rollo eliminado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar rollo: {str(e)}'}), 500



# --- ADMIN: Usuarios + Códigos fijos ---

from sqlalchemy import asc

@app.route('/admin/users-and-codes', methods=['GET'])
@login_required
@role_required('admin')
def admin_users_and_codes():
    # Lista de usuarios
    users = User.query.order_by(User.id.desc()).all()

    # Lista de códigos (fijos y cualesquiera extra que existan)
    codes = VerificationCode.query.order_by(asc(VerificationCode.role), asc(VerificationCode.code)).all()

    fixed_codes = [
        ("REQ-AB12CD", "requisitador"),
        ("ALM-34EF56", "almacenista"),
        ("ADM-7890AA", "admin"),
    ]

    return render_template('admin_users_codes.html',
                           users=users,
                           codes=codes,
                           fixed_codes=fixed_codes)


# ===== ADMINISTRACIÓN DE CATEGORÍAS =====
@app.route('/admin/categories')
@login_required
@role_required('admin')
def admin_categories():
    """Página de administración de categorías"""
    categories = Category.query.order_by(Category.name).all()
    return render_template('admin_categories.html', categories=categories)


@app.route('/admin/categories/add', methods=['POST'])
@login_required
@role_required('admin')
def admin_add_category():
    """Agregar nueva categoría"""
    try:
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        is_fabric = bool(request.form.get('is_fabric'))

        if not name:
            flash('El nombre es requerido', 'error')
            return redirect(url_for('admin_categories'))

        # Verificar duplicado
        existing = Category.query.filter_by(name=name).first()
        if existing:
            flash('Ya existe una categoría con ese nombre', 'error')
            return redirect(url_for('admin_categories'))

        category = Category(
            name=name,
            description=description if description else None,
            is_fabric=is_fabric,
            is_active=True,
            synced_at=datetime.utcnow()
        )

        db.session.add(category)
        db.session.commit()
        flash(f'Categoría "{name}" creada exitosamente', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear categoría: {str(e)}', 'error')

    return redirect(url_for('admin_categories'))


@app.route('/admin/categories/<int:category_id>/edit', methods=['POST'])
@login_required
@role_required('admin')
def admin_edit_category(category_id):
    """Editar categoría existente"""
    try:
        category = Category.query.get_or_404(category_id)

        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        is_fabric = bool(request.form.get('is_fabric'))
        is_active = bool(request.form.get('is_active'))

        if not name:
            flash('El nombre es requerido', 'error')
            return redirect(url_for('admin_categories'))

        # Verificar duplicado (exceptuando la categoría actual)
        existing = Category.query.filter(
            Category.name == name,
            Category.id != category_id
        ).first()
        if existing:
            flash('Ya existe otra categoría con ese nombre', 'error')
            return redirect(url_for('admin_categories'))

        category.name = name
        category.description = description if description else None
        category.is_fabric = is_fabric
        category.is_active = is_active
        category.synced_at = datetime.utcnow()

        db.session.commit()
        flash(f'Categoría "{name}" actualizada', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar: {str(e)}', 'error')

    return redirect(url_for('admin_categories'))


@app.route('/admin/categories/<int:category_id>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def admin_toggle_category(category_id):
    """Activar/Desactivar categoría"""
    try:
        category = Category.query.get_or_404(category_id)
        category.is_active = not category.is_active
        db.session.commit()

        status = "activada" if category.is_active else "desactivada"
        return jsonify({'success': True, 'message': f'Categoría {status}'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/admin/categories/<int:category_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def admin_delete_category(category_id):
    """Eliminar categoría (solo si no está sincronizada)"""
    try:
        category = Category.query.get_or_404(category_id)

        # No permitir eliminar categorías sincronizadas
        if category.remote_id:
            return jsonify({
                'success': False,
                'message': 'No se pueden eliminar categorías sincronizadas'
            })

        # Verificar si hay materiales usando esta categoría
        materials_count = Material.query.filter_by(category_id=category_id).count()
        if materials_count > 0:
            return jsonify({
                'success': False,
                'message': f'Esta categoría tiene {materials_count} materiales asociados'
            })

        name = category.name
        db.session.delete(category)
        db.session.commit()

        return jsonify({'success': True, 'message': f'Categoría "{name}" eliminada'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/admin/codes/reseed', methods=['POST'])
@login_required
@role_required('admin')
def admin_codes_reseed():
    try:
        seed_verification_codes_fixed()  # <- la que te compartí antes (códigos permanentes)
        return jsonify({'success': True, 'message': 'Códigos fijos verificados/insertados.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@app.route('/admin/users/<int:user_id>/set-role', methods=['POST'])
@login_required
@role_required('admin')
def admin_user_set_role(user_id):
    try:
        role = (request.form.get('role') or '').strip()
        if role not in ('requisitador', 'almacenista', 'admin'):
            return jsonify({'success': False, 'message': 'Rol inválido'}), 400

        u = User.query.get_or_404(user_id)
        u.role = role
        db.session.commit()
        return jsonify({'success': True, 'message': f'Rol actualizado a {role}'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@app.route('/admin/users/<int:user_id>/toggle-leader', methods=['POST'])
@login_required
@role_required('admin')
def admin_user_toggle_leader(user_id):
    try:
        u = User.query.get_or_404(user_id)
        u.is_leader = not bool(u.is_leader)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Estado de líder actualizado', 'is_leader': u.is_leader})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@role_required('admin')
def admin_user_reset_password(user_id):
    try:
        u = User.query.get_or_404(user_id)
        # contraseña temporal simple (ajústala si quieres más compleja)
        import secrets, string
        alphabet = string.ascii_letters + string.digits
        temp_pw = ''.join(secrets.choice(alphabet) for _ in range(10))
        u.password_hash = generate_password_hash(temp_pw)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Contraseña temporal generada', 'temp_password': temp_pw})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# ======== API: EXPORTAR ROLLOS ========

@app.route('/api/fabric-rolls/export')
@login_required
def api_export_fabric_rolls():
    try:
        rolls = FabricRoll.query.join(Material).all()
        data = []
        for r in rolls:
            unit_cost = r.material.unit_cost or 0
            data.append({
                'Rollo': r.roll_number,
                'Material': r.material.code,
                'Nombre': r.material.name,
                'Total (m)': r.total_length,
                'Disponible (m)': r.remaining_length,
                'Ancho (cm)': r.width,
                'Estado': r.status,
                'Valor ($)': (r.remaining_length or 0) * unit_cost,
                'Notas': r.notes or ''
            })
        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'rollos_tela_{timestamp}.xlsx'
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        df.to_excel(filepath, index=False, sheet_name='Rollos')
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al exportar: {str(e)}'}), 500


# ======== HISTORIAL: REDIRECCIÓN PRÁCTICA ========

@app.route('/fabric-rolls/<int:roll_id>/history')
@login_required
def roll_history_redirect(roll_id):
    roll = FabricRoll.query.get_or_404(roll_id)
    # Reusa tu visor de movimientos filtrando por código de material
    return redirect(url_for('stock_movements', material=roll.material.code))


# Comando para ejecutar desde la línea de comandos
# Comando para ejecutar desde la línea de comandos
if __name__ == '__main__':
    import sys

    # Asegura que existan tablas mínimas antes de cualquier comando
    with app.app_context():
        try:
            db.create_all()
        except Exception as e:
            print(f"⚠️ No se pudo crear tablas antes de iniciar: {e}")

    if len(sys.argv) > 1:
        cmd = sys.argv[1].strip().lower()

        if cmd == 'sync-remote':
            with app.app_context():
                init_remote_sync()
            sys.exit(0)

        elif cmd == 'migrate-data':
            with app.app_context():
                migrate_existing_data()
            sys.exit(0)

        elif cmd == 'seed-codes':
            # Solo sembrar/asegurar códigos de verificación
            with app.app_context():
                seed_verification_codes_if_missing()
            sys.exit(0)

        elif cmd == 'init-db':
            with app.app_context():
                if init_db():
                    print("✅ Inicialización completada exitosamente")
                    # Asegurar códigos base
                    seed_verification_codes_if_missing()
                    # Preguntar si desea sincronizar con la base remota
                    try:
                        response = input("¿Desea sincronizar materiales desde la base remota? (s/n): ")
                        if response.lower() == 's':
                            init_remote_sync()
                    except EOFError:
                        # en entornos sin stdin
                        pass
                else:
                    print("❌ Error en la inicialización de la base de datos")
            sys.exit(0)

        else:
            print(f"ℹ️ Comando no reconocido: {cmd}")
            print("   Usa: python app.py [init-db | seed-codes | migrate-data | sync-remote]")
            sys.exit(1)

    # Ejecución normal del servidor
    with app.app_context():
        # Verificación segura de DB y siembra de códigos de verificación
        safe_init_db()
        seed_verification_codes_fixed()

    # Iniciar scheduler de sincronización automática
    # use_reloader=False evita duplicar el scheduler en modo debug
    start_scheduler()

    app.run(debug=True, use_reloader=False)